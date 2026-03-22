"""Mandi Master V9 - FIXED: Compact Balance Sheet + Capital/Loan Ledger showing ALL parties"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_RED = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
HEADER_GREEN = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_PURPLE = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="0D47A1")
SUBTITLE_FONT = Font(bold=True, size=11, color="1565C0")
INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
CALC_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
MONEY_FMT = '₹#,##0'
NUM_FMT = '#,##0'
DATE_FMT = 'DD-MMM-YYYY'

def col_w(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def hdr(ws, row, c1, c2, fill=None):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def create_v9(path):
    wb = Workbook()
    wb.remove(wb.active)
    
    # Named ranges for INDIRECT dropdown - PARTY lists
    wb.defined_names.add(DefinedName("BEPAARI", attr_text="Masters!$B$3:$B$52"))
    wb.defined_names.add(DefinedName("DUKANDAR", attr_text="Masters!$I$3:$I$52"))
    wb.defined_names.add(DefinedName("CAPITAL", attr_text="Masters!$O$3:$O$32"))
    wb.defined_names.add(DefinedName("LOAN", attr_text="Masters!$O$3:$O$32"))
    wb.defined_names.add(DefinedName("ADVANCE", attr_text="Masters!$T$3:$T$22"))
    wb.defined_names.add(DefinedName("EXPENSE", attr_text="Lookups!$B$3:$B$3"))
    wb.defined_names.add(DefinedName("ZAKAT", attr_text="Lookups!$B$3:$B$3"))
    wb.defined_names.add(DefinedName("OTHER", attr_text="Lookups!$B$3:$B$3"))
    
    # Named ranges for Type-specific Sub-Types
    wb.defined_names.add(DefinedName("ST_BEPAARI", attr_text="Lookups!$D$3:$D$8"))
    wb.defined_names.add(DefinedName("ST_DUKANDAR", attr_text="Lookups!$E$3:$E$5"))
    wb.defined_names.add(DefinedName("ST_CAPITAL", attr_text="Lookups!$F$3:$F$4"))
    wb.defined_names.add(DefinedName("ST_LOAN", attr_text="Lookups!$G$3:$G$4"))
    wb.defined_names.add(DefinedName("ST_ADVANCE", attr_text="Lookups!$H$3:$H$4"))
    wb.defined_names.add(DefinedName("ST_EXPENSE", attr_text="Lookups!$I$3:$I$10"))
    wb.defined_names.add(DefinedName("ST_ZAKAT", attr_text="Lookups!$J$3:$J$4"))
    wb.defined_names.add(DefinedName("ST_OTHER", attr_text="Lookups!$K$3:$K$4"))
    
    # ========== LOOKUPS ==========
    ws = wb.create_sheet("Lookups", 0)
    ws.sheet_properties.tabColor = "9E9E9E"
    ws.cell(row=1, column=1, value="LOOKUP TABLES").font = TITLE_FONT
    ws.cell(row=2, column=2, value="NO_PARTY")
    ws.cell(row=3, column=2, value="-")
    
    for i, st in enumerate(["PAYMENT", "MOTOR", "BHUSSA", "GAWALI", "CASH_ADV", "OTHER"], 3):
        ws.cell(row=i, column=4, value=st)
    for i, st in enumerate(["RECEIPT", "DISCOUNT", "OTHER"], 3):
        ws.cell(row=i, column=5, value=st)
    for i, st in enumerate(["TAKEN", "WITHDRAWN"], 3):
        ws.cell(row=i, column=6, value=st)
    for i, st in enumerate(["TAKEN", "REPAID"], 3):
        ws.cell(row=i, column=7, value=st)
    for i, st in enumerate(["GIVEN", "RECEIVED"], 3):
        ws.cell(row=i, column=8, value=st)
    for i, st in enumerate(["MANDI", "TRAVEL", "FOOD", "SALARY", "BF_DISC", "JB_PAID", "MISC", "OTHER"], 3):
        ws.cell(row=i, column=9, value=st)
    for i, st in enumerate(["PROVISION", "PAID"], 3):
        ws.cell(row=i, column=10, value=st)
    for i, st in enumerate(["MISC", "OTHER"], 3):
        ws.cell(row=i, column=11, value=st)
    
    # ========== MASTERS ==========
    ws = wb.create_sheet("Masters", 1)
    ws.sheet_properties.tabColor = "1565C0"
    
    # BEPAARI MASTER
    ws.cell(row=1, column=1, value="BEPAARI MASTER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Bepaari Name", "Comm%", "Opening Bal", "Phone", "Remarks"], 1):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 1, 6)
    bepaaris = ["ABUL HASAN", "SHARAFAT", "SHARIK", "JUNAID", "ANNU SHARIF", "GAYYUR", "ARIF KK", "KISHOR", "SHOEB BYCULLA", "MAULANA", "ISLAM D", "SALIM RAJ", "MEHMOOD MADH", "ANAS RAJ", "MOBIN CHIPLUN", "AYAZ", "JALIL", "NEERAJ", "IRFAN JOGESHWARI", "CHINTU SHUJALPUR"]
    for i, name in enumerate(bepaaris, 1):
        ws.cell(row=i+2, column=1, value=i)
        ws.cell(row=i+2, column=2, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=3, value=4).fill = INPUT_FILL
        ws.cell(row=i+2, column=4, value=0).fill = INPUT_FILL
        ws.cell(row=i+2, column=4).number_format = MONEY_FMT
    for r in range(len(bepaaris)+3, 53):
        ws.cell(row=r, column=1, value=r-2)
        for c in range(2, 7): ws.cell(row=r, column=c).fill = INPUT_FILL
    
    # DUKANDAR MASTER
    ws.cell(row=1, column=8, value="DUKANDAR MASTER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Dukandar Name", "Opening Bal", "Phone", "Remarks"], 8):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 8, 12)
    dukandars = ["SAKIM", "SHARIK", "YUNUS", "MUDASSIR", "ARIF", "SHOEB BYCULLA", "ISLAM D", "MEHMOOD MADH", "MOBIN CHIPLUN", "JALIL", "IRFAN JOGESHWARI", "YUNUS DOMBIVLI", "IFTEKHAR PUNE", "LALA/ASIF", "JAFER KASHIMIRA", "CHAND", "ASIF GOREGAON", "CHINTU SHUJALPUR", "VAZIR", "AKRAM", "NADEEM BHIWANDI", "SHAKIL KON", "PARVEZ"]
    for i, name in enumerate(dukandars, 1):
        ws.cell(row=i+2, column=8, value=i)
        ws.cell(row=i+2, column=9, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=10, value=0).fill = INPUT_FILL
        ws.cell(row=i+2, column=10).number_format = MONEY_FMT
    for r in range(len(dukandars)+3, 53):
        ws.cell(row=r, column=8, value=r-2)
        for c in range(9, 13): ws.cell(row=r, column=c).fill = INPUT_FILL
    
    # CAPITAL/LOAN PARTIES - 30 rows max
    ws.cell(row=1, column=14, value="CAPITAL/LOAN PARTIES").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Party Name", "Opening Bal", "Type"], 14):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 14, 17)
    cap_parties = [["MHN", 3900000, "CAPITAL"], ["SHAKIL GHODEGAON", 2000000, "LOAN"]]
    for i, (name, bal, typ) in enumerate(cap_parties, 1):
        ws.cell(row=i+2, column=14, value=i)
        ws.cell(row=i+2, column=15, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=16, value=bal).fill = INPUT_FILL
        ws.cell(row=i+2, column=16).number_format = MONEY_FMT
        ws.cell(row=i+2, column=17, value=typ).fill = INPUT_FILL
    for r in range(len(cap_parties)+3, 33):
        ws.cell(row=r, column=14, value=r-2)
        for c in range(15, 18): ws.cell(row=r, column=c).fill = INPUT_FILL
    type_dv = DataValidation(type="list", formula1='"CAPITAL,LOAN,AMANAT"', allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add('Q3:Q32')
    
    # ADVANCE PARTIES
    ws.cell(row=1, column=19, value="ADVANCE PARTIES").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Party Name", "Opening Bal"], 19):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 19, 21)
    for r in range(3, 23):
        ws.cell(row=r, column=19, value=r-2)
        for c in range(20, 22): ws.cell(row=r, column=c).fill = INPUT_FILL
    
    # SETTINGS
    ws.cell(row=1, column=23, value="SETTINGS").font = TITLE_FONT
    settings = [["Commission Rate (%)", 4], ["KK Fixed (₹/Bepari)", 100], ["JB Rate (₹/Goat)", 10], ["", ""], ["Opening Cash", 0], ["Opening Bank", 0], ["", ""], ["ZAKAT Opening Payable", 0]]
    for r, (lbl, val) in enumerate(settings, 2):
        ws.cell(row=r, column=23, value=lbl).font = Font(bold=True) if lbl else None
        c = ws.cell(row=r, column=24, value=val)
        c.fill = INPUT_FILL
        if "Cash" in str(lbl) or "Bank" in str(lbl) or "ZAKAT" in str(lbl): c.number_format = MONEY_FMT
    
    for c, w in {1:4, 2:20, 3:8, 4:12, 5:10, 6:10, 7:2, 8:4, 9:20, 10:12, 11:10, 12:10, 13:2, 14:4, 15:20, 16:14, 17:10, 18:2, 19:4, 20:20, 21:12, 22:2, 23:22, 24:12}.items():
        col_w(ws, c, w)
    
    # ========== DAILY_SALES ==========
    ws = wb.create_sheet("Daily_Sales", 2)
    ws.sheet_properties.tabColor = "4CAF50"
    ws.cell(row=1, column=1, value="DAILY SALES (Bepaari to Dukandar)").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Date", "Bepaari", "Dukandar", "Qty", "Rate", "Gross", "Discount", "Net"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 9)
    bep_dv = DataValidation(type="list", formula1='BEPAARI', allow_blank=True)
    duk_dv = DataValidation(type="list", formula1='DUKANDAR', allow_blank=True)
    ws.add_data_validation(bep_dv)
    ws.add_data_validation(duk_dv)
    for r in range(4, 504):
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3).fill = INPUT_FILL
        bep_dv.add(f'C{r}')
        ws.cell(row=r, column=4).fill = INPUT_FILL
        duk_dv.add(f'D{r}')
        ws.cell(row=r, column=5).fill = INPUT_FILL
        ws.cell(row=r, column=6).fill = INPUT_FILL
        ws.cell(row=r, column=6).number_format = MONEY_FMT
        ws.cell(row=r, column=7, value=f'=IF(E{r}="","",E{r}*F{r})')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=7).fill = CALC_FILL
        ws.cell(row=r, column=8).fill = INPUT_FILL
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        ws.cell(row=r, column=9, value=f'=IF(G{r}="","",G{r}-IF(H{r}="",0,H{r}))')
        ws.cell(row=r, column=9).number_format = MONEY_FMT
        ws.cell(row=r, column=9).fill = CALC_FILL
    for c, w in enumerate([4, 11, 20, 20, 6, 10, 12, 10, 12], 1): col_w(ws, c, w)
    
    # ========== CASH_BOOK ==========
    ws = wb.create_sheet("Cash_Book", 3)
    ws.sheet_properties.tabColor = "FF5722"
    ws.cell(row=1, column=1, value="CASH BOOK (All Transactions)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Select TYPE first → Sub-Type & Party dropdowns update automatically").font = Font(italic=True, size=9)
    
    for c, h in enumerate(["Sr", "Date", "Type", "Sub-Type", "Party", "Particulars", "Amount", "Mode"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 8)
    
    main_dv = DataValidation(type="list", formula1='"BEPAARI,DUKANDAR,CAPITAL,LOAN,ADVANCE,EXPENSE,ZAKAT,OTHER"', allow_blank=True)
    mode_dv = DataValidation(type="list", formula1='"CASH,BANK,UPI,TRANSFER"', allow_blank=True)
    ws.add_data_validation(main_dv)
    ws.add_data_validation(mode_dv)
    
    for r in range(4, 1004):
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3).fill = INPUT_FILL
        main_dv.add(f'C{r}')
        ws.cell(row=r, column=4).fill = INPUT_FILL
        subtype_dv = DataValidation(type="list", formula1=f'INDIRECT("ST_"&$C{r})', allow_blank=True)
        ws.add_data_validation(subtype_dv)
        subtype_dv.add(f'D{r}')
        ws.cell(row=r, column=5).fill = INPUT_FILL
        party_dv = DataValidation(type="list", formula1=f'INDIRECT($C{r})', allow_blank=True)
        ws.add_data_validation(party_dv)
        party_dv.add(f'E{r}')
        ws.cell(row=r, column=6).fill = INPUT_FILL
        ws.cell(row=r, column=7).fill = INPUT_FILL
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=8).fill = INPUT_FILL
        mode_dv.add(f'H{r}')
    
    # Cash/Bank Summary
    ws.cell(row=1, column=10, value="CASH/BANK SUMMARY").font = SUBTITLE_FONT
    ws.cell(row=2, column=9, value="Opening Cash")
    ws.cell(row=2, column=10, value="=Masters!$X$6").number_format = MONEY_FMT
    ws.cell(row=3, column=9, value="Opening Bank")
    ws.cell(row=3, column=10, value="=Masters!$X$7").number_format = MONEY_FMT
    
    # Cash OUT
    ws.cell(row=5, column=9, value="Cash OUT")
    ws.cell(row=5, column=10, value='=SUMIFS(G:G,H:H,"CASH",D:D,"PAYMENT")+SUMIFS(G:G,H:H,"CASH",D:D,"MOTOR")+SUMIFS(G:G,H:H,"CASH",D:D,"BHUSSA")+SUMIFS(G:G,H:H,"CASH",D:D,"GAWALI")+SUMIFS(G:G,H:H,"CASH",D:D,"CASH_ADV")+SUMIFS(G:G,H:H,"CASH",D:D,"DISCOUNT")+SUMIFS(G:G,H:H,"CASH",D:D,"WITHDRAWN")+SUMIFS(G:G,H:H,"CASH",D:D,"REPAID")+SUMIFS(G:G,H:H,"CASH",D:D,"GIVEN")+SUMIFS(G:G,H:H,"CASH",D:D,"MANDI")+SUMIFS(G:G,H:H,"CASH",D:D,"TRAVEL")+SUMIFS(G:G,H:H,"CASH",D:D,"FOOD")+SUMIFS(G:G,H:H,"CASH",D:D,"SALARY")+SUMIFS(G:G,H:H,"CASH",D:D,"BF_DISC")+SUMIFS(G:G,H:H,"CASH",D:D,"JB_PAID")+SUMIFS(G:G,H:H,"CASH",D:D,"PAID")+SUMIFS(G:G,H:H,"CASH",D:D,"MISC")+SUMIFS(G:G,H:H,"CASH",D:D,"OTHER")')
    ws.cell(row=5, column=10).number_format = MONEY_FMT
    
    # Bank OUT
    ws.cell(row=6, column=9, value="Bank OUT")
    bank_out = []
    for mode in ["BANK", "UPI", "TRANSFER"]:
        for st in ["PAYMENT", "MOTOR", "BHUSSA", "GAWALI", "CASH_ADV", "DISCOUNT", "WITHDRAWN", "REPAID", "GIVEN", "MANDI", "TRAVEL", "FOOD", "SALARY", "BF_DISC", "JB_PAID", "PAID", "MISC", "OTHER"]:
            bank_out.append(f'SUMIFS(G:G,H:H,"{mode}",D:D,"{st}")')
    ws.cell(row=6, column=10, value='=' + '+'.join(bank_out)).number_format = MONEY_FMT
    
    # Cash IN
    ws.cell(row=7, column=9, value="Cash IN")
    ws.cell(row=7, column=10, value='=SUMIFS(G:G,H:H,"CASH",D:D,"RECEIPT")+SUMIFS(G:G,H:H,"CASH",D:D,"TAKEN")+SUMIFS(G:G,H:H,"CASH",D:D,"RECEIVED")+SUMIFS(G:G,H:H,"CASH",D:D,"PROVISION")').number_format = MONEY_FMT
    
    # Bank IN
    ws.cell(row=8, column=9, value="Bank IN")
    bank_in = []
    for mode in ["BANK", "UPI", "TRANSFER"]:
        for st in ["RECEIPT", "TAKEN", "RECEIVED", "PROVISION"]:
            bank_in.append(f'SUMIFS(G:G,H:H,"{mode}",D:D,"{st}")')
    ws.cell(row=8, column=10, value='=' + '+'.join(bank_in)).number_format = MONEY_FMT
    
    ws.cell(row=10, column=9, value="CLOSING CASH").font = Font(bold=True)
    ws.cell(row=10, column=10, value='=J2-J5+J7').number_format = MONEY_FMT
    ws.cell(row=10, column=10).font = Font(bold=True)
    ws.cell(row=10, column=10).fill = CALC_FILL
    
    ws.cell(row=11, column=9, value="CLOSING BANK").font = Font(bold=True)
    ws.cell(row=11, column=10, value='=J3-J6+J8').number_format = MONEY_FMT
    ws.cell(row=11, column=10).font = Font(bold=True)
    ws.cell(row=11, column=10).fill = CALC_FILL
    
    for c, w in enumerate([4, 11, 12, 14, 20, 20, 12, 10, 2, 14, 14], 1): col_w(ws, c, w)
    
    # ========== BEPAARI_LEDGER ==========
    ws = wb.create_sheet("Bepaari_Ledger", 4)
    ws.sheet_properties.tabColor = "9C27B0"
    ws.cell(row=1, column=1, value="BEPAARI LEDGER").font = TITLE_FONT
    hdrs = ["Sr", "Bepaari", "Comm%", "Opening", "Gross Sales", "Qty", "Commission", "KK", "JB", "Motor", "Bhussa", "Gawali", "Cash/Adv", "Other", "Tot.Deduct", "Net Payable", "Payments", "BALANCE"]
    for c, h in enumerate(hdrs, 1): ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 18)
    
    for r in range(4, 54):
        m = r - 1
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2, value=f'=IF(Masters!$B${m}="","",Masters!$B${m})')
        ws.cell(row=r, column=3, value=f'=IF(B{r}="","",Masters!$C${m})')
        ws.cell(row=r, column=4, value=f'=IF(B{r}="",0,Masters!$D${m})').number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f'=SUMIF(Daily_Sales!$C:$C,B{r},Daily_Sales!$G:$G)').number_format = MONEY_FMT
        ws.cell(row=r, column=6, value=f'=SUMIF(Daily_Sales!$C:$C,B{r},Daily_Sales!$E:$E)').number_format = NUM_FMT
        ws.cell(row=r, column=7, value=f'=IF(B{r}="",0,E{r}*C{r}/100)').number_format = MONEY_FMT
        ws.cell(row=r, column=8, value=f'=IF(F{r}>0,Masters!$X$3,0)').number_format = MONEY_FMT
        ws.cell(row=r, column=8).fill = CALC_FILL
        ws.cell(row=r, column=9, value=f'=F{r}*Masters!$X$4').number_format = MONEY_FMT
        ws.cell(row=r, column=9).fill = CALC_FILL
        for col, sub in [(10,"MOTOR"), (11,"BHUSSA"), (12,"GAWALI"), (13,"CASH_ADV"), (14,"OTHER")]:
            ws.cell(row=r, column=col, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"{sub}",Cash_Book!$E:$E,B{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=15, value=f'=SUM(G{r}:N{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=15).fill = CALC_FILL
        ws.cell(row=r, column=16, value=f'=IF(B{r}="","",E{r}-O{r}+D{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=16).fill = CALC_FILL
        ws.cell(row=r, column=17, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,B{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=18, value=f'=IF(B{r}="","",P{r}-Q{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=18).font = Font(bold=True)
    
    ws.cell(row=54, column=1, value="TOTAL").font = Font(bold=True)
    for c in [4,5,6,7,8,9,10,11,12,13,14,15,16,17,18]:
        ws.cell(row=54, column=c, value=f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}53)').number_format = MONEY_FMT
        ws.cell(row=54, column=c).font = Font(bold=True)
    
    # Balance Sheet helper
    ws.cell(row=56, column=15, value="For Balance Sheet:").font = Font(bold=True, italic=True)
    ws.cell(row=57, column=16, value="Bepaari Payable (we owe)")
    ws.cell(row=57, column=18, value='=SUMIF(R4:R53,">0",R4:R53)').number_format = MONEY_FMT
    ws.cell(row=57, column=18).fill = CALC_FILL
    ws.cell(row=58, column=16, value="Bepaari Overpaid (they owe)")
    ws.cell(row=58, column=18, value='=-SUMIF(R4:R53,"<0",R4:R53)').number_format = MONEY_FMT
    ws.cell(row=58, column=18).fill = CALC_FILL
    
    for c, w in enumerate([4,16,6,10,12,6,10,8,8,9,9,9,10,8,11,12,11,12], 1): col_w(ws, c, w)
    
    # ========== DUKANDAR_LEDGER ==========
    ws = wb.create_sheet("Dukandar_Ledger", 5)
    ws.sheet_properties.tabColor = "E91E63"
    ws.cell(row=1, column=1, value="DUKANDAR LEDGER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Dukandar", "Opening", "Purchases", "Discounts", "Net Receivable", "Receipts", "BALANCE"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 8)
    
    for r in range(4, 54):
        m = r - 1
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2, value=f'=IF(Masters!$I${m}="","",Masters!$I${m})')
        ws.cell(row=r, column=3, value=f'=IF(B{r}="",0,Masters!$J${m})').number_format = MONEY_FMT
        ws.cell(row=r, column=4, value=f'=SUMIF(Daily_Sales!$D:$D,B{r},Daily_Sales!$G:$G)').number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f'=SUMIF(Daily_Sales!$D:$D,B{r},Daily_Sales!$H:$H)').number_format = MONEY_FMT
        ws.cell(row=r, column=6, value=f'=IF(B{r}="","",D{r}-E{r}+C{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=6).fill = CALC_FILL
        ws.cell(row=r, column=7, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"DUKANDAR",Cash_Book!$D:$D,"RECEIPT",Cash_Book!$E:$E,B{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=8, value=f'=IF(B{r}="","",F{r}-G{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=8).font = Font(bold=True)
    
    ws.cell(row=54, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=54, column=2, value="PATTI")
    for c in [3,4,5,6,7,8]:
        ws.cell(row=54, column=c, value=f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}53)').number_format = MONEY_FMT
        ws.cell(row=54, column=c).font = Font(bold=True)
    
    # Balance Sheet helper
    ws.cell(row=56, column=5, value="For Balance Sheet:").font = Font(bold=True, italic=True)
    ws.cell(row=57, column=6, value="Dukandar Receivable (they owe)")
    ws.cell(row=57, column=8, value='=SUMIF(H4:H53,">0",H4:H53)').number_format = MONEY_FMT
    ws.cell(row=57, column=8).fill = CALC_FILL
    ws.cell(row=58, column=6, value="Dukandar Overpaid (we owe)")
    ws.cell(row=58, column=8, value='=-SUMIF(H4:H53,"<0",H4:H53)').number_format = MONEY_FMT
    ws.cell(row=58, column=8).fill = CALC_FILL
    
    for c, w in enumerate([4,20,12,14,12,14,12,14], 1): col_w(ws, c, w)
    
    # ========== CAPITAL_LOAN_LEDGER (NEW - Shows ALL parties) ==========
    ws = wb.create_sheet("Capital_Loan_Ledger", 6)
    ws.sheet_properties.tabColor = "6A1B9A"
    ws.cell(row=1, column=1, value="CAPITAL & LOAN LEDGER (All Investors/Lenders)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="This sheet shows ALL parties from Masters with their current balances").font = Font(italic=True, size=9)
    
    for c, h in enumerate(["Sr", "Party Name", "Type", "Opening Bal", "Taken (+)", "Withdrawn/Repaid (-)", "CURRENT BALANCE"], 1):
        ws.cell(row=4, column=c, value=h)
    hdr(ws, 4, 1, 7, HEADER_PURPLE)
    
    # Loop through all 30 possible Capital/Loan parties from Masters
    for r in range(5, 35):
        m = r - 2  # Masters row (3 to 32)
        ws.cell(row=r, column=1, value=r-4)
        ws.cell(row=r, column=2, value=f'=IF(Masters!$O${m}="","",Masters!$O${m})')
        ws.cell(row=r, column=3, value=f'=IF(B{r}="","",Masters!$Q${m})')
        ws.cell(row=r, column=4, value=f'=IF(B{r}="",0,Masters!$P${m})').number_format = MONEY_FMT
        # TAKEN = Capital TAKEN + Loan TAKEN
        ws.cell(row=r, column=5, value=f'=IF(B{r}="",0,SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,C{r},Cash_Book!$D:$D,"TAKEN",Cash_Book!$E:$E,B{r}))').number_format = MONEY_FMT
        # WITHDRAWN (for Capital) or REPAID (for Loan)
        ws.cell(row=r, column=6, value=f'=IF(B{r}="",0,SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,C{r},Cash_Book!$D:$D,"WITHDRAWN",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,C{r},Cash_Book!$D:$D,"REPAID",Cash_Book!$E:$E,B{r}))').number_format = MONEY_FMT
        # Current Balance = Opening + Taken - Withdrawn/Repaid
        ws.cell(row=r, column=7, value=f'=IF(B{r}="","",D{r}+E{r}-F{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=7).font = Font(bold=True)
        ws.cell(row=r, column=7).fill = CALC_FILL
    
    # TOTALS by Type
    ws.cell(row=36, column=1, value="TOTALS BY TYPE:").font = Font(bold=True)
    
    ws.cell(row=37, column=2, value="CAPITAL Total")
    ws.cell(row=37, column=4, value='=SUMIF(C5:C34,"CAPITAL",D5:D34)').number_format = MONEY_FMT
    ws.cell(row=37, column=5, value='=SUMIF(C5:C34,"CAPITAL",E5:E34)').number_format = MONEY_FMT
    ws.cell(row=37, column=6, value='=SUMIF(C5:C34,"CAPITAL",F5:F34)').number_format = MONEY_FMT
    ws.cell(row=37, column=7, value='=SUMIF(C5:C34,"CAPITAL",G5:G34)').number_format = MONEY_FMT
    ws.cell(row=37, column=7).font = Font(bold=True)
    ws.cell(row=37, column=7).fill = CALC_FILL
    
    ws.cell(row=38, column=2, value="LOAN Total")
    ws.cell(row=38, column=4, value='=SUMIF(C5:C34,"LOAN",D5:D34)').number_format = MONEY_FMT
    ws.cell(row=38, column=5, value='=SUMIF(C5:C34,"LOAN",E5:E34)').number_format = MONEY_FMT
    ws.cell(row=38, column=6, value='=SUMIF(C5:C34,"LOAN",F5:F34)').number_format = MONEY_FMT
    ws.cell(row=38, column=7, value='=SUMIF(C5:C34,"LOAN",G5:G34)').number_format = MONEY_FMT
    ws.cell(row=38, column=7).font = Font(bold=True)
    ws.cell(row=38, column=7).fill = CALC_FILL
    
    ws.cell(row=39, column=2, value="AMANAT Total")
    ws.cell(row=39, column=4, value='=SUMIF(C5:C34,"AMANAT",D5:D34)').number_format = MONEY_FMT
    ws.cell(row=39, column=5, value='=SUMIF(C5:C34,"AMANAT",E5:E34)').number_format = MONEY_FMT
    ws.cell(row=39, column=6, value='=SUMIF(C5:C34,"AMANAT",F5:F34)').number_format = MONEY_FMT
    ws.cell(row=39, column=7, value='=SUMIF(C5:C34,"AMANAT",G5:G34)').number_format = MONEY_FMT
    ws.cell(row=39, column=7).font = Font(bold=True)
    ws.cell(row=39, column=7).fill = CALC_FILL
    
    ws.cell(row=41, column=2, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=41, column=4, value='=SUM(D37:D39)').number_format = MONEY_FMT
    ws.cell(row=41, column=4).font = Font(bold=True)
    ws.cell(row=41, column=5, value='=SUM(E37:E39)').number_format = MONEY_FMT
    ws.cell(row=41, column=5).font = Font(bold=True)
    ws.cell(row=41, column=6, value='=SUM(F37:F39)').number_format = MONEY_FMT
    ws.cell(row=41, column=6).font = Font(bold=True)
    ws.cell(row=41, column=7, value='=SUM(G37:G39)').number_format = MONEY_FMT
    ws.cell(row=41, column=7).font = Font(bold=True, size=12)
    ws.cell(row=41, column=7).fill = CALC_FILL
    
    for c, w in enumerate([4, 22, 10, 14, 14, 18, 16], 1): col_w(ws, c, w)
    
    # ========== BEPAARI_AAKDA ==========
    ws = wb.create_sheet("Bepaari_Aakda", 7)
    ws.sheet_properties.tabColor = "795548"
    ws.cell(row=1, column=1, value="BEPAARI SETTLEMENT SLIP (AAKDA)").font = TITLE_FONT
    ws.cell(row=3, column=1, value="SELECT BEPAARI:")
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ak_dv = DataValidation(type="list", formula1='BEPAARI', allow_blank=False)
    ws.add_data_validation(ak_dv)
    ak_dv.add('B3')
    aakda = [["", ""], ["A. SALES", ""], ["Qty Sold", '=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$E:$E)'], ["Gross Sales", '=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$G:$G)'], ["", ""], ["B. DEDUCTIONS", ""], ["Commission %", '=IFERROR(INDEX(Masters!$C:$C,MATCH($B$3,Masters!$B:$B,0)),4)'], ["Commission Amt", '=B9*B12/100'], ["KK (Fixed)", '=IF(B8>0,Masters!$X$3,0)'], ["JB", '=B8*Masters!$X$4'], ["Motor", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"MOTOR",Cash_Book!$E:$E,$B$3)'], ["Bhussa", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"BHUSSA",Cash_Book!$E:$E,$B$3)'], ["Gawali", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"GAWALI",Cash_Book!$E:$E,$B$3)'], ["Cash/Adv", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"CASH_ADV",Cash_Book!$E:$E,$B$3)'], ["TOTAL DEDUCTIONS", '=SUM(B13:B19)'], ["", ""], ["C. CALCULATION", ""], ["Gross Sales", '=B9'], ["(-) Deductions", '=B20'], ["(+) Opening Bal", '=IFERROR(INDEX(Masters!$D:$D,MATCH($B$3,Masters!$B:$B,0)),0)'], ["NET PAYABLE", '=B23-B24+B25'], ["", ""], ["D. PAYMENTS", ""], ["Payments Made", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,$B$3)'], ["", ""], ["BALANCE DUE", '=B26-B29']]
    for r, (lbl, frm) in enumerate(aakda, 6):
        ws.cell(row=r, column=1, value=lbl)
        if frm:
            c = ws.cell(row=r, column=2, value=frm)
            c.number_format = MONEY_FMT if "Qty" not in lbl and "%" not in lbl else NUM_FMT
        if any(x in lbl for x in ["A.", "B.", "C.", "D.", "TOTAL", "NET", "BALANCE"]):
            ws.cell(row=r, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl or "BALANCE" in lbl:
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = CALC_FILL
    col_w(ws, 1, 22)
    col_w(ws, 2, 14)
    
    # ========== BALANCE_SHEET - COMPACT (No blank rows) ==========
    ws = wb.create_sheet("Balance_Sheet", 8)
    ws.sheet_properties.tabColor = "607D8B"
    ws.cell(row=1, column=1, value="BALANCE SHEET").font = TITLE_FONT
    ws.cell(row=1, column=5, value="LIABILITIES = ASSETS").font = Font(italic=True, size=10, color="FF0000")
    
    # ===== LIABILITIES =====
    ws.cell(row=3, column=1, value="LIABILITIES (We Owe)").font = SUBTITLE_FONT
    ws.cell(row=3, column=1).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    ws.cell(row=4, column=1, value="Particulars")
    ws.cell(row=4, column=2, value="Amount")
    hdr(ws, 4, 1, 2, HEADER_RED)
    
    row = 5
    # CAPITAL - pulls from Capital_Loan_Ledger
    ws.cell(row=row, column=1, value="CAPITAL (from Investors)").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Capital_Loan_Ledger!G37').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    row += 1
    
    # LOANS
    ws.cell(row=row, column=1, value="LOANS (from Lenders)").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Capital_Loan_Ledger!G38').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    row += 1
    
    # AMANAT
    ws.cell(row=row, column=1, value="AMANAT (Trust Deposits)").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Capital_Loan_Ledger!G39').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    row += 2
    
    # BEPAARI PAYABLES
    ws.cell(row=row, column=1, value="BEPAARI PAYABLES").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Bepaari_Ledger!R57').number_format = MONEY_FMT
    bepaari_pay_row = row
    row += 1
    
    # DUKANDAR ADVANCES (overpaid)
    ws.cell(row=row, column=1, value="DUKANDAR ADVANCES (Overpaid)").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Dukandar_Ledger!H58').number_format = MONEY_FMT
    duk_adv_row = row
    row += 2
    
    # COMMISSION NET
    ws.cell(row=row, column=1, value="COMMISSION (Net)").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Bepaari_Ledger!G54-Dukandar_Ledger!E54').number_format = MONEY_FMT
    comm_row = row
    row += 2
    
    # JB NET
    ws.cell(row=row, column=1, value="JB (Net)").font = Font(bold=True)
    jb_start = row
    row += 1
    ws.cell(row=row, column=1, value="  Collected from Bepaari")
    ws.cell(row=row, column=2, value='=Bepaari_Ledger!I54').number_format = MONEY_FMT
    jb_coll_row = row
    row += 1
    ws.cell(row=row, column=1, value="  (-) Paid to Market")
    ws.cell(row=row, column=2, value='=-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"JB_PAID")').number_format = MONEY_FMT
    jb_paid_row = row
    row += 1
    ws.cell(row=row, column=1, value="  JB Net Total").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f'=B{jb_coll_row}+B{jb_paid_row}').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    jb_net_row = row
    row += 2
    
    # KK COLLECTED
    ws.cell(row=row, column=1, value="KK COLLECTED")
    ws.cell(row=row, column=2, value='=Bepaari_Ledger!H54').number_format = MONEY_FMT
    kk_row = row
    row += 2
    
    # ZAKAT PAYABLE
    ws.cell(row=row, column=1, value="ZAKAT PAYABLE").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Masters!$X$9+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ZAKAT",Cash_Book!$D:$D,"PROVISION")-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ZAKAT",Cash_Book!$D:$D,"PAID")').number_format = MONEY_FMT
    zakat_row = row
    row += 2
    
    # TOTAL LIABILITIES
    total_liab_row = row
    ws.cell(row=row, column=1, value="TOTAL LIABILITIES").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=f'=B5+B6+B7+B{bepaari_pay_row}+B{duk_adv_row}+B{comm_row}+B{jb_net_row}+B{kk_row}+B{zakat_row}').number_format = MONEY_FMT
    ws.cell(row=row, column=2).font = Font(bold=True, size=12)
    ws.cell(row=row, column=2).fill = CALC_FILL
    
    # ===== ASSETS =====
    ws.cell(row=3, column=4, value="ASSETS (We Have / Owed to Us)").font = SUBTITLE_FONT
    ws.cell(row=3, column=4).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    ws.cell(row=4, column=4, value="Particulars")
    ws.cell(row=4, column=5, value="Amount")
    hdr(ws, 4, 4, 5, HEADER_GREEN)
    
    arow = 5
    # CASH
    ws.cell(row=arow, column=4, value="CASH BALANCE").font = Font(bold=True)
    ws.cell(row=arow, column=5, value='=Cash_Book!J10').number_format = MONEY_FMT
    ws.cell(row=arow, column=5).fill = CALC_FILL
    cash_row = arow
    arow += 1
    
    # BANK
    ws.cell(row=arow, column=4, value="BANK BALANCE").font = Font(bold=True)
    ws.cell(row=arow, column=5, value='=Cash_Book!J11').number_format = MONEY_FMT
    ws.cell(row=arow, column=5).fill = CALC_FILL
    bank_row = arow
    arow += 2
    
    # PATTI (Dukandar Receivable)
    ws.cell(row=arow, column=4, value="PATTI (Dukandar Receivable)").font = Font(bold=True)
    ws.cell(row=arow, column=5, value='=Dukandar_Ledger!H57').number_format = MONEY_FMT
    patti_row = arow
    arow += 1
    
    # BEPAARI ADVANCES (overpaid)
    ws.cell(row=arow, column=4, value="BEPAARI ADVANCES (Overpaid)").font = Font(bold=True)
    ws.cell(row=arow, column=5, value='=Bepaari_Ledger!R58').number_format = MONEY_FMT
    bep_adv_row = arow
    arow += 2
    
    # ADVANCES GIVEN
    ws.cell(row=arow, column=4, value="ADVANCES GIVEN (Net)").font = Font(bold=True)
    ws.cell(row=arow, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"GIVEN")-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"RECEIVED")').number_format = MONEY_FMT
    adv_row = arow
    arow += 2
    
    # EXPENSES
    ws.cell(row=arow, column=4, value="EXPENSES").font = Font(bold=True)
    arow += 1
    ws.cell(row=arow, column=4, value="  Mandi/Business")
    ws.cell(row=arow, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MANDI")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"TRAVEL")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"FOOD")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"SALARY")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"OTHER")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MISC")').number_format = MONEY_FMT
    mandi_row = arow
    arow += 1
    ws.cell(row=arow, column=4, value="  BF Discount")
    ws.cell(row=arow, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"BF_DISC")').number_format = MONEY_FMT
    bf_row = arow
    arow += 2
    
    # TOTAL ASSETS
    ws.cell(row=total_liab_row, column=4, value="TOTAL ASSETS").font = Font(bold=True, size=12)
    ws.cell(row=total_liab_row, column=5, value=f'=E{cash_row}+E{bank_row}+E{patti_row}+E{bep_adv_row}+E{adv_row}+E{mandi_row}+E{bf_row}').number_format = MONEY_FMT
    ws.cell(row=total_liab_row, column=5).font = Font(bold=True, size=12)
    ws.cell(row=total_liab_row, column=5).fill = CALC_FILL
    
    # DIFFERENCE
    ws.cell(row=total_liab_row+2, column=1, value="DIFFERENCE (Should be 0)").font = Font(bold=True, color="FF0000")
    ws.cell(row=total_liab_row+2, column=2, value=f'=B{total_liab_row}-E{total_liab_row}').number_format = MONEY_FMT
    ws.cell(row=total_liab_row+2, column=2).font = Font(bold=True, color="FF0000")
    
    # Note about detailed breakdown
    ws.cell(row=total_liab_row+4, column=1, value="Note: For individual Capital/Loan party breakdown, see 'Capital_Loan_Ledger' sheet").font = Font(italic=True, size=9)
    
    col_w(ws, 1, 30)
    col_w(ws, 2, 14)
    col_w(ws, 3, 3)
    col_w(ws, 4, 30)
    col_w(ws, 5, 14)
    
    # ========== COMMISSION_SUMMARY ==========
    ws = wb.create_sheet("Commission_Summary", 9)
    ws.sheet_properties.tabColor = "FFC107"
    ws.cell(row=1, column=1, value="COMMISSION & PROFIT SUMMARY").font = TITLE_FONT
    items = [["", ""], ["GROSS SALES", "=Bepaari_Ledger!E54"], ["TOTAL GOATS SOLD", "=Bepaari_Ledger!F54"], ["", ""], ["INCOME:", ""], ["  Commission Earned", "=Bepaari_Ledger!G54"], ["  (-) Discounts Given", "=-Dukandar_Ledger!E54"], ["  Net Commission", "=B6+B7"], ["  JB Collected", "=Bepaari_Ledger!I54"], ["  KK Collected", "=Bepaari_Ledger!H54"], ["TOTAL INCOME", "=B8+B9+B10"], ["", ""], ["EXPENSES:", ""], ["  Mandi/Business", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MANDI")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"TRAVEL")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"FOOD")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"SALARY")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"OTHER")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MISC")'], ["  BF Discount", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"BF_DISC")'], ["  JB Paid", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"JB_PAID")'], ["TOTAL EXPENSES", "=B14+B15+B16"], ["", ""], ["NET PROFIT/LOSS", "=B11-B17"]]
    for r, (lbl, frm) in enumerate(items, 2):
        ws.cell(row=r, column=1, value=lbl)
        if frm:
            c = ws.cell(row=r, column=2, value=frm)
            c.number_format = MONEY_FMT if "GOATS" not in lbl else NUM_FMT
        if any(x in lbl for x in ["GROSS", "TOTAL", "NET", "INCOME:", "EXPENSES:"]):
            ws.cell(row=r, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl:
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = CALC_FILL
    col_w(ws, 1, 26)
    col_w(ws, 2, 14)
    
    wb.save(path)
    print(f"Mandi Master V9 created: {path}")

if __name__ == "__main__":
    create_v9("/app/frontend/public/Mandi_Master_V9.xlsx")
