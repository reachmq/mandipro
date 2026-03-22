"""
Mandi Master Excel Generator - VERSION 4.0
Simplified: Daily_Sales + Cash_Book = Everything
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

# Styles
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_RED = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
HEADER_GREEN = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="0D47A1")
SUBTITLE_FONT = Font(bold=True, size=11, color="1565C0")
INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
CALC_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
MONEY_FMT = '₹#,##0'
NUM_FMT = '#,##0'
DATE_FMT = 'DD-MMM-YYYY'
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def col_w(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def hdr(ws, row, c1, c2, fill=None):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def create_v4(path):
    wb = Workbook()
    wb.remove(wb.active)
    
    # ================================================================
    # 1. MASTERS
    # ================================================================
    ws = wb.create_sheet("Masters", 0)
    ws.sheet_properties.tabColor = "1565C0"
    
    # BEPAARI (A-F)
    ws.cell(row=1, column=1, value="BEPAARI MASTER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Bepaari Name", "Comm%", "Opening Bal", "Phone", "Remarks"], 1):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 1, 6)
    
    bepaaris = [
        "ABUL HASAN", "SHARAFAT", "SHARIK", "JUNAID", "ANNU SHARIF", "GAYYUR", 
        "ARIF KK", "KISHOR", "SHOEB BYCULLA", "MAULANA", "ISLAM D", "SALIM RAJ",
        "MEHMOOD MADH", "ANAS RAJ", "MOBIN CHIPLUN", "AYAZ", "JALIL", "NEERAJ",
        "IRFAN JOGESHWARI", "CHINTU SHUJALPUR"
    ]
    for i, name in enumerate(bepaaris, 1):
        ws.cell(row=i+2, column=1, value=i)
        ws.cell(row=i+2, column=2, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=3, value=4).fill = INPUT_FILL  # 4%
        ws.cell(row=i+2, column=4, value=0).fill = INPUT_FILL
        ws.cell(row=i+2, column=4).number_format = MONEY_FMT
    for r in range(len(bepaaris)+3, 80):
        ws.cell(row=r, column=1, value=r-2)
        for c in range(2, 7):
            ws.cell(row=r, column=c).fill = INPUT_FILL
    
    # DUKANDAR (H-L)
    ws.cell(row=1, column=8, value="DUKANDAR MASTER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Dukandar Name", "Opening Bal", "Phone", "Remarks"], 8):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 8, 12)
    
    dukandars = [
        "SAKIM", "SHARIK", "YUNUS", "MUDASSIR", "ARIF", "SHOEB BYCULLA", "ISLAM D",
        "MEHMOOD MADH", "MOBIN CHIPLUN", "JALIL", "IRFAN JOGESHWARI", "YUNUS DOMBIVLI",
        "IFTEKHAR PUNE", "LALA/ASIF", "JAFER KASHIMIRA", "CHAND", "ASIF GOREGAON",
        "CHINTU SHUJALPUR", "VAZIR", "AKRAM", "NADEEM BHIWANDI", "SHAKIL KON", "PARVEZ"
    ]
    for i, name in enumerate(dukandars, 1):
        ws.cell(row=i+2, column=8, value=i)
        ws.cell(row=i+2, column=9, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=10, value=0).fill = INPUT_FILL
        ws.cell(row=i+2, column=10).number_format = MONEY_FMT
    for r in range(len(dukandars)+3, 80):
        ws.cell(row=r, column=8, value=r-2)
        for c in range(9, 13):
            ws.cell(row=r, column=c).fill = INPUT_FILL
    
    # CAPITAL/LOAN PARTIES (N-Q)
    ws.cell(row=1, column=14, value="CAPITAL/LOAN PARTIES").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Party Name", "Opening Bal", "Type"], 14):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 14, 17)
    
    cap_parties = ["MHN", "SHAKIL GHODEGAON"]
    for i, name in enumerate(cap_parties, 1):
        ws.cell(row=i+2, column=14, value=i)
        ws.cell(row=i+2, column=15, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=16, value=0).fill = INPUT_FILL
        ws.cell(row=i+2, column=16).number_format = MONEY_FMT
        ws.cell(row=i+2, column=17, value="CAPITAL" if name=="MHN" else "LOAN").fill = INPUT_FILL
    for r in range(len(cap_parties)+3, 25):
        ws.cell(row=r, column=14, value=r-2)
        for c in range(15, 18):
            ws.cell(row=r, column=c).fill = INPUT_FILL
    
    # Type dropdown for Capital/Loan
    type_dv = DataValidation(type="list", formula1='"CAPITAL,LOAN,AMANAT"', allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add('Q3:Q24')
    
    # ADVANCE PARTIES (S-U) - People we give advances to
    ws.cell(row=1, column=19, value="ADVANCE PARTIES").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Party Name", "Opening Bal"], 19):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 19, 21)
    for r in range(3, 25):
        ws.cell(row=r, column=19, value=r-2)
        for c in range(20, 22):
            ws.cell(row=r, column=c).fill = INPUT_FILL
    
    # SETTINGS (W-X)
    ws.cell(row=1, column=23, value="SETTINGS").font = TITLE_FONT
    settings = [
        ["Commission Rate (%)", 4],
        ["KK Fixed (₹/Bepari)", 100],
        ["JB Rate (₹/Goat)", 10],
        ["", ""],
        ["Opening Cash", 0],
        ["Opening Bank", 0],
    ]
    for r, (lbl, val) in enumerate(settings, 2):
        ws.cell(row=r, column=23, value=lbl).font = Font(bold=True) if lbl else None
        c = ws.cell(row=r, column=24, value=val)
        c.fill = INPUT_FILL
        if "Cash" in str(lbl) or "Bank" in str(lbl):
            c.number_format = MONEY_FMT
    
    # Column widths
    for c, w in {1:4, 2:20, 3:8, 4:12, 5:10, 6:10, 7:2, 8:4, 9:20, 10:12, 11:10, 12:10, 13:2,
                 14:4, 15:20, 16:12, 17:10, 18:2, 19:4, 20:20, 21:12, 22:2, 23:22, 24:12}.items():
        col_w(ws, c, w)
    
    # ================================================================
    # 2. DAILY_SALES
    # ================================================================
    ws = wb.create_sheet("Daily_Sales", 1)
    ws.sheet_properties.tabColor = "4CAF50"
    
    ws.cell(row=1, column=1, value="DAILY SALES (Bepaari → Dukandar)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="This sheet feeds: Bepaari Ledger, Dukandar Ledger, Commission calculation").font = Font(italic=True, size=9)
    
    for c, h in enumerate(["Sr", "Date", "Bepaari", "Dukandar", "Qty", "Rate", "Gross", "Discount", "Net"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 9)
    
    # Dropdowns
    bep_dv = DataValidation(type="list", formula1='Masters!$B$3:$B$79', allow_blank=True)
    duk_dv = DataValidation(type="list", formula1='Masters!$I$3:$I$79', allow_blank=True)
    ws.add_data_validation(bep_dv)
    ws.add_data_validation(duk_dv)
    
    for r in range(4, 1004):
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=2).fill = INPUT_FILL
        
        ws.cell(row=r, column=3).fill = INPUT_FILL
        bep_dv.add(f'C{r}')
        
        ws.cell(row=r, column=4).fill = INPUT_FILL
        duk_dv.add(f'D{r}')
        
        ws.cell(row=r, column=5).fill = INPUT_FILL  # Qty
        ws.cell(row=r, column=6).fill = INPUT_FILL  # Rate
        ws.cell(row=r, column=6).number_format = MONEY_FMT
        
        # Gross = Qty × Rate
        ws.cell(row=r, column=7, value=f'=IF(E{r}="","",E{r}*F{r})')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=7).fill = CALC_FILL
        
        ws.cell(row=r, column=8).fill = INPUT_FILL  # Discount
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        
        # Net = Gross - Discount
        ws.cell(row=r, column=9, value=f'=IF(G{r}="","",G{r}-IF(H{r}="",0,H{r}))')
        ws.cell(row=r, column=9).number_format = MONEY_FMT
        ws.cell(row=r, column=9).fill = CALC_FILL
    
    for c, w in enumerate([4, 11, 20, 20, 6, 10, 12, 10, 12], 1):
        col_w(ws, c, w)
    
    # ================================================================
    # 3. CASH_BOOK (The Heart of the System)
    # ================================================================
    ws = wb.create_sheet("Cash_Book", 2)
    ws.sheet_properties.tabColor = "FF5722"
    
    ws.cell(row=1, column=1, value="CASH BOOK (All Transactions)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="This sheet feeds: Bepaari Ledger, Dukandar Ledger, Balance Sheet, P&L").font = Font(italic=True, size=9)
    
    for c, h in enumerate(["Sr", "Date", "Type", "Sub-Type", "Party", "Particulars", "Cash OUT", "Bank OUT", "Cash IN", "Bank IN"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 10)
    
    # Main Type dropdown
    main_types = "BEPAARI,DUKANDAR,CAPITAL,LOAN,ADVANCE,EXPENSE,ZAKAT,OTHER"
    main_dv = DataValidation(type="list", formula1=f'"{main_types}"', allow_blank=True)
    ws.add_data_validation(main_dv)
    
    # Sub-Type dropdown (user selects based on main type)
    sub_types = "PAYMENT,MOTOR,BHUSSA,GAWALI,CASH_ADV,OTHER,RECEIPT,TAKEN,REPAID,GIVEN,RECEIVED,MANDI,TRAVEL,FOOD,SALARY,JB_PAID,BF_DISC"
    sub_dv = DataValidation(type="list", formula1=f'"{sub_types}"', allow_blank=True)
    ws.add_data_validation(sub_dv)
    
    for r in range(4, 2004):  # 2000 rows
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=2).fill = INPUT_FILL
        
        ws.cell(row=r, column=3).fill = INPUT_FILL  # Type
        main_dv.add(f'C{r}')
        
        ws.cell(row=r, column=4).fill = INPUT_FILL  # Sub-Type
        sub_dv.add(f'D{r}')
        
        ws.cell(row=r, column=5).fill = INPUT_FILL  # Party
        ws.cell(row=r, column=6).fill = INPUT_FILL  # Particulars
        
        for col in [7, 8, 9, 10]:
            ws.cell(row=r, column=col).fill = INPUT_FILL
            ws.cell(row=r, column=col).number_format = MONEY_FMT
    
    # Running totals
    ws.cell(row=1, column=12, value="CASH/BANK SUMMARY").font = SUBTITLE_FONT
    tots = [
        ["Opening Cash", "=Masters!$X$6"],
        ["Opening Bank", "=Masters!$X$7"],
        ["Total Cash OUT", "=SUM(G:G)"],
        ["Total Bank OUT", "=SUM(H:H)"],
        ["Total Cash IN", "=SUM(I:I)"],
        ["Total Bank IN", "=SUM(J:J)"],
        ["CLOSING CASH", "=L2-L4+L6"],
        ["CLOSING BANK", "=L3-L5+L7"],
    ]
    for r, (lbl, frm) in enumerate(tots, 2):
        ws.cell(row=r, column=11, value=lbl)
        ws.cell(row=r, column=12, value=frm).number_format = MONEY_FMT
        if "CLOSING" in lbl:
            ws.cell(row=r, column=11).font = Font(bold=True)
            ws.cell(row=r, column=12).font = Font(bold=True)
            ws.cell(row=r, column=12).fill = CALC_FILL
    
    for c, w in enumerate([4, 11, 10, 10, 20, 20, 12, 12, 12, 12, 2, 14, 14], 1):
        col_w(ws, c, w)
    
    # ================================================================
    # 4. BEPAARI_LEDGER
    # ================================================================
    ws = wb.create_sheet("Bepaari_Ledger", 3)
    ws.sheet_properties.tabColor = "9C27B0"
    
    ws.cell(row=1, column=1, value="BEPAARI LEDGER (Auto-calculated)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Data from: Daily_Sales (sales) + Cash_Book (payments, expenses)").font = Font(italic=True, size=9)
    
    hdrs = ["Sr", "Bepaari", "Comm%", "Opening", "Gross Sales", "Qty", "Commission", "KK", "JB",
            "Motor", "Bhussa", "Gawali", "Cash/Adv", "Other", "Tot.Deduct", "Net Payable", "Payments", "BALANCE"]
    for c, h in enumerate(hdrs, 1):
        ws.cell(row=4, column=c, value=h)
    hdr(ws, 4, 1, 18)
    
    for r in range(5, 82):  # 77 bepaaris
        m = r - 2  # Masters row
        
        ws.cell(row=r, column=1, value=r-4)
        
        # Name from Masters
        ws.cell(row=r, column=2, value=f'=IF(Masters!$B${m}="","",Masters!$B${m})')
        
        # Comm%
        ws.cell(row=r, column=3, value=f'=IF(B{r}="","",Masters!$C${m})')
        
        # Opening
        ws.cell(row=r, column=4, value=f'=IF(B{r}="",0,Masters!$D${m})')
        ws.cell(row=r, column=4).number_format = MONEY_FMT
        
        # Gross Sales (from Daily_Sales)
        ws.cell(row=r, column=5, value=f'=SUMIF(Daily_Sales!$C:$C,B{r},Daily_Sales!$G:$G)')
        ws.cell(row=r, column=5).number_format = MONEY_FMT
        
        # Qty
        ws.cell(row=r, column=6, value=f'=SUMIF(Daily_Sales!$C:$C,B{r},Daily_Sales!$E:$E)')
        ws.cell(row=r, column=6).number_format = NUM_FMT
        
        # Commission
        ws.cell(row=r, column=7, value=f'=IF(B{r}="",0,E{r}*C{r}/100)')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        
        # KK - ₹100 per Bepari (if sales > 0)
        ws.cell(row=r, column=8, value=f'=IF(F{r}>0,Masters!$X$3,0)')
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        ws.cell(row=r, column=8).fill = CALC_FILL
        
        # JB - ₹10 per goat
        ws.cell(row=r, column=9, value=f'=F{r}*Masters!$X$4')
        ws.cell(row=r, column=9).number_format = MONEY_FMT
        ws.cell(row=r, column=9).fill = CALC_FILL
        
        # Motor (from Cash_Book: Type=BEPAARI, Sub=MOTOR)
        ws.cell(row=r, column=10, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"MOTOR",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"MOTOR",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=10).number_format = MONEY_FMT
        
        # Bhussa
        ws.cell(row=r, column=11, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"BHUSSA",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"BHUSSA",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=11).number_format = MONEY_FMT
        
        # Gawali
        ws.cell(row=r, column=12, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"GAWALI",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"GAWALI",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=12).number_format = MONEY_FMT
        
        # Cash/Adv
        ws.cell(row=r, column=13, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"CASH_ADV",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"CASH_ADV",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=13).number_format = MONEY_FMT
        
        # Other
        ws.cell(row=r, column=14, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"OTHER",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"OTHER",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=14).number_format = MONEY_FMT
        
        # Total Deductions = Comm + KK + JB + Motor + Bhussa + Gawali + Cash + Other
        ws.cell(row=r, column=15, value=f'=SUM(G{r}:N{r})')
        ws.cell(row=r, column=15).number_format = MONEY_FMT
        ws.cell(row=r, column=15).fill = CALC_FILL
        
        # Net Payable = Gross - Deductions + Opening
        ws.cell(row=r, column=16, value=f'=IF(B{r}="","",E{r}-O{r}+D{r})')
        ws.cell(row=r, column=16).number_format = MONEY_FMT
        ws.cell(row=r, column=16).fill = CALC_FILL
        
        # Payments (Type=BEPAARI, Sub=PAYMENT)
        ws.cell(row=r, column=17, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=17).number_format = MONEY_FMT
        
        # Balance = Net Payable - Payments
        ws.cell(row=r, column=18, value=f'=IF(B{r}="","",P{r}-Q{r})')
        ws.cell(row=r, column=18).number_format = MONEY_FMT
        ws.cell(row=r, column=18).font = Font(bold=True)
    
    # Totals
    tot = 82
    ws.cell(row=tot, column=1, value="TOTAL").font = Font(bold=True)
    for c in [4,5,6,7,8,9,10,11,12,13,14,15,16,17,18]:
        ws.cell(row=tot, column=c, value=f'=SUM({get_column_letter(c)}5:{get_column_letter(c)}81)')
        ws.cell(row=tot, column=c).number_format = MONEY_FMT
        ws.cell(row=tot, column=c).font = Font(bold=True)
    
    for c, w in enumerate([4,16,6,10,12,6,10,8,8,9,9,9,10,8,11,12,11,12], 1):
        col_w(ws, c, w)
    
    # ================================================================
    # 5. DUKANDAR_LEDGER
    # ================================================================
    ws = wb.create_sheet("Dukandar_Ledger", 4)
    ws.sheet_properties.tabColor = "E91E63"
    
    ws.cell(row=1, column=1, value="DUKANDAR LEDGER (Auto-calculated)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Data from: Daily_Sales (purchases) + Cash_Book (receipts)").font = Font(italic=True, size=9)
    
    for c, h in enumerate(["Sr", "Dukandar", "Opening", "Purchases", "Discounts", "Net Receivable", "Receipts", "BALANCE"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 8)
    
    for r in range(4, 81):
        m = r - 1  # Masters row
        
        ws.cell(row=r, column=1, value=r-3)
        
        # Name
        ws.cell(row=r, column=2, value=f'=IF(Masters!$I${m}="","",Masters!$I${m})')
        
        # Opening
        ws.cell(row=r, column=3, value=f'=IF(B{r}="",0,Masters!$J${m})')
        ws.cell(row=r, column=3).number_format = MONEY_FMT
        
        # Purchases (Gross from Daily_Sales)
        ws.cell(row=r, column=4, value=f'=SUMIF(Daily_Sales!$D:$D,B{r},Daily_Sales!$G:$G)')
        ws.cell(row=r, column=4).number_format = MONEY_FMT
        
        # Discounts
        ws.cell(row=r, column=5, value=f'=SUMIF(Daily_Sales!$D:$D,B{r},Daily_Sales!$H:$H)')
        ws.cell(row=r, column=5).number_format = MONEY_FMT
        
        # Net Receivable = Purchases - Discounts + Opening
        ws.cell(row=r, column=6, value=f'=IF(B{r}="","",D{r}-E{r}+C{r})')
        ws.cell(row=r, column=6).number_format = MONEY_FMT
        ws.cell(row=r, column=6).fill = CALC_FILL
        
        # Receipts (Type=DUKANDAR, Sub=RECEIPT)
        ws.cell(row=r, column=7, value=f'=SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"DUKANDAR",Cash_Book!$D:$D,"RECEIPT",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$J:$J,Cash_Book!$C:$C,"DUKANDAR",Cash_Book!$D:$D,"RECEIPT",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        
        # Balance
        ws.cell(row=r, column=8, value=f'=IF(B{r}="","",F{r}-G{r})')
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        ws.cell(row=r, column=8).font = Font(bold=True)
    
    # Totals
    tot = 81
    ws.cell(row=tot, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tot, column=2, value="PATTI (Total Dukandar O/S)")
    for c in [3,4,5,6,7,8]:
        ws.cell(row=tot, column=c, value=f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}80)')
        ws.cell(row=tot, column=c).number_format = MONEY_FMT
        ws.cell(row=tot, column=c).font = Font(bold=True)
    
    for c, w in enumerate([4,20,12,14,12,14,12,14], 1):
        col_w(ws, c, w)
    
    # ================================================================
    # 6. BEPAARI_AAKDA
    # ================================================================
    ws = wb.create_sheet("Bepaari_Aakda", 5)
    ws.sheet_properties.tabColor = "795548"
    
    ws.cell(row=1, column=1, value="BEPAARI SETTLEMENT SLIP (AAKDA)").font = TITLE_FONT
    
    ws.cell(row=3, column=1, value="SELECT BEPAARI:")
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ak_dv = DataValidation(type="list", formula1='Masters!$B$3:$B$79', allow_blank=False)
    ws.add_data_validation(ak_dv)
    ak_dv.add('B3')
    
    ws.cell(row=4, column=1, value="DATE:")
    ws.cell(row=4, column=2).fill = INPUT_FILL
    ws.cell(row=4, column=2).number_format = DATE_FMT
    
    aakda = [
        ["", ""],
        ["A. SALES", ""],
        ["Qty Sold", '=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$E:$E)'],
        ["Gross Sales", '=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$G:$G)'],
        ["", ""],
        ["B. DEDUCTIONS", ""],
        ["Commission %", '=IFERROR(INDEX(Masters!$C:$C,MATCH($B$3,Masters!$B:$B,0)),4)'],
        ["Commission Amt", '=B9*B12/100'],
        ["KK (Fixed)", '=IF(B8>0,Masters!$X$3,0)'],
        ["JB (₹10/goat)", '=B8*Masters!$X$4'],
        ["Motor", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"MOTOR",Cash_Book!$E:$E,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"MOTOR",Cash_Book!$E:$E,$B$3)'],
        ["Bhussa", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"BHUSSA",Cash_Book!$E:$E,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"BHUSSA",Cash_Book!$E:$E,$B$3)'],
        ["Gawali", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"GAWALI",Cash_Book!$E:$E,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"GAWALI",Cash_Book!$E:$E,$B$3)'],
        ["Cash/Advance", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"CASH_ADV",Cash_Book!$E:$E,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"CASH_ADV",Cash_Book!$E:$E,$B$3)'],
        ["TOTAL DEDUCTIONS", '=SUM(B13:B19)'],
        ["", ""],
        ["C. CALCULATION", ""],
        ["Gross Sales", '=B9'],
        ["(-) Deductions", '=B20'],
        ["(+) Opening Bal", '=IFERROR(INDEX(Masters!$D:$D,MATCH($B$3,Masters!$B:$B,0)),0)'],
        ["NET PAYABLE", '=B23-B24+B25'],
        ["", ""],
        ["D. PAYMENTS", ""],
        ["Payments Made", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,$B$3)'],
        ["", ""],
        ["BALANCE DUE", '=B26-B29'],
    ]
    
    for r, (lbl, frm) in enumerate(aakda, 6):
        ws.cell(row=r, column=1, value=lbl)
        if frm:
            c = ws.cell(row=r, column=2, value=frm)
            c.number_format = MONEY_FMT if "Qty" not in lbl and "%" not in lbl else NUM_FMT
        if any(x in lbl for x in ["A.", "B.", "C.", "D.", "TOTAL", "NET", "BALANCE"]):
            ws.cell(row=r, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl or "BALANCE" in lbl:
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = CALC_FILL
    
    col_w(ws, 1, 22)
    col_w(ws, 2, 14)
    
    # ================================================================
    # 7. BALANCE_SHEET
    # ================================================================
    ws = wb.create_sheet("Balance_Sheet", 6)
    ws.sheet_properties.tabColor = "607D8B"
    
    ws.cell(row=1, column=1, value="BALANCE SHEET").font = TITLE_FONT
    ws.cell(row=1, column=5, value="LIABILITIES = ASSETS").font = Font(italic=True, size=10, color="FF0000")
    
    # ----- LIABILITIES (Left) -----
    ws.cell(row=3, column=1, value="LIABILITIES (We Owe)").font = SUBTITLE_FONT
    ws.cell(row=3, column=1).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    ws.cell(row=4, column=1, value="Particulars")
    ws.cell(row=4, column=2, value="Amount")
    hdr(ws, 4, 1, 2, HEADER_RED)
    
    # Capital - from Cash_Book (Type=CAPITAL, Sub=TAKEN) + Opening
    ws.cell(row=5, column=1, value="CAPITAL:")
    ws.cell(row=5, column=1).font = Font(bold=True)
    
    # Dynamic Capital/Loan parties from Cash_Book
    ws.cell(row=6, column=1, value="  (From Cash_Book)")
    ws.cell(row=6, column=2, value='=SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"CAPITAL",Cash_Book!$D:$D,"TAKEN")+SUMIFS(Cash_Book!$J:$J,Cash_Book!$C:$C,"CAPITAL",Cash_Book!$D:$D,"TAKEN")+SUM(Masters!$P$3:$P$24)')
    ws.cell(row=6, column=2).number_format = MONEY_FMT
    
    ws.cell(row=8, column=1, value="LOANS/AMANAT:")
    ws.cell(row=8, column=1).font = Font(bold=True)
    ws.cell(row=9, column=1, value="  (From Cash_Book)")
    ws.cell(row=9, column=2, value='=SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"LOAN",Cash_Book!$D:$D,"TAKEN")+SUMIFS(Cash_Book!$J:$J,Cash_Book!$C:$C,"LOAN",Cash_Book!$D:$D,"TAKEN")-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"LOAN",Cash_Book!$D:$D,"REPAID")-SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"LOAN",Cash_Book!$D:$D,"REPAID")')
    ws.cell(row=9, column=2).number_format = MONEY_FMT
    
    ws.cell(row=11, column=1, value="BEPAARI PAYABLES:")
    ws.cell(row=11, column=1).font = Font(bold=True)
    ws.cell(row=12, column=1, value="  (From Bepaari_Ledger)")
    ws.cell(row=12, column=2, value='=Bepaari_Ledger!R82')
    ws.cell(row=12, column=2).number_format = MONEY_FMT
    
    ws.cell(row=14, column=1, value="COMMISSION EARNED:")
    ws.cell(row=14, column=1).font = Font(bold=True)
    ws.cell(row=15, column=1, value="  Gross Commission")
    ws.cell(row=15, column=2, value='=Bepaari_Ledger!G82')
    ws.cell(row=15, column=2).number_format = MONEY_FMT
    ws.cell(row=16, column=1, value="  (-) Discounts Given")
    ws.cell(row=16, column=2, value='=-Dukandar_Ledger!E81')
    ws.cell(row=16, column=2).number_format = MONEY_FMT
    ws.cell(row=17, column=1, value="  Net Commission")
    ws.cell(row=17, column=2, value='=B15+B16')
    ws.cell(row=17, column=2).number_format = MONEY_FMT
    ws.cell(row=17, column=2).fill = CALC_FILL
    
    ws.cell(row=19, column=1, value="JB COLLECTED:")
    ws.cell(row=19, column=2, value='=Bepaari_Ledger!I82')
    ws.cell(row=19, column=2).number_format = MONEY_FMT
    
    ws.cell(row=20, column=1, value="KK COLLECTED:")
    ws.cell(row=20, column=2, value='=Bepaari_Ledger!H82')
    ws.cell(row=20, column=2).number_format = MONEY_FMT
    
    # Total Liabilities
    ws.cell(row=22, column=1, value="TOTAL LIABILITIES").font = Font(bold=True)
    ws.cell(row=22, column=2, value='=B6+B9+B12+B17+B19+B20')
    ws.cell(row=22, column=2).number_format = MONEY_FMT
    ws.cell(row=22, column=2).font = Font(bold=True)
    ws.cell(row=22, column=2).fill = CALC_FILL
    
    # ----- ASSETS (Right) -----
    ws.cell(row=3, column=4, value="ASSETS (We Have / Owed to Us)").font = SUBTITLE_FONT
    ws.cell(row=3, column=4).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    ws.cell(row=4, column=4, value="Particulars")
    ws.cell(row=4, column=5, value="Amount")
    hdr(ws, 4, 4, 5, HEADER_GREEN)
    
    ws.cell(row=5, column=4, value="CASH BALANCE:")
    ws.cell(row=5, column=4).font = Font(bold=True)
    ws.cell(row=5, column=5, value='=Cash_Book!L8')
    ws.cell(row=5, column=5).number_format = MONEY_FMT
    
    ws.cell(row=7, column=4, value="BANK BALANCE:")
    ws.cell(row=7, column=4).font = Font(bold=True)
    ws.cell(row=7, column=5, value='=Cash_Book!L9')
    ws.cell(row=7, column=5).number_format = MONEY_FMT
    
    ws.cell(row=9, column=4, value="PATTI (Dukandar O/S):")
    ws.cell(row=9, column=4).font = Font(bold=True)
    ws.cell(row=10, column=4, value="  (From Dukandar_Ledger)")
    ws.cell(row=10, column=5, value='=Dukandar_Ledger!H81')
    ws.cell(row=10, column=5).number_format = MONEY_FMT
    
    ws.cell(row=12, column=4, value="ADVANCES GIVEN:")
    ws.cell(row=12, column=4).font = Font(bold=True)
    ws.cell(row=13, column=4, value="  (From Cash_Book)")
    ws.cell(row=13, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"GIVEN")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"GIVEN")-SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"RECEIVED")-SUMIFS(Cash_Book!$J:$J,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"RECEIVED")')
    ws.cell(row=13, column=5).number_format = MONEY_FMT
    
    ws.cell(row=15, column=4, value="EXPENSES:")
    ws.cell(row=15, column=4).font = Font(bold=True)
    ws.cell(row=16, column=4, value="  Mandi/Business Exp")
    ws.cell(row=16, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE")')
    ws.cell(row=16, column=5).number_format = MONEY_FMT
    
    ws.cell(row=17, column=4, value="  Zakat Paid")
    ws.cell(row=17, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ZAKAT")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"ZAKAT")')
    ws.cell(row=17, column=5).number_format = MONEY_FMT
    
    ws.cell(row=18, column=4, value="  JB Paid to Market")
    ws.cell(row=18, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$D:$D,"JB_PAID")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$D:$D,"JB_PAID")')
    ws.cell(row=18, column=5).number_format = MONEY_FMT
    
    ws.cell(row=19, column=4, value="  BF Discount Earned")
    ws.cell(row=19, column=5, value='=-SUMIFS(Cash_Book!$I:$I,Cash_Book!$D:$D,"BF_DISC")-SUMIFS(Cash_Book!$J:$J,Cash_Book!$D:$D,"BF_DISC")')
    ws.cell(row=19, column=5).number_format = MONEY_FMT
    
    # Total Assets
    ws.cell(row=22, column=4, value="TOTAL ASSETS").font = Font(bold=True)
    ws.cell(row=22, column=5, value='=E5+E7+E10+E13+E16+E17+E18+E19')
    ws.cell(row=22, column=5).number_format = MONEY_FMT
    ws.cell(row=22, column=5).font = Font(bold=True)
    ws.cell(row=22, column=5).fill = CALC_FILL
    
    # Difference
    ws.cell(row=24, column=1, value="DIFFERENCE (Should be 0)").font = Font(bold=True, color="FF0000")
    ws.cell(row=24, column=2, value='=B22-E22')
    ws.cell(row=24, column=2).number_format = MONEY_FMT
    ws.cell(row=24, column=2).font = Font(bold=True, color="FF0000")
    
    col_w(ws, 1, 24)
    col_w(ws, 2, 14)
    col_w(ws, 3, 3)
    col_w(ws, 4, 24)
    col_w(ws, 5, 14)
    
    # ================================================================
    # 8. COMMISSION_SUMMARY
    # ================================================================
    ws = wb.create_sheet("Commission_Summary", 7)
    ws.sheet_properties.tabColor = "FFC107"
    
    ws.cell(row=1, column=1, value="COMMISSION & PROFIT SUMMARY").font = TITLE_FONT
    
    items = [
        ["", ""],
        ["GROSS SALES", "=Bepaari_Ledger!E82"],
        ["TOTAL GOATS SOLD", "=Bepaari_Ledger!F82"],
        ["", ""],
        ["INCOME:", ""],
        ["  Commission Earned", "=Bepaari_Ledger!G82"],
        ["  (-) Discounts Given", "=-Dukandar_Ledger!E81"],
        ["  Net Commission", "=B6+B7"],
        ["  JB Collected", "=Bepaari_Ledger!I82"],
        ["  KK Collected", "=Bepaari_Ledger!H82"],
        ["  BF Discount", '=SUMIFS(Cash_Book!$I:$I,Cash_Book!$D:$D,"BF_DISC")+SUMIFS(Cash_Book!$J:$J,Cash_Book!$D:$D,"BF_DISC")'],
        ["TOTAL INCOME", "=B8+B9+B10+B11"],
        ["", ""],
        ["EXPENSES:", ""],
        ["  Mandi/Business", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE")'],
        ["  Zakat", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ZAKAT")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"ZAKAT")'],
        ["  JB Paid to Market", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$D:$D,"JB_PAID")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$D:$D,"JB_PAID")'],
        ["TOTAL EXPENSES", "=B15+B16+B17"],
        ["", ""],
        ["NET PROFIT/LOSS", "=B12-B18"],
    ]
    
    for r, (lbl, frm) in enumerate(items, 2):
        ws.cell(row=r, column=1, value=lbl)
        if frm:
            c = ws.cell(row=r, column=2, value=frm)
            c.number_format = MONEY_FMT if "GOATS" not in lbl else NUM_FMT
        if any(x in lbl for x in ["GROSS", "TOTAL", "NET", "INCOME:", "EXPENSES:"]):
            ws.cell(row=r, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl:
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = CALC_FILL
    
    col_w(ws, 1, 24)
    col_w(ws, 2, 14)
    
    # ================================================================
    # SAVE
    # ================================================================
    wb.save(path)
    print(f"✅ Mandi Master V4 created: {path}")
    return path


if __name__ == "__main__":
    create_v4("/app/Mandi_Master_V4.xlsx")
