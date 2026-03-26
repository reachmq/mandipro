"""Mandi Master V9.1 - Simple V9 + B/F Balances + Flat Commission + User Data"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_RED = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
HEADER_GREEN = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_PURPLE = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="0D47A1")
SUBTITLE_FONT = Font(bold=True, size=11, color="1565C0")
INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
CALC_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
MONEY_FMT = '₹#,##0'
NUM_FMT = '#,##0'
DATE_FMT = 'DD-MMM-YYYY'

def col_w(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def hdr(ws, row, c1, c2, fill=None):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def create_v91(path):
    wb = Workbook()
    wb.remove(wb.active)
    
    # Simple named ranges - PARTY lists only
    wb.defined_names.add(DefinedName("BEPAARI", attr_text="Masters!$B$3:$B$52"))
    wb.defined_names.add(DefinedName("DUKANDAR", attr_text="Masters!$J$3:$J$52"))
    wb.defined_names.add(DefinedName("CAPITAL", attr_text="Masters!$P$3:$P$32"))
    wb.defined_names.add(DefinedName("LOAN", attr_text="Masters!$P$3:$P$32"))
    wb.defined_names.add(DefinedName("AMANAT", attr_text="Masters!$P$3:$P$32"))
    wb.defined_names.add(DefinedName("ADVANCE", attr_text="Masters!$U$3:$U$22"))
    
    # ========== MASTERS ==========
    ws = wb.create_sheet("Masters", 0)
    ws.sheet_properties.tabColor = "1565C0"
    
    # BEPAARI MASTER - With Flat Comm column
    ws.cell(row=1, column=1, value="BEPAARI MASTER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Bepaari Name", "Comm%", "Flat ₹/Goat", "Opening Bal", "Phone", "Remarks"], 1):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 1, 7)
    
    bepaaris = [
        ["ABUL HASAN", 4, "", 7550],
        ["SHARAFAT", 4, "", 0],
        ["JUNAID", 4, "", 0],
        ["ANNU SHARIF", 4, "", 0],
        ["GAYYUR", 4, "", 142530],
        ["KISHOR", 4, "", -152964],
        ["MAULANA", 4, "", 35330],
        ["SALIM RAJ", 4, "", 0],
        ["ANAS RAJ", 4, "", 0],
        ["AYAZ", 4, "", 1530],
        ["NEERAJ", 4, "", 0],
        ["CHINTU SHUJALPUR", 4, "", 0],
        ["SANNA", 4, "", -14048],
        ["WAHID", 4, "", -4661],
        ["MANSOOR", 4, "", -9845],
        ["AYAZ PREVIOUS BAL", 4, "", -88000],
        ["SHAKIL SHUJALPUR", 4, "", 660],
        ["USMAN", 4, "", 40440],
    ]
    for i, (name, comm, flat, bal) in enumerate(bepaaris, 1):
        ws.cell(row=i+2, column=1, value=i)
        ws.cell(row=i+2, column=2, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=3, value=comm).fill = INPUT_FILL
        ws.cell(row=i+2, column=4, value=flat if flat else "").fill = INPUT_FILL
        ws.cell(row=i+2, column=4).number_format = MONEY_FMT
        ws.cell(row=i+2, column=5, value=bal).fill = INPUT_FILL
        ws.cell(row=i+2, column=5).number_format = MONEY_FMT
    for r in range(len(bepaaris)+3, 53):
        ws.cell(row=r, column=1, value=r-2)
        for c in range(2, 8): ws.cell(row=r, column=c).fill = INPUT_FILL
        ws.cell(row=r, column=3, value=4).fill = INPUT_FILL
        ws.cell(row=r, column=5).number_format = MONEY_FMT
    
    # DUKANDAR MASTER - Column I onwards (9+)
    ws.cell(row=1, column=9, value="DUKANDAR MASTER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Dukandar Name", "Opening Bal", "Phone", "Remarks"], 9):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 9, 13)
    
    dukandars = [
        ["SAKIM", 113500], ["SHARIK", 879000], ["YUNUS", 94931], ["MUDASSIR", 516778],
        ["ARIF", 187650], ["SHOEB BYCULLA", 0], ["ISLAM D", 100000], ["MEHMOOD MADH", 15000],
        ["MOBIN CHIPLUN", 486250], ["JALIL", 90950], ["IRFAN JOGESHWARI", 0], ["YUNUS DOMBIVLI", 59662],
        ["IFTEKHAR PUNE", 2000], ["LALA/ASIF", 0], ["JAFER KASHIMIRA", 92500], ["HANIF KALAM", 0],
        ["SAUD K", 0], ["AKRAM", 0], ["WAHID", 4000], ["VAZIR", 134750],
        ["ISMAIL SOPARA", 0], ["ARIF CHINCHOTI", 0], ["NADEEM BHIWANDI", 265000], ["GABBAR", 33050],
        ["BABU GABBAR", 50000], ["BABU TALOJA", 2600], ["SHAKIL KON", 34000], ["PARVEZ", 0],
        ["I CHO CHO", 20500], ["MD VIRAR", 1900], ["MAYA", 11000], ["NAEEM THORPE", 0],
        ["CHAND", 59000], ["ASIF GOREGAON", 57550], ["SOHAIL MAROL", 0], ["ANEES KM", 70000],
        ["RAFIQ AHMED HAJI", 356250], ["ABP - MAJEED THANA", 0], ["ABP - MAJNU", 0], ["MEHBOOB JOG", 98500],
        ["SIRAJ KURLA", 388000], ["ZUBAIR AZEEZ", 30000], ["BABBU BYC", 192000], ["NAWAZ V", 148800],
        ["LIYAQAT", 7500],
    ]
    for i, (name, bal) in enumerate(dukandars, 1):
        ws.cell(row=i+2, column=9, value=i)
        ws.cell(row=i+2, column=10, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=11, value=bal).fill = INPUT_FILL
        ws.cell(row=i+2, column=11).number_format = MONEY_FMT
    for r in range(len(dukandars)+3, 53):
        ws.cell(row=r, column=9, value=r-2)
        for c in range(10, 14): ws.cell(row=r, column=c).fill = INPUT_FILL
        ws.cell(row=r, column=11).number_format = MONEY_FMT
    
    # CAPITAL/LOAN/AMANAT - Column O onwards (15+)
    ws.cell(row=1, column=15, value="CAPITAL/LOAN/AMANAT").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Party Name", "Opening Bal", "Type"], 15):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 15, 18)
    
    cap_parties = [
        ["MHN CAPITAL", 4469981, "CAPITAL"],
        ["SHAKIL GHODEGAON AMANAT", 2000000, "AMANAT"],
    ]
    for i, (name, bal, typ) in enumerate(cap_parties, 1):
        ws.cell(row=i+2, column=15, value=i)
        ws.cell(row=i+2, column=16, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=17, value=bal).fill = INPUT_FILL
        ws.cell(row=i+2, column=17).number_format = MONEY_FMT
        ws.cell(row=i+2, column=18, value=typ).fill = INPUT_FILL
    for r in range(len(cap_parties)+3, 33):
        ws.cell(row=r, column=15, value=r-2)
        for c in range(16, 19): ws.cell(row=r, column=c).fill = INPUT_FILL
        ws.cell(row=r, column=17).number_format = MONEY_FMT
    type_dv = DataValidation(type="list", formula1='"CAPITAL,LOAN,AMANAT"', allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add('R3:R32')
    
    # ADVANCE PARTIES - Column T onwards (20+)
    ws.cell(row=1, column=20, value="ADVANCE PARTIES").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Party Name", "Opening Bal"], 20):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 20, 22)
    for r in range(3, 23):
        ws.cell(row=r, column=20, value=r-2)
        for c in range(21, 23): ws.cell(row=r, column=c).fill = INPUT_FILL
        ws.cell(row=r, column=22).number_format = MONEY_FMT
    
    # SETTINGS - Column X onwards (24+) - WITH ALL B/F BALANCES
    ws.cell(row=1, column=24, value="SETTINGS & B/F BALANCES").font = TITLE_FONT
    settings = [
        ["Commission Rate (%)", 4],
        ["KK Fixed (₹/Bepari)", 25],
        ["JB Rate (₹/Goat)", 1],
        ["", ""],
        ["Opening Cash", 394906],
        ["Opening Bank", 724604],
        ["", ""],
        ["--- LIABILITIES B/F ---", ""],
        ["JB Opening (B/F)", 17100],
        ["KK Opening (B/F)", 10100],
        ["Zakat Opening (B/F)", 73111],
        ["Commission Opening (B/F)", 358808],
        ["", ""],
        ["--- ASSETS B/F ---", ""],
        ["Mandi Exp Opening (B/F)", 17450],
        ["BF Discount Opening (B/F)", 24816],
        ["MHN Personal Opening (B/F)", 294265],
    ]
    for r, (lbl, val) in enumerate(settings, 2):
        ws.cell(row=r, column=24, value=lbl)
        if lbl and not lbl.startswith("---"):
            ws.cell(row=r, column=24).font = Font(bold=True)
        elif lbl.startswith("---"):
            ws.cell(row=r, column=24).font = Font(bold=True, color="1565C0")
        c = ws.cell(row=r, column=25, value=val if val != "" else "")
        if val != "":
            c.fill = INPUT_FILL
            c.number_format = MONEY_FMT if isinstance(val, (int, float)) and val > 100 else NUM_FMT
    
    for c, w in {1:4, 2:22, 3:8, 4:12, 5:14, 6:10, 7:10, 8:2, 9:4, 10:22, 11:14, 12:10, 13:10, 14:2, 15:4, 16:26, 17:14, 18:10, 19:2, 20:4, 21:20, 22:12, 23:2, 24:28, 25:14}.items():
        col_w(ws, c, w)
    
    # ========== DAILY_SALES ==========
    ws = wb.create_sheet("Daily_Sales", 1)
    ws.sheet_properties.tabColor = "4CAF50"
    ws.cell(row=1, column=1, value="DAILY SALES (Bepaari to Dukandar)").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Date", "Bepaari", "Dukandar", "Qty", "Rate", "Gross", "Discount", "Net"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 9)
    bep_dv = DataValidation(type="list", formula1='BEPAARI', allow_blank=True)
    duk_dv = DataValidation(type="list", formula1='DUKANDAR', allow_blank=True)
    ws.add_data_validation(bep_dv)
    ws.add_data_validation(duk_dv)
    for r in range(4, 504):
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3).fill = INPUT_FILL
        bep_dv.add(f'C{r}')
        ws.cell(row=r, column=4).fill = INPUT_FILL
        duk_dv.add(f'D{r}')
        ws.cell(row=r, column=5).fill = INPUT_FILL
        ws.cell(row=r, column=6).fill = INPUT_FILL
        ws.cell(row=r, column=6).number_format = MONEY_FMT
        ws.cell(row=r, column=7, value=f'=IF(E{r}="","",E{r}*F{r})')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=7).fill = CALC_FILL
        ws.cell(row=r, column=8).fill = INPUT_FILL
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        ws.cell(row=r, column=9, value=f'=IF(G{r}="","",G{r}-IF(H{r}="",0,H{r}))')
        ws.cell(row=r, column=9).number_format = MONEY_FMT
        ws.cell(row=r, column=9).fill = CALC_FILL
    for c, w in enumerate([4, 11, 22, 22, 6, 10, 12, 10, 12], 1): col_w(ws, c, w)
    
    # ========== CASH_BOOK - SIMPLE like V9 ==========
    ws = wb.create_sheet("Cash_Book", 2)
    ws.sheet_properties.tabColor = "FF5722"
    ws.cell(row=1, column=1, value="CASH BOOK (All Transactions)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Select TYPE first → Party dropdown shows relevant list").font = Font(italic=True, size=9)
    
    for c, h in enumerate(["Sr", "Date", "Type", "Sub-Type", "Party", "Particulars", "Amount", "Mode"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 8)
    
    # Simple Type dropdown
    main_dv = DataValidation(type="list", formula1='"BEPAARI,DUKANDAR,CAPITAL,LOAN,AMANAT,ADVANCE,EXPENSE,ZAKAT,OTHER"', allow_blank=True)
    ws.add_data_validation(main_dv)
    
    # Simple Sub-Type dropdown (all options)
    sub_dv = DataValidation(type="list", formula1='"PAYMENT,RECEIPT,MOTOR,BHUSSA,GAWALI,CASH_ADV,DISCOUNT,TAKEN,WITHDRAWN,REPAID,GIVEN,RECEIVED,MANDI,TRAVEL,FOOD,SALARY,BF_DISC,JB_PAID,MHN_PERSONAL,PROVISION,PAID,MISC,OTHER"', allow_blank=True)
    ws.add_data_validation(sub_dv)
    
    # Mode dropdown
    mode_dv = DataValidation(type="list", formula1='"CASH,BANK,UPI,TRANSFER"', allow_blank=True)
    ws.add_data_validation(mode_dv)
    
    for r in range(4, 1004):
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3).fill = INPUT_FILL
        main_dv.add(f'C{r}')
        ws.cell(row=r, column=4).fill = INPUT_FILL
        sub_dv.add(f'D{r}')
        # Dynamic party dropdown based on Type
        ws.cell(row=r, column=5).fill = INPUT_FILL
        party_dv = DataValidation(type="list", formula1=f'INDIRECT($C{r})', allow_blank=True)
        ws.add_data_validation(party_dv)
        party_dv.add(f'E{r}')
        ws.cell(row=r, column=6).fill = INPUT_FILL
        ws.cell(row=r, column=7).fill = INPUT_FILL
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=8).fill = INPUT_FILL
        mode_dv.add(f'H{r}')
    
    # Cash/Bank Summary - using specific ranges
    ws.cell(row=1, column=10, value="CASH/BANK SUMMARY").font = SUBTITLE_FONT
    ws.cell(row=2, column=9, value="Opening Cash")
    ws.cell(row=2, column=10, value="=Masters!$Y$6").number_format = MONEY_FMT
    ws.cell(row=3, column=9, value="Opening Bank")
    ws.cell(row=3, column=10, value="=Masters!$Y$7").number_format = MONEY_FMT
    
    # Cash OUT
    out_types = ["PAYMENT", "MOTOR", "BHUSSA", "GAWALI", "CASH_ADV", "DISCOUNT", "WITHDRAWN", "REPAID", "GIVEN", "MANDI", "TRAVEL", "FOOD", "SALARY", "BF_DISC", "JB_PAID", "MHN_PERSONAL", "PAID", "MISC", "OTHER"]
    ws.cell(row=5, column=9, value="Cash OUT")
    cash_out = '+'.join([f'SUMIFS($G$4:$G$1003,$H$4:$H$1003,"CASH",$D$4:$D$1003,"{st}")' for st in out_types])
    ws.cell(row=5, column=10, value=f'={cash_out}').number_format = MONEY_FMT
    
    # Bank OUT
    ws.cell(row=6, column=9, value="Bank OUT")
    bank_out_parts = []
    for mode in ["BANK", "UPI", "TRANSFER"]:
        for st in out_types:
            bank_out_parts.append(f'SUMIFS($G$4:$G$1003,$H$4:$H$1003,"{mode}",$D$4:$D$1003,"{st}")')
    ws.cell(row=6, column=10, value='=' + '+'.join(bank_out_parts)).number_format = MONEY_FMT
    
    # Cash IN
    in_types = ["RECEIPT", "TAKEN", "RECEIVED", "PROVISION"]
    ws.cell(row=7, column=9, value="Cash IN")
    cash_in = '+'.join([f'SUMIFS($G$4:$G$1003,$H$4:$H$1003,"CASH",$D$4:$D$1003,"{st}")' for st in in_types])
    ws.cell(row=7, column=10, value=f'={cash_in}').number_format = MONEY_FMT
    
    # Bank IN
    ws.cell(row=8, column=9, value="Bank IN")
    bank_in_parts = []
    for mode in ["BANK", "UPI", "TRANSFER"]:
        for st in in_types:
            bank_in_parts.append(f'SUMIFS($G$4:$G$1003,$H$4:$H$1003,"{mode}",$D$4:$D$1003,"{st}")')
    ws.cell(row=8, column=10, value='=' + '+'.join(bank_in_parts)).number_format = MONEY_FMT
    
    ws.cell(row=10, column=9, value="CLOSING CASH").font = Font(bold=True)
    ws.cell(row=10, column=10, value='=J2-J5+J7').number_format = MONEY_FMT
    ws.cell(row=10, column=10).font = Font(bold=True)
    ws.cell(row=10, column=10).fill = CALC_FILL
    
    ws.cell(row=11, column=9, value="CLOSING BANK").font = Font(bold=True)
    ws.cell(row=11, column=10, value='=J3-J6+J8').number_format = MONEY_FMT
    ws.cell(row=11, column=10).font = Font(bold=True)
    ws.cell(row=11, column=10).fill = CALC_FILL
    
    for c, w in enumerate([4, 11, 12, 16, 24, 20, 12, 10, 2, 14, 14], 1): col_w(ws, c, w)
    
    # ========== BEPAARI_LEDGER ==========
    ws = wb.create_sheet("Bepaari_Ledger", 3)
    ws.sheet_properties.tabColor = "9C27B0"
    ws.cell(row=1, column=1, value="BEPAARI LEDGER").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Commission = Flat ₹/Goat if set, otherwise Comm%").font = Font(italic=True, size=9)
    
    hdrs = ["Sr", "Bepaari", "Comm%", "Flat₹", "Opening", "Gross Sales", "Qty", "Commission", "KK", "JB", "Motor", "Bhussa", "Gawali", "Cash/Adv", "Other", "Tot.Deduct", "Net Payable", "Payments", "BALANCE"]
    for c, h in enumerate(hdrs, 1): ws.cell(row=4, column=c, value=h)
    hdr(ws, 4, 1, 19)
    
    for r in range(5, 55):
        m = r - 2
        ws.cell(row=r, column=1, value=r-4)
        ws.cell(row=r, column=2, value=f'=IF(Masters!$B${m}="","",Masters!$B${m})')
        ws.cell(row=r, column=3, value=f'=IF(B{r}="","",Masters!$C${m})')
        ws.cell(row=r, column=4, value=f'=IF(B{r}="","",IF(Masters!$D${m}="","",Masters!$D${m}))')
        ws.cell(row=r, column=4).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f'=IF(B{r}="",0,Masters!$E${m})').number_format = MONEY_FMT
        ws.cell(row=r, column=6, value=f'=SUMIF(Daily_Sales!$C$4:$C$503,B{r},Daily_Sales!$G$4:$G$503)').number_format = MONEY_FMT
        ws.cell(row=r, column=7, value=f'=SUMIF(Daily_Sales!$C$4:$C$503,B{r},Daily_Sales!$E$4:$E$503)').number_format = NUM_FMT
        # Commission: Flat if set, else %
        ws.cell(row=r, column=8, value=f'=IF(B{r}="",0,IF(D{r}<>"",D{r}*G{r},F{r}*C{r}/100))').number_format = MONEY_FMT
        ws.cell(row=r, column=9, value=f'=IF(G{r}>0,Masters!$Y$3,0)').number_format = MONEY_FMT
        ws.cell(row=r, column=9).fill = CALC_FILL
        ws.cell(row=r, column=10, value=f'=G{r}*Masters!$Y$4').number_format = MONEY_FMT
        ws.cell(row=r, column=10).fill = CALC_FILL
        for col, sub in [(11,"MOTOR"), (12,"BHUSSA"), (13,"GAWALI"), (14,"CASH_ADV"), (15,"OTHER")]:
            ws.cell(row=r, column=col, value=f'=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,"BEPAARI",Cash_Book!$D$4:$D$1003,"{sub}",Cash_Book!$E$4:$E$1003,B{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=16, value=f'=SUM(H{r}:O{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=16).fill = CALC_FILL
        ws.cell(row=r, column=17, value=f'=IF(B{r}="","",F{r}-P{r}+E{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=17).fill = CALC_FILL
        ws.cell(row=r, column=18, value=f'=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,"BEPAARI",Cash_Book!$D$4:$D$1003,"PAYMENT",Cash_Book!$E$4:$E$1003,B{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=19, value=f'=IF(B{r}="","",Q{r}-R{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=19).font = Font(bold=True)
    
    ws.cell(row=55, column=1, value="TOTAL").font = Font(bold=True)
    for c in [5,6,7,8,9,10,11,12,13,14,15,16,17,18,19]:
        ws.cell(row=55, column=c, value=f'=SUM({get_column_letter(c)}5:{get_column_letter(c)}54)').number_format = MONEY_FMT
        ws.cell(row=55, column=c).font = Font(bold=True)
    
    # Balance Sheet helpers
    ws.cell(row=57, column=16, value="For Balance Sheet:").font = Font(bold=True, italic=True)
    ws.cell(row=58, column=17, value="Bepaari Payable (we owe)")
    ws.cell(row=58, column=19, value='=SUMIF(S5:S54,">0",S5:S54)').number_format = MONEY_FMT
    ws.cell(row=58, column=19).fill = CALC_FILL
    ws.cell(row=59, column=17, value="Bepaari Overpaid (they owe)")
    ws.cell(row=59, column=19, value='=-SUMIF(S5:S54,"<0",S5:S54)').number_format = MONEY_FMT
    ws.cell(row=59, column=19).fill = CALC_FILL
    
    for c, w in enumerate([4,18,6,8,10,12,6,10,8,8,9,9,9,10,8,11,12,11,12], 1): col_w(ws, c, w)
    
    # ========== DUKANDAR_LEDGER ==========
    ws = wb.create_sheet("Dukandar_Ledger", 4)
    ws.sheet_properties.tabColor = "E91E63"
    ws.cell(row=1, column=1, value="DUKANDAR LEDGER").font = TITLE_FONT
    
    for c, h in enumerate(["Sr", "Dukandar", "Opening", "Purchases", "Discounts", "Net Receivable", "Receipts", "BALANCE"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 8)
    
    for r in range(4, 54):
        m = r - 1
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2, value=f'=IF(Masters!$J${m}="","",Masters!$J${m})')
        ws.cell(row=r, column=3, value=f'=IF(B{r}="",0,Masters!$K${m})').number_format = MONEY_FMT
        ws.cell(row=r, column=4, value=f'=SUMIF(Daily_Sales!$D$4:$D$503,B{r},Daily_Sales!$G$4:$G$503)').number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f'=SUMIF(Daily_Sales!$D$4:$D$503,B{r},Daily_Sales!$H$4:$H$503)').number_format = MONEY_FMT
        ws.cell(row=r, column=6, value=f'=IF(B{r}="","",D{r}-E{r}+C{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=6).fill = CALC_FILL
        ws.cell(row=r, column=7, value=f'=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,"DUKANDAR",Cash_Book!$D$4:$D$1003,"RECEIPT",Cash_Book!$E$4:$E$1003,B{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=8, value=f'=IF(B{r}="","",F{r}-G{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=8).font = Font(bold=True)
    
    ws.cell(row=54, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=54, column=2, value="PATTI")
    for c in [3,4,5,6,7,8]:
        ws.cell(row=54, column=c, value=f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}53)').number_format = MONEY_FMT
        ws.cell(row=54, column=c).font = Font(bold=True)
    
    # Balance Sheet helpers
    ws.cell(row=56, column=5, value="For Balance Sheet:").font = Font(bold=True, italic=True)
    ws.cell(row=57, column=6, value="Dukandar Receivable (they owe)")
    ws.cell(row=57, column=8, value='=SUMIF(H4:H53,">0",H4:H53)').number_format = MONEY_FMT
    ws.cell(row=57, column=8).fill = CALC_FILL
    ws.cell(row=58, column=6, value="Dukandar Overpaid (we owe)")
    ws.cell(row=58, column=8, value='=-SUMIF(H4:H53,"<0",H4:H53)').number_format = MONEY_FMT
    ws.cell(row=58, column=8).fill = CALC_FILL
    
    for c, w in enumerate([4,22,12,14,12,14,12,14], 1): col_w(ws, c, w)
    
    # ========== CAPITAL_LOAN_LEDGER ==========
    ws = wb.create_sheet("Capital_Loan_Ledger", 5)
    ws.sheet_properties.tabColor = "6A1B9A"
    ws.cell(row=1, column=1, value="CAPITAL & LOAN LEDGER").font = TITLE_FONT
    
    for c, h in enumerate(["Sr", "Party Name", "Type", "Opening Bal", "Taken (+)", "Withdrawn/Repaid (-)", "CURRENT BALANCE"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 7, HEADER_PURPLE)
    
    for r in range(4, 34):
        m = r - 1
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2, value=f'=IF(Masters!$P${m}="","",Masters!$P${m})')
        ws.cell(row=r, column=3, value=f'=IF(B{r}="","",Masters!$R${m})')
        ws.cell(row=r, column=4, value=f'=IF(B{r}="",0,Masters!$Q${m})').number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f'=IF(B{r}="",0,SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,C{r},Cash_Book!$D$4:$D$1003,"TAKEN",Cash_Book!$E$4:$E$1003,B{r}))').number_format = MONEY_FMT
        ws.cell(row=r, column=6, value=f'=IF(B{r}="",0,SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,C{r},Cash_Book!$D$4:$D$1003,"WITHDRAWN",Cash_Book!$E$4:$E$1003,B{r})+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,C{r},Cash_Book!$D$4:$D$1003,"REPAID",Cash_Book!$E$4:$E$1003,B{r}))').number_format = MONEY_FMT
        ws.cell(row=r, column=7, value=f'=IF(B{r}="","",D{r}+E{r}-F{r})').number_format = MONEY_FMT
        ws.cell(row=r, column=7).font = Font(bold=True)
        ws.cell(row=r, column=7).fill = CALC_FILL
    
    # Totals
    ws.cell(row=35, column=1, value="TOTALS BY TYPE:").font = Font(bold=True)
    ws.cell(row=36, column=2, value="CAPITAL Total")
    ws.cell(row=36, column=7, value='=SUMIF(C4:C33,"CAPITAL",G4:G33)').number_format = MONEY_FMT
    ws.cell(row=36, column=7).font = Font(bold=True)
    ws.cell(row=36, column=7).fill = CALC_FILL
    
    ws.cell(row=37, column=2, value="LOAN Total")
    ws.cell(row=37, column=7, value='=SUMIF(C4:C33,"LOAN",G4:G33)').number_format = MONEY_FMT
    ws.cell(row=37, column=7).font = Font(bold=True)
    ws.cell(row=37, column=7).fill = CALC_FILL
    
    ws.cell(row=38, column=2, value="AMANAT Total")
    ws.cell(row=38, column=7, value='=SUMIF(C4:C33,"AMANAT",G4:G33)').number_format = MONEY_FMT
    ws.cell(row=38, column=7).font = Font(bold=True)
    ws.cell(row=38, column=7).fill = CALC_FILL
    
    for c, w in enumerate([4, 26, 10, 14, 14, 18, 16], 1): col_w(ws, c, w)
    
    # ========== BEPAARI_AAKDA ==========
    ws = wb.create_sheet("Bepaari_Aakda", 6)
    ws.sheet_properties.tabColor = "795548"
    ws.cell(row=1, column=1, value="BEPAARI SETTLEMENT SLIP (AAKDA)").font = TITLE_FONT
    ws.cell(row=3, column=1, value="SELECT BEPAARI:")
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ak_dv = DataValidation(type="list", formula1='BEPAARI', allow_blank=False)
    ws.add_data_validation(ak_dv)
    ak_dv.add('B3')
    
    # Row 5 onwards
    aakda = [
        ["", ""],
        ["A. SALES", ""],
        ["Qty Sold", '=SUMIF(Daily_Sales!$C$4:$C$503,$B$3,Daily_Sales!$E$4:$E$503)'],
        ["Gross Sales", '=SUMIF(Daily_Sales!$C$4:$C$503,$B$3,Daily_Sales!$G$4:$G$503)'],
        ["", ""],
        ["B. DEDUCTIONS", ""],
        ["Commission Type", '=IF(IFERROR(INDEX(Masters!$D$3:$D$52,MATCH($B$3,Masters!$B$3:$B$52,0)),"")="","% Based","Flat ₹/Goat")'],
        ["Commission %", '=IFERROR(INDEX(Masters!$C$3:$C$52,MATCH($B$3,Masters!$B$3:$B$52,0)),4)'],
        ["Flat ₹/Goat", '=IFERROR(INDEX(Masters!$D$3:$D$52,MATCH($B$3,Masters!$B$3:$B$52,0)),"")'],
        ["Commission Amt", '=IF(B13<>"",B13*B7,B8*B12/100)'],
        ["KK (Fixed)", '=IF(B7>0,Masters!$Y$3,0)'],
        ["JB", '=B7*Masters!$Y$4'],
        ["Motor", '=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,"BEPAARI",Cash_Book!$D$4:$D$1003,"MOTOR",Cash_Book!$E$4:$E$1003,$B$3)'],
        ["Bhussa", '=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,"BEPAARI",Cash_Book!$D$4:$D$1003,"BHUSSA",Cash_Book!$E$4:$E$1003,$B$3)'],
        ["Gawali", '=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,"BEPAARI",Cash_Book!$D$4:$D$1003,"GAWALI",Cash_Book!$E$4:$E$1003,$B$3)'],
        ["Cash/Adv", '=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,"BEPAARI",Cash_Book!$D$4:$D$1003,"CASH_ADV",Cash_Book!$E$4:$E$1003,$B$3)'],
        ["TOTAL DEDUCTIONS", '=SUM(B14:B20)'],
        ["", ""],
        ["C. CALCULATION", ""],
        ["Gross Sales", '=B8'],
        ["(-) Deductions", '=B21'],
        ["(+) Opening Bal", '=IFERROR(INDEX(Masters!$E$3:$E$52,MATCH($B$3,Masters!$B$3:$B$52,0)),0)'],
        ["NET PAYABLE", '=B24-B25+B26'],
        ["", ""],
        ["D. PAYMENTS", ""],
        ["Payments Made", '=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$C$4:$C$1003,"BEPAARI",Cash_Book!$D$4:$D$1003,"PAYMENT",Cash_Book!$E$4:$E$1003,$B$3)'],
        ["", ""],
        ["BALANCE DUE", '=B27-B30']
    ]
    for r, (lbl, frm) in enumerate(aakda, 5):
        ws.cell(row=r, column=1, value=lbl)
        if frm:
            c = ws.cell(row=r, column=2, value=frm)
            c.number_format = MONEY_FMT if "Qty" not in lbl and "%" not in lbl and "Type" not in lbl else NUM_FMT
        if any(x in lbl for x in ["A.", "B.", "C.", "D.", "TOTAL", "NET", "BALANCE"]):
            ws.cell(row=r, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl or "BALANCE" in lbl:
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = CALC_FILL
    col_w(ws, 1, 22)
    col_w(ws, 2, 14)
    
    # ========== BALANCE_SHEET - SIMPLE with B/F ==========
    ws = wb.create_sheet("Balance_Sheet", 7)
    ws.sheet_properties.tabColor = "607D8B"
    ws.cell(row=1, column=1, value="BALANCE SHEET").font = TITLE_FONT
    ws.cell(row=1, column=5, value="LIABILITIES = ASSETS").font = Font(italic=True, size=10, color="FF0000")
    
    # ===== LIABILITIES =====
    ws.cell(row=3, column=1, value="LIABILITIES (We Owe)").font = SUBTITLE_FONT
    ws.cell(row=3, column=1).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    ws.cell(row=4, column=1, value="Particulars")
    ws.cell(row=4, column=2, value="Amount")
    hdr(ws, 4, 1, 2, HEADER_RED)
    
    row = 5
    ws.cell(row=row, column=1, value="CAPITAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Capital_Loan_Ledger!G36').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    row += 1
    
    ws.cell(row=row, column=1, value="LOANS").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Capital_Loan_Ledger!G37').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    row += 1
    
    ws.cell(row=row, column=1, value="AMANAT").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Capital_Loan_Ledger!G38').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    row += 2
    
    ws.cell(row=row, column=1, value="BEPAARI PAYABLES").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Bepaari_Ledger!S58').number_format = MONEY_FMT
    row += 1
    
    ws.cell(row=row, column=1, value="DUKANDAR ADVANCES (Overpaid)").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=Dukandar_Ledger!H58').number_format = MONEY_FMT
    row += 2
    
    # JB with B/F
    ws.cell(row=row, column=1, value="JB:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="  B/F")
    ws.cell(row=row, column=2, value='=Masters!$Y$10').number_format = MONEY_FMT
    jb_bf = row
    row += 1
    ws.cell(row=row, column=1, value="  (+) Collected")
    ws.cell(row=row, column=2, value='=Bepaari_Ledger!J55').number_format = MONEY_FMT
    jb_coll = row
    row += 1
    ws.cell(row=row, column=1, value="  (-) Paid")
    ws.cell(row=row, column=2, value='=-SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"JB_PAID")').number_format = MONEY_FMT
    jb_paid = row
    row += 1
    ws.cell(row=row, column=1, value="  JB Total").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f'=B{jb_bf}+B{jb_coll}+B{jb_paid}').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    jb_total = row
    row += 2
    
    # KK with B/F
    ws.cell(row=row, column=1, value="KK:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="  B/F")
    ws.cell(row=row, column=2, value='=Masters!$Y$11').number_format = MONEY_FMT
    kk_bf = row
    row += 1
    ws.cell(row=row, column=1, value="  (+) Collected")
    ws.cell(row=row, column=2, value='=Bepaari_Ledger!I55').number_format = MONEY_FMT
    kk_coll = row
    row += 1
    ws.cell(row=row, column=1, value="  KK Total").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f'=B{kk_bf}+B{kk_coll}').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    kk_total = row
    row += 2
    
    # Commission with B/F
    ws.cell(row=row, column=1, value="COMMISSION:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="  B/F")
    ws.cell(row=row, column=2, value='=Masters!$Y$13').number_format = MONEY_FMT
    comm_bf = row
    row += 1
    ws.cell(row=row, column=1, value="  (+) Earned")
    ws.cell(row=row, column=2, value='=Bepaari_Ledger!H55').number_format = MONEY_FMT
    comm_earned = row
    row += 1
    ws.cell(row=row, column=1, value="  (-) Discounts")
    ws.cell(row=row, column=2, value='=-Dukandar_Ledger!E54').number_format = MONEY_FMT
    comm_disc = row
    row += 1
    ws.cell(row=row, column=1, value="  Commission Total").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f'=B{comm_bf}+B{comm_earned}+B{comm_disc}').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    comm_total = row
    row += 2
    
    # Zakat with B/F
    ws.cell(row=row, column=1, value="ZAKAT PAYABLE:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="  B/F + Provision - Paid")
    ws.cell(row=row, column=2, value='=Masters!$Y$12+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"PROVISION")-SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"PAID")').number_format = MONEY_FMT
    ws.cell(row=row, column=2).fill = CALC_FILL
    zakat_row = row
    row += 2
    
    total_liab_row = row
    ws.cell(row=row, column=1, value="TOTAL LIABILITIES").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=f'=B5+B6+B7+B9+B10+B{jb_total}+B{kk_total}+B{comm_total}+B{zakat_row}').number_format = MONEY_FMT
    ws.cell(row=row, column=2).font = Font(bold=True, size=12)
    ws.cell(row=row, column=2).fill = CALC_FILL
    
    # ===== ASSETS =====
    ws.cell(row=3, column=4, value="ASSETS (We Have / Owed to Us)").font = SUBTITLE_FONT
    ws.cell(row=3, column=4).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    ws.cell(row=4, column=4, value="Particulars")
    ws.cell(row=4, column=5, value="Amount")
    hdr(ws, 4, 4, 5, HEADER_GREEN)
    
    arow = 5
    ws.cell(row=arow, column=4, value="CASH BALANCE").font = Font(bold=True)
    ws.cell(row=arow, column=5, value='=Cash_Book!J10').number_format = MONEY_FMT
    ws.cell(row=arow, column=5).fill = CALC_FILL
    cash_row = arow
    arow += 1
    
    ws.cell(row=arow, column=4, value="BANK BALANCE").font = Font(bold=True)
    ws.cell(row=arow, column=5, value='=Cash_Book!J11').number_format = MONEY_FMT
    ws.cell(row=arow, column=5).fill = CALC_FILL
    bank_row = arow
    arow += 2
    
    ws.cell(row=arow, column=4, value="PATTI (Dukandar Receivable)").font = Font(bold=True)
    ws.cell(row=arow, column=5, value='=Dukandar_Ledger!H57').number_format = MONEY_FMT
    patti_row = arow
    arow += 1
    
    ws.cell(row=arow, column=4, value="BEPAARI ADVANCES (Overpaid)").font = Font(bold=True)
    ws.cell(row=arow, column=5, value='=Bepaari_Ledger!S59').number_format = MONEY_FMT
    bep_adv_row = arow
    arow += 2
    
    ws.cell(row=arow, column=4, value="ADVANCES GIVEN (Net)").font = Font(bold=True)
    ws.cell(row=arow, column=5, value='=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"GIVEN")-SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"RECEIVED")').number_format = MONEY_FMT
    adv_row = arow
    arow += 2
    
    # Mandi Exp with B/F
    ws.cell(row=arow, column=4, value="MANDI EXPENSES:").font = Font(bold=True)
    arow += 1
    ws.cell(row=arow, column=4, value="  B/F")
    ws.cell(row=arow, column=5, value='=Masters!$Y$16').number_format = MONEY_FMT
    mandi_bf = arow
    arow += 1
    ws.cell(row=arow, column=4, value="  (+) Current")
    ws.cell(row=arow, column=5, value='=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"MANDI")+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"TRAVEL")+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"FOOD")+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"SALARY")+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"MISC")+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"OTHER")').number_format = MONEY_FMT
    mandi_curr = arow
    arow += 1
    ws.cell(row=arow, column=4, value="  Mandi Total").font = Font(bold=True)
    ws.cell(row=arow, column=5, value=f'=E{mandi_bf}+E{mandi_curr}').number_format = MONEY_FMT
    ws.cell(row=arow, column=5).fill = CALC_FILL
    mandi_total = arow
    arow += 2
    
    # BF Disc with B/F
    ws.cell(row=arow, column=4, value="BF DISCOUNT:").font = Font(bold=True)
    arow += 1
    ws.cell(row=arow, column=4, value="  B/F")
    ws.cell(row=arow, column=5, value='=Masters!$Y$17').number_format = MONEY_FMT
    bf_bf = arow
    arow += 1
    ws.cell(row=arow, column=4, value="  (+) Current")
    ws.cell(row=arow, column=5, value='=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"BF_DISC")').number_format = MONEY_FMT
    bf_curr = arow
    arow += 1
    ws.cell(row=arow, column=4, value="  BF Disc Total").font = Font(bold=True)
    ws.cell(row=arow, column=5, value=f'=E{bf_bf}+E{bf_curr}').number_format = MONEY_FMT
    ws.cell(row=arow, column=5).fill = CALC_FILL
    bf_total = arow
    arow += 2
    
    # MHN Personal with B/F
    ws.cell(row=arow, column=4, value="MHN PERSONAL:").font = Font(bold=True)
    arow += 1
    ws.cell(row=arow, column=4, value="  B/F")
    ws.cell(row=arow, column=5, value='=Masters!$Y$18').number_format = MONEY_FMT
    mhn_bf = arow
    arow += 1
    ws.cell(row=arow, column=4, value="  (+) Paid")
    ws.cell(row=arow, column=5, value='=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"MHN_PERSONAL")').number_format = MONEY_FMT
    mhn_curr = arow
    arow += 1
    ws.cell(row=arow, column=4, value="  MHN Total").font = Font(bold=True)
    ws.cell(row=arow, column=5, value=f'=E{mhn_bf}+E{mhn_curr}').number_format = MONEY_FMT
    ws.cell(row=arow, column=5).fill = CALC_FILL
    mhn_total = arow
    
    # Total Assets
    ws.cell(row=total_liab_row, column=4, value="TOTAL ASSETS").font = Font(bold=True, size=12)
    ws.cell(row=total_liab_row, column=5, value=f'=E{cash_row}+E{bank_row}+E{patti_row}+E{bep_adv_row}+E{adv_row}+E{mandi_total}+E{bf_total}+E{mhn_total}').number_format = MONEY_FMT
    ws.cell(row=total_liab_row, column=5).font = Font(bold=True, size=12)
    ws.cell(row=total_liab_row, column=5).fill = CALC_FILL
    
    # Difference
    ws.cell(row=total_liab_row+2, column=1, value="DIFFERENCE (Should be 0)").font = Font(bold=True, color="FF0000")
    ws.cell(row=total_liab_row+2, column=2, value=f'=B{total_liab_row}-E{total_liab_row}').number_format = MONEY_FMT
    ws.cell(row=total_liab_row+2, column=2).font = Font(bold=True, color="FF0000")
    
    col_w(ws, 1, 28)
    col_w(ws, 2, 14)
    col_w(ws, 3, 3)
    col_w(ws, 4, 28)
    col_w(ws, 5, 14)
    
    # ========== COMMISSION_SUMMARY ==========
    ws = wb.create_sheet("Commission_Summary", 8)
    ws.sheet_properties.tabColor = "FFC107"
    ws.cell(row=1, column=1, value="COMMISSION & PROFIT SUMMARY").font = TITLE_FONT
    
    items = [
        ["", ""],
        ["GROSS SALES", "=Bepaari_Ledger!F55"],
        ["TOTAL GOATS SOLD", "=Bepaari_Ledger!G55"],
        ["", ""],
        ["INCOME:", ""],
        ["  Commission Earned", "=Bepaari_Ledger!H55"],
        ["  (-) Discounts Given", "=-Dukandar_Ledger!E54"],
        ["  Net Commission", "=B6+B7"],
        ["  JB Collected", "=Bepaari_Ledger!J55"],
        ["  KK Collected", "=Bepaari_Ledger!I55"],
        ["TOTAL INCOME", "=B8+B9+B10"],
        ["", ""],
        ["EXPENSES:", ""],
        ["  Mandi/Business", '=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"MANDI")+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"TRAVEL")+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"FOOD")+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"SALARY")+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"MISC")+SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"OTHER")'],
        ["  BF Discount", '=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"BF_DISC")'],
        ["  JB Paid", '=SUMIFS(Cash_Book!$G$4:$G$1003,Cash_Book!$D$4:$D$1003,"JB_PAID")'],
        ["TOTAL EXPENSES", "=B14+B15+B16"],
        ["", ""],
        ["NET PROFIT/LOSS", "=B11-B17"]
    ]
    for r, (lbl, frm) in enumerate(items, 2):
        ws.cell(row=r, column=1, value=lbl)
        if frm:
            c = ws.cell(row=r, column=2, value=frm)
            c.number_format = MONEY_FMT if "GOATS" not in lbl else NUM_FMT
        if any(x in lbl for x in ["GROSS", "TOTAL", "NET", "INCOME:", "EXPENSES:"]):
            ws.cell(row=r, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl:
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = CALC_FILL
    col_w(ws, 1, 26)
    col_w(ws, 2, 14)
    
    wb.save(path)
    print(f"Mandi Master V9.1 created: {path}")

if __name__ == "__main__":
    create_v91("/app/frontend/public/Mandi_Master_V9.1.xlsx")
