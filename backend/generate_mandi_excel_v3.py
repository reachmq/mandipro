"""
Mandi Master Excel Generator - VERSION 3.0
Complete redesign based on user feedback and actual paper-based accounting
FIXES: Dropdowns, Cash Book, Mode of Payment, Auto KK/JB, Balance Sheet
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

# Style definitions
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_RED = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
HEADER_GREEN = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="0D47A1")
SUBTITLE_FONT = Font(bold=True, size=11, color="1565C0")
INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
CALC_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
MONEY_FORMAT = '₹#,##0'
NUMBER_FORMAT = '#,##0'
DATE_FORMAT = 'DD-MMM-YYYY'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def style_header(ws, row, start_col, end_col, fill=None):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def create_mandi_excel_v3(output_path):
    wb = Workbook()
    wb.remove(wb.active)
    
    # ================================================================
    # SHEET 1: MASTERS
    # ================================================================
    ws_m = wb.create_sheet("Masters", 0)
    ws_m.sheet_properties.tabColor = "1565C0"
    
    # ----- BEPAARI MASTER (A-F) -----
    ws_m.cell(row=1, column=1, value="BEPAARI MASTER").font = TITLE_FONT
    bep_headers = ["Sr.", "Bepaari Name", "Comm.Type", "Comm.Value", "Opening Bal.", "Remarks"]
    for col, h in enumerate(bep_headers, 1):
        ws_m.cell(row=2, column=col, value=h)
    style_header(ws_m, 2, 1, 6)
    
    # Sample data from user's file
    sample_bepaaris = [
        [1, "ABUL HASAN", "%", 4, -70210, ""],
        [2, "SHARAFAT", "%", 4, 0, ""],
        [3, "SHARIK", "%", 4, 0, ""],
        [4, "JUNAID", "%", 4, -121590, ""],
        [5, "ANNU SHARIF", "%", 4, 45286, ""],
        [6, "GAYYUR", "%", 4, -184866, ""],
        [7, "ARIF KK", "%", 4, 0, ""],
        [8, "KISHOR", "%", 4, -360084, ""],
        [9, "SHOEB BYCULLA", "%", 4, 0, ""],
        [10, "MAULANA", "%", 4, 13630, ""],
        [11, "ISLAM D", "%", 4, 41812, ""],
        [12, "SALIM RAJ", "%", 4, 1526, ""],
        [13, "MEHMOOD MADH", "%", 4, 1530, ""],
        [14, "ANAS RAJ", "%", 4, 25132, ""],
        [15, "MOBIN CHIPLUN", "%", 4, 0, ""],
        [16, "AYAZ", "%", 4, -14048, ""],
        [17, "JALIL", "%", 4, -4661, ""],
        [18, "NEERAJ", "%", 4, -9845, ""],
        [19, "IRFAN JOGESHWARI", "%", 4, -88000, ""],
        [20, "CHINTU SHUJALPUR", "%", 4, 0, ""],
    ]
    for row_idx, data in enumerate(sample_bepaaris, 3):
        for col_idx, val in enumerate(data, 1):
            cell = ws_m.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = INPUT_FILL
            if col_idx == 5:
                cell.number_format = MONEY_FORMAT
    
    # Empty rows for more
    for row_idx in range(23, 70):
        ws_m.cell(row=row_idx, column=1, value=row_idx - 2)
        for col in range(2, 7):
            ws_m.cell(row=row_idx, column=col).fill = INPUT_FILL
    
    # Commission type dropdown
    comm_dv = DataValidation(type="list", formula1='"%, Per Piece"', allow_blank=True)
    ws_m.add_data_validation(comm_dv)
    comm_dv.add('C3:C69')
    
    # ----- DUKANDAR MASTER (H-L) -----
    ws_m.cell(row=1, column=8, value="DUKANDAR MASTER").font = TITLE_FONT
    duk_headers = ["Sr.", "Dukandar Name", "Opening Bal.", "Phone", "Remarks"]
    for col, h in enumerate(duk_headers, 8):
        ws_m.cell(row=2, column=col, value=h)
    style_header(ws_m, 2, 8, 12)
    
    sample_dukandars = [
        [1, "SAKIM", 143000, "", ""],
        [2, "SHARIK", 362250, "", ""],
        [3, "YUNUS", 168543, "", ""],
        [4, "MUDASSIR", 515528, "", ""],
        [5, "ARIF", 36150, "", ""],
        [6, "SHOEB BYCULLA", 24000, "", ""],
        [7, "ISLAM D", 101200, "", ""],
        [8, "MEHMOOD MADH", 15000, "", ""],
        [9, "MOBIN CHIPLUN", 184500, "", ""],
        [10, "JALIL", 150950, "", ""],
        [11, "IRFAN JOGESHWARI", 0, "", ""],
        [12, "YUNUS DOMBIVLI", 59662, "", ""],
        [13, "IFTEKHAR PUNE", 47000, "", ""],
        [14, "LALA/ASIF", 83000, "", ""],
        [15, "JAFER KASHIMIRA", 147000, "", ""],
        [16, "CHAND", 0, "", ""],
        [17, "ASIF GOREGAON", 0, "", ""],
        [18, "CHINTU SHUJALPUR", 0, "", ""],  # Added as Dukandar too
    ]
    for row_idx, data in enumerate(sample_dukandars, 3):
        for col_idx, val in enumerate(data, 8):
            cell = ws_m.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = INPUT_FILL
            if col_idx == 10:
                cell.number_format = MONEY_FORMAT
    
    for row_idx in range(21, 70):
        ws_m.cell(row=row_idx, column=8, value=row_idx - 2)
        for col in range(9, 13):
            ws_m.cell(row=row_idx, column=col).fill = INPUT_FILL
    
    # ----- LOAN/AMANAT PARTIES (N-Q) -----
    ws_m.cell(row=1, column=14, value="LOAN/AMANAT PARTIES").font = TITLE_FONT
    loan_headers = ["Sr.", "Party Name", "Opening Bal.", "Type"]
    for col, h in enumerate(loan_headers, 14):
        ws_m.cell(row=2, column=col, value=h)
    style_header(ws_m, 2, 14, 17)
    
    sample_loans = [
        [1, "SHAKIL GHODEGAON", 500000, "AMANAT"],
    ]
    for row_idx, data in enumerate(sample_loans, 3):
        for col_idx, val in enumerate(data, 14):
            cell = ws_m.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = INPUT_FILL
            if col_idx == 16:
                cell.number_format = MONEY_FORMAT
    
    for row_idx in range(5, 20):
        ws_m.cell(row=row_idx, column=14, value=row_idx - 2)
        for col in range(15, 18):
            ws_m.cell(row=row_idx, column=col).fill = INPUT_FILL
    
    # ----- FIXED VALUES & BALANCES (S-T) -----
    ws_m.cell(row=1, column=19, value="SETTINGS & OPENING BALANCES").font = TITLE_FONT
    
    fixed_items = [
        ["COMMISSION RATE (%)", 4],
        ["KARKUNI (KK) Fixed ₹/Bepari", 100],
        ["JAGAH BHADA (JB) ₹/Goat", 10],
        ["", ""],
        ["OPENING CASH BALANCE", 0],
        ["OPENING BANK BALANCE", 0],
        ["CAPITAL ACCOUNT", 3819981],
    ]
    for row_idx, (label, val) in enumerate(fixed_items, 2):
        ws_m.cell(row=row_idx, column=19, value=label)
        if label:
            ws_m.cell(row=row_idx, column=19).font = Font(bold=True)
        cell = ws_m.cell(row=row_idx, column=20, value=val)
        cell.fill = INPUT_FILL
        if "BALANCE" in str(label) or "CAPITAL" in str(label):
            cell.number_format = MONEY_FORMAT
    
    # Column widths
    widths = {1:4, 2:20, 3:10, 4:10, 5:14, 6:12, 7:2, 8:4, 9:20, 10:14, 11:10, 12:12, 13:2,
              14:4, 15:20, 16:14, 17:10, 18:2, 19:26, 20:14}
    for col, w in widths.items():
        set_col_width(ws_m, col, w)
    
    # ================================================================
    # SHEET 2: DAILY_SALES
    # ================================================================
    ws_s = wb.create_sheet("Daily_Sales", 1)
    ws_s.sheet_properties.tabColor = "4CAF50"
    
    ws_s.cell(row=1, column=1, value="DAILY SALES ENTRY (Bepaari → Dukandar)").font = TITLE_FONT
    ws_s.cell(row=2, column=1, value="Select names from DROPDOWN. Yellow = Enter, Green = Auto-calculated").font = Font(italic=True, size=9)
    
    sales_headers = ["Sr.", "Date", "Bepaari", "Dukandar", "Qty", "Rate", "Gross", "Discount", "Net", "Remarks"]
    for col, h in enumerate(sales_headers, 1):
        ws_s.cell(row=3, column=col, value=h)
    style_header(ws_s, 3, 1, 10)
    
    # DROPDOWN for Bepaari (from Masters B3:B69)
    bep_dv = DataValidation(type="list", formula1='Masters!$B$3:$B$69', allow_blank=True)
    bep_dv.error = "Select from Bepaari list"
    ws_s.add_data_validation(bep_dv)
    
    # DROPDOWN for Dukandar (from Masters I3:I69)
    duk_dv = DataValidation(type="list", formula1='Masters!$I$3:$I$69', allow_blank=True)
    duk_dv.error = "Select from Dukandar list"
    ws_s.add_data_validation(duk_dv)
    
    for row in range(4, 504):
        ws_s.cell(row=row, column=1, value=row-3)
        ws_s.cell(row=row, column=2).number_format = DATE_FORMAT
        ws_s.cell(row=row, column=2).fill = INPUT_FILL
        
        # Bepaari dropdown
        ws_s.cell(row=row, column=3).fill = INPUT_FILL
        bep_dv.add(f'C{row}')
        
        # Dukandar dropdown
        ws_s.cell(row=row, column=4).fill = INPUT_FILL
        duk_dv.add(f'D{row}')
        
        ws_s.cell(row=row, column=5).fill = INPUT_FILL  # Qty
        ws_s.cell(row=row, column=6).fill = INPUT_FILL  # Rate
        ws_s.cell(row=row, column=6).number_format = MONEY_FORMAT
        
        # Gross = Qty × Rate
        ws_s.cell(row=row, column=7, value=f'=IF(E{row}="","",E{row}*F{row})')
        ws_s.cell(row=row, column=7).number_format = MONEY_FORMAT
        ws_s.cell(row=row, column=7).fill = CALC_FILL
        
        ws_s.cell(row=row, column=8).fill = INPUT_FILL  # Discount
        ws_s.cell(row=row, column=8).number_format = MONEY_FORMAT
        
        # Net = Gross - Discount
        ws_s.cell(row=row, column=9, value=f'=IF(G{row}="","",G{row}-IF(H{row}="",0,H{row}))')
        ws_s.cell(row=row, column=9).number_format = MONEY_FORMAT
        ws_s.cell(row=row, column=9).fill = CALC_FILL
        
        ws_s.cell(row=row, column=10).fill = INPUT_FILL  # Remarks
    
    widths = [4, 11, 18, 18, 6, 10, 12, 10, 12, 18]
    for col, w in enumerate(widths, 1):
        set_col_width(ws_s, col, w)
    
    # ================================================================
    # SHEET 3: CASH_BOOK (Single sheet for ALL Cash/Bank transactions)
    # ================================================================
    ws_cb = wb.create_sheet("Cash_Book", 2)
    ws_cb.sheet_properties.tabColor = "FF5722"
    
    ws_cb.cell(row=1, column=1, value="DAILY CASH & BANK BOOK (All Transactions in One Place)").font = TITLE_FONT
    ws_cb.cell(row=2, column=1, value="Trans Types: BEP_PAY, DUK_REC, EXP_*, LOAN_*, ZAKAT, BF_DISC, JB_PAID, ADV_*, CAPITAL").font = Font(italic=True, size=9)
    
    cb_headers = ["Sr.", "Date", "Type", "Party/Head", "Particulars", "Mode", "Cash OUT", "Bank OUT", "Cash IN", "Bank IN"]
    for col, h in enumerate(cb_headers, 1):
        ws_cb.cell(row=3, column=col, value=h)
    style_header(ws_cb, 3, 1, 10)
    
    # Transaction type dropdown
    trans_list = "BEP_PAY,DUK_REC,EXP_MOTOR,EXP_CONV,EXP_FOOD,EXP_TRAVEL,EXP_SALARY,EXP_OTHER,LOAN_TAKEN,LOAN_REPAY,ADV_GIVEN,ADV_RECV,ZAKAT,BF_DISC,JB_PAID,CAPITAL"
    trans_dv = DataValidation(type="list", formula1=f'"{trans_list}"', allow_blank=True)
    ws_cb.add_data_validation(trans_dv)
    
    # Mode dropdown (CASH, BANK/UPI combined)
    mode_dv = DataValidation(type="list", formula1='"CASH,BANK/UPI"', allow_blank=True)
    ws_cb.add_data_validation(mode_dv)
    
    for row in range(4, 1004):
        ws_cb.cell(row=row, column=1, value=row-3)
        ws_cb.cell(row=row, column=2).number_format = DATE_FORMAT
        ws_cb.cell(row=row, column=2).fill = INPUT_FILL
        
        ws_cb.cell(row=row, column=3).fill = INPUT_FILL
        trans_dv.add(f'C{row}')
        
        ws_cb.cell(row=row, column=4).fill = INPUT_FILL  # Party
        ws_cb.cell(row=row, column=5).fill = INPUT_FILL  # Particulars
        
        ws_cb.cell(row=row, column=6).fill = INPUT_FILL  # Mode
        mode_dv.add(f'F{row}')
        
        for col in [7, 8, 9, 10]:
            ws_cb.cell(row=row, column=col).fill = INPUT_FILL
            ws_cb.cell(row=row, column=col).number_format = MONEY_FORMAT
    
    # Running totals on right side
    ws_cb.cell(row=1, column=12, value="RUNNING TOTALS").font = SUBTITLE_FONT
    totals = [
        ["Opening Cash", "=Masters!$T$6"],
        ["Opening Bank", "=Masters!$T$7"],
        ["Total Cash OUT", "=SUM(G:G)"],
        ["Total Bank OUT", "=SUM(H:H)"],
        ["Total Cash IN", "=SUM(I:I)"],
        ["Total Bank IN", "=SUM(J:J)"],
        ["CLOSING CASH", "=L2-L4+L6"],
        ["CLOSING BANK", "=L3-L5+L7"],
    ]
    for row_idx, (lbl, frm) in enumerate(totals, 2):
        ws_cb.cell(row=row_idx, column=11, value=lbl)
        ws_cb.cell(row=row_idx, column=12, value=frm)
        ws_cb.cell(row=row_idx, column=12).number_format = MONEY_FORMAT
        if "CLOSING" in lbl:
            ws_cb.cell(row=row_idx, column=11).font = Font(bold=True)
            ws_cb.cell(row=row_idx, column=12).font = Font(bold=True)
            ws_cb.cell(row=row_idx, column=12).fill = CALC_FILL
    
    widths = [4, 11, 12, 20, 18, 10, 12, 12, 12, 12, 2, 14, 14]
    for col, w in enumerate(widths, 1):
        set_col_width(ws_cb, col, w)
    
    # ================================================================
    # SHEET 4: BEPAARI_EXPENSES (with Mode of Payment)
    # ================================================================
    ws_be = wb.create_sheet("Bepaari_Expenses", 3)
    ws_be.sheet_properties.tabColor = "F44336"
    
    ws_be.cell(row=1, column=1, value="BEPAARI EXPENSES (Deducted from their payment)").font = TITLE_FONT
    ws_be.cell(row=2, column=1, value="KK (₹100/Bepari) & JB (₹10/goat) auto-calculated in Ledger. Enter only ACTUAL expenses here.").font = Font(italic=True, size=9)
    
    be_headers = ["Sr.", "Date", "Bepaari", "Expense Type", "Amount", "Mode", "Remarks"]
    for col, h in enumerate(be_headers, 1):
        ws_be.cell(row=3, column=col, value=h)
    style_header(ws_be, 3, 1, 7)
    
    # Bepaari dropdown
    bep_exp_dv = DataValidation(type="list", formula1='Masters!$B$3:$B$69', allow_blank=True)
    ws_be.add_data_validation(bep_exp_dv)
    
    # Expense type dropdown
    exp_dv = DataValidation(type="list", formula1='"MOTOR,BHUSSA,GAWALI,CASH,OTHER"', allow_blank=True)
    ws_be.add_data_validation(exp_dv)
    
    # Mode dropdown
    mode_exp_dv = DataValidation(type="list", formula1='"CASH,BANK/UPI"', allow_blank=True)
    ws_be.add_data_validation(mode_exp_dv)
    
    for row in range(4, 504):
        ws_be.cell(row=row, column=1, value=row-3)
        ws_be.cell(row=row, column=2).number_format = DATE_FORMAT
        ws_be.cell(row=row, column=2).fill = INPUT_FILL
        
        ws_be.cell(row=row, column=3).fill = INPUT_FILL
        bep_exp_dv.add(f'C{row}')
        
        ws_be.cell(row=row, column=4).fill = INPUT_FILL
        exp_dv.add(f'D{row}')
        
        ws_be.cell(row=row, column=5).fill = INPUT_FILL
        ws_be.cell(row=row, column=5).number_format = MONEY_FORMAT
        
        ws_be.cell(row=row, column=6).fill = INPUT_FILL
        mode_exp_dv.add(f'F{row}')
        
        ws_be.cell(row=row, column=7).fill = INPUT_FILL
    
    widths = [4, 11, 20, 12, 12, 10, 20]
    for col, w in enumerate(widths, 1):
        set_col_width(ws_be, col, w)
    
    # ================================================================
    # SHEET 5: BEPAARI_LEDGER (with auto KK & JB)
    # ================================================================
    ws_bl = wb.create_sheet("Bepaari_Ledger", 4)
    ws_bl.sheet_properties.tabColor = "9C27B0"
    
    ws_bl.cell(row=1, column=1, value="BEPAARI LEDGER (Auto-calculated)").font = TITLE_FONT
    ws_bl.cell(row=2, column=1, value="KK = ₹100 per Bepari (if sales > 0), JB = ₹10 per goat sold").font = Font(italic=True, size=9)
    
    bl_headers = ["Sr.", "Bepaari", "Type", "Rate", "Opening", "Gross Sales", "Qty", "Commission", 
                  "KK (Auto)", "JB (Auto)", "Motor", "Bhussa", "Gawali", "Cash/Adv", "Other",
                  "Tot.Deduct", "Net Payable", "Payments", "BALANCE"]
    for col, h in enumerate(bl_headers, 1):
        ws_bl.cell(row=4, column=col, value=h)
    style_header(ws_bl, 4, 1, 19)
    
    for row in range(5, 72):  # 67 bepaaris
        m_row = row - 2  # Masters row
        
        ws_bl.cell(row=row, column=1, value=row-4)
        
        # Name from Masters
        ws_bl.cell(row=row, column=2, value=f'=IF(Masters!$B${m_row}="","",Masters!$B${m_row})')
        
        # Comm Type
        ws_bl.cell(row=row, column=3, value=f'=IF(B{row}="","",Masters!$C${m_row})')
        
        # Comm Rate
        ws_bl.cell(row=row, column=4, value=f'=IF(B{row}="","",Masters!$D${m_row})')
        
        # Opening
        ws_bl.cell(row=row, column=5, value=f'=IF(B{row}="",0,Masters!$E${m_row})')
        ws_bl.cell(row=row, column=5).number_format = MONEY_FORMAT
        
        # Gross Sales
        ws_bl.cell(row=row, column=6, value=f'=SUMIF(Daily_Sales!$C:$C,B{row},Daily_Sales!$G:$G)')
        ws_bl.cell(row=row, column=6).number_format = MONEY_FORMAT
        
        # Qty
        ws_bl.cell(row=row, column=7, value=f'=SUMIF(Daily_Sales!$C:$C,B{row},Daily_Sales!$E:$E)')
        ws_bl.cell(row=row, column=7).number_format = NUMBER_FORMAT
        
        # Commission (% or Per Piece)
        ws_bl.cell(row=row, column=8, value=f'=IF(B{row}="",0,IF(C{row}="%",F{row}*D{row}/100,G{row}*D{row}))')
        ws_bl.cell(row=row, column=8).number_format = MONEY_FORMAT
        
        # KK - Auto ₹100 per Bepari (if sales > 0)
        ws_bl.cell(row=row, column=9, value=f'=IF(G{row}>0,Masters!$T$3,0)')
        ws_bl.cell(row=row, column=9).number_format = MONEY_FORMAT
        ws_bl.cell(row=row, column=9).fill = CALC_FILL
        
        # JB - Auto ₹10 per goat
        ws_bl.cell(row=row, column=10, value=f'=G{row}*Masters!$T$4')
        ws_bl.cell(row=row, column=10).number_format = MONEY_FORMAT
        ws_bl.cell(row=row, column=10).fill = CALC_FILL
        
        # Motor (from Bepaari_Expenses)
        ws_bl.cell(row=row, column=11, value=f'=SUMIFS(Bepaari_Expenses!$E:$E,Bepaari_Expenses!$C:$C,B{row},Bepaari_Expenses!$D:$D,"MOTOR")')
        ws_bl.cell(row=row, column=11).number_format = MONEY_FORMAT
        
        # Bhussa
        ws_bl.cell(row=row, column=12, value=f'=SUMIFS(Bepaari_Expenses!$E:$E,Bepaari_Expenses!$C:$C,B{row},Bepaari_Expenses!$D:$D,"BHUSSA")')
        ws_bl.cell(row=row, column=12).number_format = MONEY_FORMAT
        
        # Gawali
        ws_bl.cell(row=row, column=13, value=f'=SUMIFS(Bepaari_Expenses!$E:$E,Bepaari_Expenses!$C:$C,B{row},Bepaari_Expenses!$D:$D,"GAWALI")')
        ws_bl.cell(row=row, column=13).number_format = MONEY_FORMAT
        
        # Cash/Adv
        ws_bl.cell(row=row, column=14, value=f'=SUMIFS(Bepaari_Expenses!$E:$E,Bepaari_Expenses!$C:$C,B{row},Bepaari_Expenses!$D:$D,"CASH")')
        ws_bl.cell(row=row, column=14).number_format = MONEY_FORMAT
        
        # Other
        ws_bl.cell(row=row, column=15, value=f'=SUMIFS(Bepaari_Expenses!$E:$E,Bepaari_Expenses!$C:$C,B{row},Bepaari_Expenses!$D:$D,"OTHER")')
        ws_bl.cell(row=row, column=15).number_format = MONEY_FORMAT
        
        # Total Deductions
        ws_bl.cell(row=row, column=16, value=f'=SUM(H{row}:O{row})')
        ws_bl.cell(row=row, column=16).number_format = MONEY_FORMAT
        ws_bl.cell(row=row, column=16).fill = CALC_FILL
        
        # Net Payable = Gross - Deductions + Opening
        ws_bl.cell(row=row, column=17, value=f'=IF(B{row}="","",F{row}-P{row}+E{row})')
        ws_bl.cell(row=row, column=17).number_format = MONEY_FORMAT
        ws_bl.cell(row=row, column=17).fill = CALC_FILL
        
        # Payments (from Cash_Book BEP_PAY)
        ws_bl.cell(row=row, column=18, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEP_PAY",Cash_Book!$D:$D,B{row})+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEP_PAY",Cash_Book!$D:$D,B{row})')
        ws_bl.cell(row=row, column=18).number_format = MONEY_FORMAT
        
        # Balance = Net Payable - Payments
        ws_bl.cell(row=row, column=19, value=f'=IF(B{row}="","",Q{row}-R{row})')
        ws_bl.cell(row=row, column=19).number_format = MONEY_FORMAT
        ws_bl.cell(row=row, column=19).font = Font(bold=True)
    
    # Totals
    tot_row = 72
    ws_bl.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in [5,6,7,8,9,10,11,12,13,14,15,16,17,18,19]:
        ws_bl.cell(row=tot_row, column=col, value=f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}71)')
        ws_bl.cell(row=tot_row, column=col).number_format = MONEY_FORMAT
        ws_bl.cell(row=tot_row, column=col).font = Font(bold=True)
    
    widths = [4,16,6,6,11,12,6,10,9,10,9,9,9,10,9,11,12,12,12]
    for col, w in enumerate(widths, 1):
        set_col_width(ws_bl, col, w)
    
    # ================================================================
    # SHEET 6: DUKANDAR_LEDGER
    # ================================================================
    ws_dl = wb.create_sheet("Dukandar_Ledger", 5)
    ws_dl.sheet_properties.tabColor = "E91E63"
    
    ws_dl.cell(row=1, column=1, value="DUKANDAR LEDGER (Auto-calculated)").font = TITLE_FONT
    
    dl_headers = ["Sr.", "Dukandar", "Opening", "Purchases", "Discounts", "Net Receivable", "Receipts", "BALANCE"]
    for col, h in enumerate(dl_headers, 1):
        ws_dl.cell(row=3, column=col, value=h)
    style_header(ws_dl, 3, 1, 8)
    
    for row in range(4, 71):
        m_row = row - 1  # Masters row
        
        ws_dl.cell(row=row, column=1, value=row-3)
        
        # Name
        ws_dl.cell(row=row, column=2, value=f'=IF(Masters!$I${m_row}="","",Masters!$I${m_row})')
        
        # Opening
        ws_dl.cell(row=row, column=3, value=f'=IF(B{row}="",0,Masters!$J${m_row})')
        ws_dl.cell(row=row, column=3).number_format = MONEY_FORMAT
        
        # Purchases (Gross from Daily_Sales)
        ws_dl.cell(row=row, column=4, value=f'=SUMIF(Daily_Sales!$D:$D,B{row},Daily_Sales!$G:$G)')
        ws_dl.cell(row=row, column=4).number_format = MONEY_FORMAT
        
        # Discounts
        ws_dl.cell(row=row, column=5, value=f'=SUMIF(Daily_Sales!$D:$D,B{row},Daily_Sales!$H:$H)')
        ws_dl.cell(row=row, column=5).number_format = MONEY_FORMAT
        
        # Net Receivable = Purchases - Discounts + Opening
        ws_dl.cell(row=row, column=6, value=f'=IF(B{row}="","",D{row}-E{row}+C{row})')
        ws_dl.cell(row=row, column=6).number_format = MONEY_FORMAT
        ws_dl.cell(row=row, column=6).fill = CALC_FILL
        
        # Receipts (from Cash_Book DUK_REC)
        ws_dl.cell(row=row, column=7, value=f'=SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"DUK_REC",Cash_Book!$D:$D,B{row})+SUMIFS(Cash_Book!$J:$J,Cash_Book!$C:$C,"DUK_REC",Cash_Book!$D:$D,B{row})')
        ws_dl.cell(row=row, column=7).number_format = MONEY_FORMAT
        
        # Balance = Net Receivable - Receipts
        ws_dl.cell(row=row, column=8, value=f'=IF(B{row}="","",F{row}-G{row})')
        ws_dl.cell(row=row, column=8).number_format = MONEY_FORMAT
        ws_dl.cell(row=row, column=8).font = Font(bold=True)
    
    # Totals
    tot_row = 71
    ws_dl.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in [3,4,5,6,7,8]:
        ws_dl.cell(row=tot_row, column=col, value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}70)')
        ws_dl.cell(row=tot_row, column=col).number_format = MONEY_FORMAT
        ws_dl.cell(row=tot_row, column=col).font = Font(bold=True)
    
    widths = [4,20,12,14,12,14,12,14]
    for col, w in enumerate(widths, 1):
        set_col_width(ws_dl, col, w)
    
    # ================================================================
    # SHEET 7: BEPAARI_AAKDA (Settlement Slip) - FIXED
    # ================================================================
    ws_ak = wb.create_sheet("Bepaari_Aakda", 6)
    ws_ak.sheet_properties.tabColor = "795548"
    
    ws_ak.cell(row=1, column=1, value="BEPAARI SETTLEMENT SLIP (AAKDA)").font = TITLE_FONT
    
    ws_ak.cell(row=3, column=1, value="SELECT BEPAARI:")
    ws_ak.cell(row=3, column=2).fill = INPUT_FILL
    # DROPDOWN for bepaari
    ak_bep_dv = DataValidation(type="list", formula1='Masters!$B$3:$B$69', allow_blank=False)
    ws_ak.add_data_validation(ak_bep_dv)
    ak_bep_dv.add('B3')
    
    ws_ak.cell(row=4, column=1, value="DATE:")
    ws_ak.cell(row=4, column=2).fill = INPUT_FILL
    ws_ak.cell(row=4, column=2).number_format = DATE_FORMAT
    
    ws_ak.cell(row=6, column=1, value="SETTLEMENT DETAILS").font = SUBTITLE_FONT
    
    aakda = [
        ["", ""],
        ["A. SALES SUMMARY", ""],
        ["Total Qty (Goats)", '=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$E:$E)'],
        ["Gross Sales Value", '=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$G:$G)'],
        ["", ""],
        ["B. DEDUCTIONS", ""],
        ["Commission Type", '=IFERROR(INDEX(Masters!$C:$C,MATCH($B$3,Masters!$B:$B,0)),"")'],
        ["Commission Rate/Value", '=IFERROR(INDEX(Masters!$D:$D,MATCH($B$3,Masters!$B:$B,0)),"")'],
        ["Commission Amount", '=IF(B13="%",B10*B14/100,IF(B13="Per Piece",B9*B14,0))'],
        ["Karkuni (KK) Fixed", '=IF(B9>0,Masters!$T$3,0)'],
        ["Jagah Bhada (JB)", '=B9*Masters!$T$4'],
        ["Motor Bhada", '=SUMIFS(Bepaari_Expenses!$E:$E,Bepaari_Expenses!$C:$C,$B$3,Bepaari_Expenses!$D:$D,"MOTOR")'],
        ["Bhussa", '=SUMIFS(Bepaari_Expenses!$E:$E,Bepaari_Expenses!$C:$C,$B$3,Bepaari_Expenses!$D:$D,"BHUSSA")'],
        ["Gawali", '=SUMIFS(Bepaari_Expenses!$E:$E,Bepaari_Expenses!$C:$C,$B$3,Bepaari_Expenses!$D:$D,"GAWALI")'],
        ["Cash/Advance", '=SUMIFS(Bepaari_Expenses!$E:$E,Bepaari_Expenses!$C:$C,$B$3,Bepaari_Expenses!$D:$D,"CASH")'],
        ["Other Expenses", '=SUMIFS(Bepaari_Expenses!$E:$E,Bepaari_Expenses!$C:$C,$B$3,Bepaari_Expenses!$D:$D,"OTHER")'],
        ["TOTAL DEDUCTIONS", '=SUM(B15:B22)'],
        ["", ""],
        ["C. NET CALCULATION", ""],
        ["Gross Sales", '=B10'],
        ["Less: Total Deductions", '=B23'],
        ["Add: Opening Balance", '=IFERROR(INDEX(Masters!$E:$E,MATCH($B$3,Masters!$B:$B,0)),0)'],
        ["NET PAYABLE TO BEPAARI", '=B26-B27+B28'],
        ["", ""],
        ["D. PAYMENTS MADE", ""],
        ["Total Payments", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEP_PAY",Cash_Book!$D:$D,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEP_PAY",Cash_Book!$D:$D,$B$3)'],
        ["", ""],
        ["BALANCE DUE", '=B29-B32'],
    ]
    
    for row_idx, (lbl, frm) in enumerate(aakda, 7):
        ws_ak.cell(row=row_idx, column=1, value=lbl)
        if frm:
            ws_ak.cell(row=row_idx, column=2, value=frm)
            ws_ak.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        
        if any(x in lbl for x in ["A.", "B.", "C.", "D.", "TOTAL", "NET PAYABLE", "BALANCE"]):
            ws_ak.cell(row=row_idx, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl or "BALANCE" in lbl:
                ws_ak.cell(row=row_idx, column=2).font = Font(bold=True)
                ws_ak.cell(row=row_idx, column=2).fill = CALC_FILL
    
    set_col_width(ws_ak, 1, 26)
    set_col_width(ws_ak, 2, 16)
    
    # ================================================================
    # SHEET 8: DAILY_BALANCE_SHEET
    # ================================================================
    ws_bs = wb.create_sheet("Balance_Sheet", 7)
    ws_bs.sheet_properties.tabColor = "607D8B"
    
    ws_bs.cell(row=1, column=1, value="DAILY BALANCE SHEET").font = TITLE_FONT
    ws_bs.cell(row=1, column=5, value="Assets = Liabilities (Both sides must tally)").font = Font(italic=True, size=10)
    
    # LEFT - LIABILITIES
    ws_bs.cell(row=3, column=1, value="LIABILITIES (We Owe)").font = SUBTITLE_FONT
    ws_bs.cell(row=3, column=1).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    ws_bs.cell(row=4, column=1, value="Particulars")
    ws_bs.cell(row=4, column=2, value="Amount")
    style_header(ws_bs, 4, 1, 2, fill=HEADER_RED)
    
    liab = [
        ["CAPITAL", "=Masters!$T$8"],
        ["", ""],
        ["LOANS/AMANAT:", ""],
        ["  (from Masters)", "=SUM(Masters!$P$3:$P$19)"],
        ["", ""],
        ["BEPAARI PAYABLES:", ""],
    ]
    row_idx = 5
    for lbl, frm in liab:
        ws_bs.cell(row=row_idx, column=1, value=lbl)
        if frm:
            ws_bs.cell(row=row_idx, column=2, value=frm)
            ws_bs.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        if ":" in lbl or lbl == "CAPITAL":
            ws_bs.cell(row=row_idx, column=1).font = Font(bold=True)
        row_idx += 1
    
    # Dynamic bepaari balances
    for i in range(1, 68):
        bep_row = i + 4
        ws_bs.cell(row=row_idx, column=1, value=f'=IF(Bepaari_Ledger!B{bep_row}="","",Bepaari_Ledger!B{bep_row})')
        ws_bs.cell(row=row_idx, column=2, value=f'=IF(Bepaari_Ledger!B{bep_row}="","",Bepaari_Ledger!S{bep_row})')
        ws_bs.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        row_idx += 1
    
    liab_tot_row = row_idx
    ws_bs.cell(row=liab_tot_row, column=1, value="TOTAL LIABILITIES").font = Font(bold=True)
    ws_bs.cell(row=liab_tot_row, column=2, value=f'=SUM(B5:B{liab_tot_row-1})')
    ws_bs.cell(row=liab_tot_row, column=2).number_format = MONEY_FORMAT
    ws_bs.cell(row=liab_tot_row, column=2).font = Font(bold=True)
    ws_bs.cell(row=liab_tot_row, column=2).fill = CALC_FILL
    
    # RIGHT - ASSETS
    ws_bs.cell(row=3, column=4, value="ASSETS (We Have / We Are Owed)").font = SUBTITLE_FONT
    ws_bs.cell(row=3, column=4).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    ws_bs.cell(row=4, column=4, value="Particulars")
    ws_bs.cell(row=4, column=5, value="Amount")
    style_header(ws_bs, 4, 4, 5, fill=HEADER_GREEN)
    
    assets = [
        ["CASH BALANCE", "=Cash_Book!L8"],
        ["BANK BALANCE", "=Cash_Book!L9"],
        ["", ""],
        ["DUKANDAR RECEIVABLES:", ""],
    ]
    row_idx = 5
    for lbl, frm in assets:
        ws_bs.cell(row=row_idx, column=4, value=lbl)
        if frm:
            ws_bs.cell(row=row_idx, column=5, value=frm)
            ws_bs.cell(row=row_idx, column=5).number_format = MONEY_FORMAT
        if ":" in lbl or "BALANCE" in lbl:
            ws_bs.cell(row=row_idx, column=4).font = Font(bold=True)
        row_idx += 1
    
    # Dynamic dukandar balances
    for i in range(1, 68):
        duk_row = i + 3
        ws_bs.cell(row=row_idx, column=4, value=f'=IF(Dukandar_Ledger!B{duk_row}="","",Dukandar_Ledger!B{duk_row})')
        ws_bs.cell(row=row_idx, column=5, value=f'=IF(Dukandar_Ledger!B{duk_row}="","",Dukandar_Ledger!H{duk_row})')
        ws_bs.cell(row=row_idx, column=5).number_format = MONEY_FORMAT
        row_idx += 1
    
    # Extra items
    extras = [
        ["", ""],
        ["COMMISSION EARNED", "=Bepaari_Ledger!H72"],
        ["Less: Discounts Given", "=-Dukandar_Ledger!E71"],
        ["Net Commission", "=E75+E76"],
        ["", ""],
        ["JB Collected", "=Bepaari_Ledger!J72"],
        ["ZAKAT PAID", '=-SUMIF(Cash_Book!$C:$C,"ZAKAT",Cash_Book!$G:$G)'],
        ["BF (Cash Disc)", '=SUMIF(Cash_Book!$C:$C,"BF_DISC",Cash_Book!$I:$I)'],
    ]
    for lbl, frm in extras:
        ws_bs.cell(row=row_idx, column=4, value=lbl)
        if frm:
            ws_bs.cell(row=row_idx, column=5, value=frm)
            ws_bs.cell(row=row_idx, column=5).number_format = MONEY_FORMAT
        if "Net" in str(lbl) or "COMMISSION" in str(lbl):
            ws_bs.cell(row=row_idx, column=4).font = Font(bold=True)
        row_idx += 1
    
    asset_tot_row = liab_tot_row  # Align
    ws_bs.cell(row=asset_tot_row, column=4, value="TOTAL ASSETS").font = Font(bold=True)
    ws_bs.cell(row=asset_tot_row, column=5, value=f'=SUM(E5:E{asset_tot_row-1})')
    ws_bs.cell(row=asset_tot_row, column=5).number_format = MONEY_FORMAT
    ws_bs.cell(row=asset_tot_row, column=5).font = Font(bold=True)
    ws_bs.cell(row=asset_tot_row, column=5).fill = CALC_FILL
    
    # Difference check
    ws_bs.cell(row=liab_tot_row+2, column=1, value="DIFFERENCE (Should be 0)").font = Font(bold=True, color="FF0000")
    ws_bs.cell(row=liab_tot_row+2, column=2, value=f'=B{liab_tot_row}-E{asset_tot_row}')
    ws_bs.cell(row=liab_tot_row+2, column=2).number_format = MONEY_FORMAT
    ws_bs.cell(row=liab_tot_row+2, column=2).font = Font(bold=True, color="FF0000")
    
    set_col_width(ws_bs, 1, 22)
    set_col_width(ws_bs, 2, 14)
    set_col_width(ws_bs, 3, 3)
    set_col_width(ws_bs, 4, 22)
    set_col_width(ws_bs, 5, 14)
    
    # ================================================================
    # SHEET 9: COMMISSION_SUMMARY
    # ================================================================
    ws_cs = wb.create_sheet("Commission_Summary", 8)
    ws_cs.sheet_properties.tabColor = "FFC107"
    
    ws_cs.cell(row=1, column=1, value="COMMISSION SUMMARY").font = TITLE_FONT
    
    comm = [
        ["", ""],
        ["GROSS SALES (All Bepaaris)", "=Bepaari_Ledger!F72"],
        ["", ""],
        ["COMMISSION EARNED (Gross)", "=Bepaari_Ledger!H72"],
        ["Less: Discounts to Dukandars", "=Dukandar_Ledger!E71"],
        ["NET COMMISSION", "=B4-B5"],
        ["", ""],
        ["ADDITIONAL INCOME:", ""],
        ["  Jagah Bhada Collected", "=Bepaari_Ledger!J72"],
        ["  Karkuni Collected", "=Bepaari_Ledger!I72"],
        ["", ""],
        ["TOTAL INCOME", "=B6+B9+B10"],
        ["", ""],
        ["EXPENSES (from Cash_Book):", ""],
        ["  Motor/Transport", '=SUMIF(Cash_Book!$C:$C,"EXP_MOTOR",Cash_Book!$G:$G)+SUMIF(Cash_Book!$C:$C,"EXP_MOTOR",Cash_Book!$H:$H)'],
        ["  Conveyance/OLA", '=SUMIF(Cash_Book!$C:$C,"EXP_CONV",Cash_Book!$G:$G)+SUMIF(Cash_Book!$C:$C,"EXP_CONV",Cash_Book!$H:$H)'],
        ["  Food", '=SUMIF(Cash_Book!$C:$C,"EXP_FOOD",Cash_Book!$G:$G)+SUMIF(Cash_Book!$C:$C,"EXP_FOOD",Cash_Book!$H:$H)'],
        ["  Traveling", '=SUMIF(Cash_Book!$C:$C,"EXP_TRAVEL",Cash_Book!$G:$G)+SUMIF(Cash_Book!$C:$C,"EXP_TRAVEL",Cash_Book!$H:$H)'],
        ["  Salary", '=SUMIF(Cash_Book!$C:$C,"EXP_SALARY",Cash_Book!$G:$G)+SUMIF(Cash_Book!$C:$C,"EXP_SALARY",Cash_Book!$H:$H)'],
        ["  Zakat", '=SUMIF(Cash_Book!$C:$C,"ZAKAT",Cash_Book!$G:$G)+SUMIF(Cash_Book!$C:$C,"ZAKAT",Cash_Book!$H:$H)'],
        ["  Other", '=SUMIF(Cash_Book!$C:$C,"EXP_OTHER",Cash_Book!$G:$G)+SUMIF(Cash_Book!$C:$C,"EXP_OTHER",Cash_Book!$H:$H)'],
        ["TOTAL EXPENSES", "=SUM(B15:B21)"],
        ["", ""],
        ["NET PROFIT/LOSS", "=B12-B22"],
    ]
    
    for row_idx, (lbl, frm) in enumerate(comm, 2):
        ws_cs.cell(row=row_idx, column=1, value=lbl)
        if frm:
            ws_cs.cell(row=row_idx, column=2, value=frm)
            ws_cs.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        
        if any(x in lbl for x in ["GROSS", "NET", "TOTAL"]):
            ws_cs.cell(row=row_idx, column=1).font = Font(bold=True)
            ws_cs.cell(row=row_idx, column=2).font = Font(bold=True)
            if "NET" in lbl or "TOTAL INCOME" in lbl:
                ws_cs.cell(row=row_idx, column=2).fill = CALC_FILL
    
    set_col_width(ws_cs, 1, 30)
    set_col_width(ws_cs, 2, 16)
    
    # ================================================================
    # SHEET 10: DAILY_SUMMARY (Fixed - no ₹ on goat count)
    # ================================================================
    ws_ds = wb.create_sheet("Daily_Summary", 9)
    ws_ds.sheet_properties.tabColor = "00BCD4"
    
    ws_ds.cell(row=1, column=1, value="DAILY SUMMARY").font = TITLE_FONT
    
    ws_ds.cell(row=3, column=1, value="SELECT DATE:")
    ws_ds.cell(row=3, column=2).fill = INPUT_FILL
    ws_ds.cell(row=3, column=2).number_format = DATE_FORMAT
    
    summary = [
        ["", ""],
        ["TRADING SUMMARY", ""],
        ["Total Goats Sold", '=SUMIF(Daily_Sales!$B:$B,$B$3,Daily_Sales!$E:$E)'],
        ["Gross Sales Value", '=SUMIF(Daily_Sales!$B:$B,$B$3,Daily_Sales!$G:$G)'],
        ["Total Discounts Given", '=SUMIF(Daily_Sales!$B:$B,$B$3,Daily_Sales!$H:$H)'],
        ["Net Sales", "=B8-B9"],
        ["", ""],
        ["Commission Earned (4%)", "=B8*Masters!$T$2/100"],
    ]
    
    for row_idx, (lbl, frm) in enumerate(summary, 4):
        ws_ds.cell(row=row_idx, column=1, value=lbl)
        if frm:
            ws_ds.cell(row=row_idx, column=2, value=frm)
            # No ₹ for goat count
            if "Goats" in lbl:
                ws_ds.cell(row=row_idx, column=2).number_format = NUMBER_FORMAT
            else:
                ws_ds.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        
        if "SUMMARY" in lbl or "Net" in lbl:
            ws_ds.cell(row=row_idx, column=1).font = Font(bold=True)
    
    set_col_width(ws_ds, 1, 26)
    set_col_width(ws_ds, 2, 14)
    
    # ================================================================
    # SAVE
    # ================================================================
    wb.save(output_path)
    print(f"✅ Mandi Master V3 Excel created: {output_path}")
    return output_path


if __name__ == "__main__":
    output_file = "/app/Mandi_Master_V3.xlsx"
    create_mandi_excel_v3(output_file)
