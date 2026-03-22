"""Mandi Master V8 - Type-specific Sub-Type dropdowns + Excess payments in Balance Sheet"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_RED = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
HEADER_GREEN = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="0D47A1")
SUBTITLE_FONT = Font(bold=True, size=11, color="1565C0")
INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
CALC_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
MONEY_FMT = '₹#,##0'
NUM_FMT = '#,##0'
DATE_FMT = 'DD-MMM-YYYY'

def col_w(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def hdr(ws, row, c1, c2, fill=None):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def create_v8(path):
    wb = Workbook()
    wb.remove(wb.active)
    
    # Named ranges for INDIRECT dropdown - PARTY lists
    wb.defined_names.add(DefinedName("BEPAARI", attr_text="Masters!$B$3:$B$79"))
    wb.defined_names.add(DefinedName("DUKANDAR", attr_text="Masters!$I$3:$I$79"))
    wb.defined_names.add(DefinedName("CAPITAL", attr_text="Masters!$O$3:$O$34"))
    wb.defined_names.add(DefinedName("LOAN", attr_text="Masters!$O$3:$O$34"))
    wb.defined_names.add(DefinedName("ADVANCE", attr_text="Masters!$T$3:$T$24"))
    wb.defined_names.add(DefinedName("EXPENSE", attr_text="Lookups!$B$3:$B$3"))  # No party needed
    wb.defined_names.add(DefinedName("ZAKAT", attr_text="Lookups!$B$3:$B$3"))    # No party needed
    wb.defined_names.add(DefinedName("OTHER", attr_text="Lookups!$B$3:$B$3"))    # No party needed
    
    # Named ranges for INDIRECT dropdown - SUB-TYPE lists (Type-specific)
    wb.defined_names.add(DefinedName("ST_BEPAARI", attr_text="Lookups!$D$3:$D$8"))
    wb.defined_names.add(DefinedName("ST_DUKANDAR", attr_text="Lookups!$E$3:$E$5"))
    wb.defined_names.add(DefinedName("ST_CAPITAL", attr_text="Lookups!$F$3:$F$4"))
    wb.defined_names.add(DefinedName("ST_LOAN", attr_text="Lookups!$G$3:$G$4"))
    wb.defined_names.add(DefinedName("ST_ADVANCE", attr_text="Lookups!$H$3:$H$4"))
    wb.defined_names.add(DefinedName("ST_EXPENSE", attr_text="Lookups!$I$3:$I$10"))
    wb.defined_names.add(DefinedName("ST_ZAKAT", attr_text="Lookups!$J$3:$J$4"))
    wb.defined_names.add(DefinedName("ST_OTHER", attr_text="Lookups!$K$3:$K$4"))
    
    # ========== LOOKUPS (Hidden helper sheet for dropdowns) ==========
    ws = wb.create_sheet("Lookups", 0)
    ws.sheet_properties.tabColor = "9E9E9E"
    
    ws.cell(row=1, column=1, value="LOOKUP TABLES FOR DROPDOWNS").font = TITLE_FONT
    
    # Column B - Empty party placeholder
    ws.cell(row=2, column=2, value="NO_PARTY")
    ws.cell(row=3, column=2, value="-")
    
    # Column D - BEPAARI Sub-Types
    ws.cell(row=2, column=4, value="ST_BEPAARI").font = Font(bold=True)
    for i, st in enumerate(["PAYMENT", "MOTOR", "BHUSSA", "GAWALI", "CASH_ADV", "OTHER"], 3):
        ws.cell(row=i, column=4, value=st)
    
    # Column E - DUKANDAR Sub-Types
    ws.cell(row=2, column=5, value="ST_DUKANDAR").font = Font(bold=True)
    for i, st in enumerate(["RECEIPT", "DISCOUNT", "OTHER"], 3):
        ws.cell(row=i, column=5, value=st)
    
    # Column F - CAPITAL Sub-Types
    ws.cell(row=2, column=6, value="ST_CAPITAL").font = Font(bold=True)
    for i, st in enumerate(["TAKEN", "WITHDRAWN"], 3):
        ws.cell(row=i, column=6, value=st)
    
    # Column G - LOAN Sub-Types
    ws.cell(row=2, column=7, value="ST_LOAN").font = Font(bold=True)
    for i, st in enumerate(["TAKEN", "REPAID"], 3):
        ws.cell(row=i, column=7, value=st)
    
    # Column H - ADVANCE Sub-Types
    ws.cell(row=2, column=8, value="ST_ADVANCE").font = Font(bold=True)
    for i, st in enumerate(["GIVEN", "RECEIVED"], 3):
        ws.cell(row=i, column=8, value=st)
    
    # Column I - EXPENSE Sub-Types
    ws.cell(row=2, column=9, value="ST_EXPENSE").font = Font(bold=True)
    for i, st in enumerate(["MANDI", "TRAVEL", "FOOD", "SALARY", "BF_DISC", "JB_PAID", "MISC", "OTHER"], 3):
        ws.cell(row=i, column=9, value=st)
    
    # Column J - ZAKAT Sub-Types
    ws.cell(row=2, column=10, value="ST_ZAKAT").font = Font(bold=True)
    for i, st in enumerate(["PROVISION", "PAID"], 3):
        ws.cell(row=i, column=10, value=st)
    
    # Column K - OTHER Sub-Types
    ws.cell(row=2, column=11, value="ST_OTHER").font = Font(bold=True)
    for i, st in enumerate(["MISC", "OTHER"], 3):
        ws.cell(row=i, column=11, value=st)
    
    # ========== MASTERS ==========
    ws = wb.create_sheet("Masters", 1)
    ws.sheet_properties.tabColor = "1565C0"
    
    ws.cell(row=1, column=1, value="BEPAARI MASTER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Bepaari Name", "Comm%", "Opening Bal", "Phone", "Remarks"], 1):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 1, 6)
    bepaaris = ["ABUL HASAN", "SHARAFAT", "SHARIK", "JUNAID", "ANNU SHARIF", "GAYYUR", "ARIF KK", "KISHOR", "SHOEB BYCULLA", "MAULANA", "ISLAM D", "SALIM RAJ", "MEHMOOD MADH", "ANAS RAJ", "MOBIN CHIPLUN", "AYAZ", "JALIL", "NEERAJ", "IRFAN JOGESHWARI", "CHINTU SHUJALPUR"]
    for i, name in enumerate(bepaaris, 1):
        ws.cell(row=i+2, column=1, value=i)
        ws.cell(row=i+2, column=2, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=3, value=4).fill = INPUT_FILL
        ws.cell(row=i+2, column=4, value=0).fill = INPUT_FILL
        ws.cell(row=i+2, column=4).number_format = MONEY_FMT
    for r in range(len(bepaaris)+3, 80):
        ws.cell(row=r, column=1, value=r-2)
        for c in range(2, 7): ws.cell(row=r, column=c).fill = INPUT_FILL
    
    ws.cell(row=1, column=8, value="DUKANDAR MASTER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Dukandar Name", "Opening Bal", "Phone", "Remarks"], 8):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 8, 12)
    dukandars = ["SAKIM", "SHARIK", "YUNUS", "MUDASSIR", "ARIF", "SHOEB BYCULLA", "ISLAM D", "MEHMOOD MADH", "MOBIN CHIPLUN", "JALIL", "IRFAN JOGESHWARI", "YUNUS DOMBIVLI", "IFTEKHAR PUNE", "LALA/ASIF", "JAFER KASHIMIRA", "CHAND", "ASIF GOREGAON", "CHINTU SHUJALPUR", "VAZIR", "AKRAM", "NADEEM BHIWANDI", "SHAKIL KON", "PARVEZ"]
    for i, name in enumerate(dukandars, 1):
        ws.cell(row=i+2, column=8, value=i)
        ws.cell(row=i+2, column=9, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=10, value=0).fill = INPUT_FILL
        ws.cell(row=i+2, column=10).number_format = MONEY_FMT
    for r in range(len(dukandars)+3, 80):
        ws.cell(row=r, column=8, value=r-2)
        for c in range(9, 13): ws.cell(row=r, column=c).fill = INPUT_FILL
    
    # CAPITAL/LOAN PARTIES - Extended to 32 rows
    ws.cell(row=1, column=14, value="CAPITAL/LOAN PARTIES").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Party Name", "Opening Bal", "Type"], 14):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 14, 17)
    cap_parties = [["MHN", 3900000, "CAPITAL"], ["SHAKIL GHODEGAON", 2000000, "LOAN"]]
    for i, (name, bal, typ) in enumerate(cap_parties, 1):
        ws.cell(row=i+2, column=14, value=i)
        ws.cell(row=i+2, column=15, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=16, value=bal).fill = INPUT_FILL
        ws.cell(row=i+2, column=16).number_format = MONEY_FMT
        ws.cell(row=i+2, column=17, value=typ).fill = INPUT_FILL
    for r in range(len(cap_parties)+3, 35):
        ws.cell(row=r, column=14, value=r-2)
        for c in range(15, 18): ws.cell(row=r, column=c).fill = INPUT_FILL
    type_dv = DataValidation(type="list", formula1='"CAPITAL,LOAN,AMANAT"', allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add('Q3:Q34')
    
    ws.cell(row=1, column=19, value="ADVANCE PARTIES").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Party Name", "Opening Bal"], 19):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 19, 21)
    for r in range(3, 25):
        ws.cell(row=r, column=19, value=r-2)
        for c in range(20, 22): ws.cell(row=r, column=c).fill = INPUT_FILL
    
    ws.cell(row=1, column=23, value="SETTINGS").font = TITLE_FONT
    settings = [["Commission Rate (%)", 4], ["KK Fixed (₹/Bepari)", 100], ["JB Rate (₹/Goat)", 10], ["", ""], ["Opening Cash", 0], ["Opening Bank", 0], ["", ""], ["ZAKAT Opening Payable", 0]]
    for r, (lbl, val) in enumerate(settings, 2):
        ws.cell(row=r, column=23, value=lbl).font = Font(bold=True) if lbl else None
        c = ws.cell(row=r, column=24, value=val)
        c.fill = INPUT_FILL
        if "Cash" in str(lbl) or "Bank" in str(lbl) or "ZAKAT" in str(lbl): c.number_format = MONEY_FMT
    for c, w in {1:4, 2:20, 3:8, 4:12, 5:10, 6:10, 7:2, 8:4, 9:20, 10:12, 11:10, 12:10, 13:2, 14:4, 15:20, 16:12, 17:10, 18:2, 19:4, 20:20, 21:12, 22:2, 23:22, 24:12}.items():
        col_w(ws, c, w)
    
    # ========== DAILY_SALES ==========
    ws = wb.create_sheet("Daily_Sales", 2)
    ws.sheet_properties.tabColor = "4CAF50"
    ws.cell(row=1, column=1, value="DAILY SALES (Bepaari to Dukandar)").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Date", "Bepaari", "Dukandar", "Qty", "Rate", "Gross", "Discount", "Net"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 9)
    bep_dv = DataValidation(type="list", formula1='BEPAARI', allow_blank=True)
    duk_dv = DataValidation(type="list", formula1='DUKANDAR', allow_blank=True)
    ws.add_data_validation(bep_dv)
    ws.add_data_validation(duk_dv)
    for r in range(4, 1004):
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3).fill = INPUT_FILL
        bep_dv.add(f'C{r}')
        ws.cell(row=r, column=4).fill = INPUT_FILL
        duk_dv.add(f'D{r}')
        ws.cell(row=r, column=5).fill = INPUT_FILL
        ws.cell(row=r, column=6).fill = INPUT_FILL
        ws.cell(row=r, column=6).number_format = MONEY_FMT
        ws.cell(row=r, column=7, value=f'=IF(E{r}="","",E{r}*F{r})')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=7).fill = CALC_FILL
        ws.cell(row=r, column=8).fill = INPUT_FILL
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        ws.cell(row=r, column=9, value=f'=IF(G{r}="","",G{r}-IF(H{r}="",0,H{r}))')
        ws.cell(row=r, column=9).number_format = MONEY_FMT
        ws.cell(row=r, column=9).fill = CALC_FILL
    for c, w in enumerate([4, 11, 20, 20, 6, 10, 12, 10, 12], 1): col_w(ws, c, w)
    
    # ========== CASH_BOOK - Type-specific Sub-Type dropdowns ==========
    ws = wb.create_sheet("Cash_Book", 3)
    ws.sheet_properties.tabColor = "FF5722"
    ws.cell(row=1, column=1, value="CASH BOOK (All Transactions)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Select TYPE first → Sub-Type & Party dropdowns update automatically").font = Font(italic=True, size=9)
    
    # Columns: Sr, Date, Type, Sub-Type, Party, Particulars, Amount, Mode
    for c, h in enumerate(["Sr", "Date", "Type", "Sub-Type", "Party", "Particulars", "Amount", "Mode"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 8)
    
    # Type dropdown
    main_dv = DataValidation(type="list", formula1='"BEPAARI,DUKANDAR,CAPITAL,LOAN,ADVANCE,EXPENSE,ZAKAT,OTHER"', allow_blank=True)
    ws.add_data_validation(main_dv)
    
    # Mode of Payment dropdown
    mode_dv = DataValidation(type="list", formula1='"CASH,BANK,UPI,TRANSFER"', allow_blank=True)
    ws.add_data_validation(mode_dv)
    
    for r in range(4, 2004):
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3).fill = INPUT_FILL
        main_dv.add(f'C{r}')
        
        # DYNAMIC SUB-TYPE DROPDOWN using INDIRECT with "ST_" prefix
        ws.cell(row=r, column=4).fill = INPUT_FILL
        subtype_dv = DataValidation(type="list", formula1=f'INDIRECT("ST_"&$C{r})', allow_blank=True)
        ws.add_data_validation(subtype_dv)
        subtype_dv.add(f'D{r}')
        
        # DYNAMIC PARTY DROPDOWN using INDIRECT
        ws.cell(row=r, column=5).fill = INPUT_FILL
        party_dv = DataValidation(type="list", formula1=f'INDIRECT($C{r})', allow_blank=True)
        ws.add_data_validation(party_dv)
        party_dv.add(f'E{r}')
        
        ws.cell(row=r, column=6).fill = INPUT_FILL
        ws.cell(row=r, column=7).fill = INPUT_FILL
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=8).fill = INPUT_FILL
        mode_dv.add(f'H{r}')
    
    # CASH/BANK SUMMARY
    ws.cell(row=1, column=10, value="CASH/BANK SUMMARY").font = SUBTITLE_FONT
    
    ws.cell(row=2, column=9, value="Opening Cash")
    ws.cell(row=2, column=10, value="=Masters!$X$6")
    ws.cell(row=2, column=10).number_format = MONEY_FMT
    
    ws.cell(row=3, column=9, value="Opening Bank")
    ws.cell(row=3, column=10, value="=Masters!$X$7")
    ws.cell(row=3, column=10).number_format = MONEY_FMT
    
    # Outgoing sub-types: PAYMENT, MOTOR, BHUSSA, GAWALI, CASH_ADV, DISCOUNT, WITHDRAWN, REPAID, GIVEN, MANDI, TRAVEL, FOOD, SALARY, BF_DISC, JB_PAID, PAID, MISC, OTHER
    # Incoming sub-types: RECEIPT, TAKEN, RECEIVED, PROVISION
    
    # Cash OUT
    ws.cell(row=5, column=9, value="Cash OUT")
    cash_out_formula = '=SUMIFS(G:G,H:H,"CASH",D:D,"PAYMENT")+SUMIFS(G:G,H:H,"CASH",D:D,"MOTOR")+SUMIFS(G:G,H:H,"CASH",D:D,"BHUSSA")+SUMIFS(G:G,H:H,"CASH",D:D,"GAWALI")+SUMIFS(G:G,H:H,"CASH",D:D,"CASH_ADV")+SUMIFS(G:G,H:H,"CASH",D:D,"DISCOUNT")+SUMIFS(G:G,H:H,"CASH",D:D,"WITHDRAWN")+SUMIFS(G:G,H:H,"CASH",D:D,"REPAID")+SUMIFS(G:G,H:H,"CASH",D:D,"GIVEN")+SUMIFS(G:G,H:H,"CASH",D:D,"MANDI")+SUMIFS(G:G,H:H,"CASH",D:D,"TRAVEL")+SUMIFS(G:G,H:H,"CASH",D:D,"FOOD")+SUMIFS(G:G,H:H,"CASH",D:D,"SALARY")+SUMIFS(G:G,H:H,"CASH",D:D,"BF_DISC")+SUMIFS(G:G,H:H,"CASH",D:D,"JB_PAID")+SUMIFS(G:G,H:H,"CASH",D:D,"PAID")+SUMIFS(G:G,H:H,"CASH",D:D,"MISC")+SUMIFS(G:G,H:H,"CASH",D:D,"OTHER")'
    ws.cell(row=5, column=10, value=cash_out_formula)
    ws.cell(row=5, column=10).number_format = MONEY_FMT
    
    # Bank OUT (BANK + UPI + TRANSFER)
    ws.cell(row=6, column=9, value="Bank OUT")
    bank_out_parts = []
    for mode in ["BANK", "UPI", "TRANSFER"]:
        for st in ["PAYMENT", "MOTOR", "BHUSSA", "GAWALI", "CASH_ADV", "DISCOUNT", "WITHDRAWN", "REPAID", "GIVEN", "MANDI", "TRAVEL", "FOOD", "SALARY", "BF_DISC", "JB_PAID", "PAID", "MISC", "OTHER"]:
            bank_out_parts.append(f'SUMIFS(G:G,H:H,"{mode}",D:D,"{st}")')
    ws.cell(row=6, column=10, value='=' + '+'.join(bank_out_parts))
    ws.cell(row=6, column=10).number_format = MONEY_FMT
    
    # Cash IN
    ws.cell(row=7, column=9, value="Cash IN")
    ws.cell(row=7, column=10, value='=SUMIFS(G:G,H:H,"CASH",D:D,"RECEIPT")+SUMIFS(G:G,H:H,"CASH",D:D,"TAKEN")+SUMIFS(G:G,H:H,"CASH",D:D,"RECEIVED")+SUMIFS(G:G,H:H,"CASH",D:D,"PROVISION")')
    ws.cell(row=7, column=10).number_format = MONEY_FMT
    
    # Bank IN
    ws.cell(row=8, column=9, value="Bank IN")
    bank_in_parts = []
    for mode in ["BANK", "UPI", "TRANSFER"]:
        for st in ["RECEIPT", "TAKEN", "RECEIVED", "PROVISION"]:
            bank_in_parts.append(f'SUMIFS(G:G,H:H,"{mode}",D:D,"{st}")')
    ws.cell(row=8, column=10, value='=' + '+'.join(bank_in_parts))
    ws.cell(row=8, column=10).number_format = MONEY_FMT
    
    ws.cell(row=10, column=9, value="CLOSING CASH").font = Font(bold=True)
    ws.cell(row=10, column=10, value='=J2-J5+J7')
    ws.cell(row=10, column=10).number_format = MONEY_FMT
    ws.cell(row=10, column=10).font = Font(bold=True)
    ws.cell(row=10, column=10).fill = CALC_FILL
    
    ws.cell(row=11, column=9, value="CLOSING BANK").font = Font(bold=True)
    ws.cell(row=11, column=10, value='=J3-J6+J8')
    ws.cell(row=11, column=10).number_format = MONEY_FMT
    ws.cell(row=11, column=10).font = Font(bold=True)
    ws.cell(row=11, column=10).fill = CALC_FILL
    
    for c, w in enumerate([4, 11, 12, 14, 20, 20, 12, 10, 2, 14, 14], 1): col_w(ws, c, w)
    
    # ========== BEPAARI_LEDGER ==========
    ws = wb.create_sheet("Bepaari_Ledger", 4)
    ws.sheet_properties.tabColor = "9C27B0"
    ws.cell(row=1, column=1, value="BEPAARI LEDGER (Auto-calculated)").font = TITLE_FONT
    hdrs = ["Sr", "Bepaari", "Comm%", "Opening", "Gross Sales", "Qty", "Commission", "KK", "JB", "Motor", "Bhussa", "Gawali", "Cash/Adv", "Other", "Tot.Deduct", "Net Payable", "Payments", "BALANCE"]
    for c, h in enumerate(hdrs, 1): ws.cell(row=4, column=c, value=h)
    hdr(ws, 4, 1, 18)
    for r in range(5, 82):
        m = r - 2
        ws.cell(row=r, column=1, value=r-4)
        ws.cell(row=r, column=2, value=f'=IF(Masters!$B${m}="","",Masters!$B${m})')
        ws.cell(row=r, column=3, value=f'=IF(B{r}="","",Masters!$C${m})')
        ws.cell(row=r, column=4, value=f'=IF(B{r}="",0,Masters!$D${m})')
        ws.cell(row=r, column=4).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f'=SUMIF(Daily_Sales!$C:$C,B{r},Daily_Sales!$G:$G)')
        ws.cell(row=r, column=5).number_format = MONEY_FMT
        ws.cell(row=r, column=6, value=f'=SUMIF(Daily_Sales!$C:$C,B{r},Daily_Sales!$E:$E)')
        ws.cell(row=r, column=6).number_format = NUM_FMT
        ws.cell(row=r, column=7, value=f'=IF(B{r}="",0,E{r}*C{r}/100)')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=8, value=f'=IF(F{r}>0,Masters!$X$3,0)')
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        ws.cell(row=r, column=8).fill = CALC_FILL
        ws.cell(row=r, column=9, value=f'=F{r}*Masters!$X$4')
        ws.cell(row=r, column=9).number_format = MONEY_FMT
        ws.cell(row=r, column=9).fill = CALC_FILL
        for col, sub in [(10,"MOTOR"), (11,"BHUSSA"), (12,"GAWALI"), (13,"CASH_ADV"), (14,"OTHER")]:
            ws.cell(row=r, column=col, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"{sub}",Cash_Book!$E:$E,B{r})')
            ws.cell(row=r, column=col).number_format = MONEY_FMT
        ws.cell(row=r, column=15, value=f'=SUM(G{r}:N{r})')
        ws.cell(row=r, column=15).number_format = MONEY_FMT
        ws.cell(row=r, column=15).fill = CALC_FILL
        ws.cell(row=r, column=16, value=f'=IF(B{r}="","",E{r}-O{r}+D{r})')
        ws.cell(row=r, column=16).number_format = MONEY_FMT
        ws.cell(row=r, column=16).fill = CALC_FILL
        ws.cell(row=r, column=17, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=17).number_format = MONEY_FMT
        ws.cell(row=r, column=18, value=f'=IF(B{r}="","",P{r}-Q{r})')
        ws.cell(row=r, column=18).number_format = MONEY_FMT
        ws.cell(row=r, column=18).font = Font(bold=True)
    ws.cell(row=82, column=1, value="TOTAL").font = Font(bold=True)
    for c in [4,5,6,7,8,9,10,11,12,13,14,15,16,17,18]:
        ws.cell(row=82, column=c, value=f'=SUM({get_column_letter(c)}5:{get_column_letter(c)}81)')
        ws.cell(row=82, column=c).number_format = MONEY_FMT
        ws.cell(row=82, column=c).font = Font(bold=True)
    
    # Additional totals for Balance Sheet: Payable (positive) vs Overpaid (negative)
    ws.cell(row=84, column=15, value="For Balance Sheet:").font = Font(bold=True, italic=True)
    ws.cell(row=85, column=16, value="Bepaari Payable (we owe)")
    ws.cell(row=85, column=18, value='=SUMIF(R5:R81,">0",R5:R81)')
    ws.cell(row=85, column=18).number_format = MONEY_FMT
    ws.cell(row=85, column=18).fill = CALC_FILL
    ws.cell(row=86, column=16, value="Bepaari Overpaid (they owe us)")
    ws.cell(row=86, column=18, value='=-SUMIF(R5:R81,"<0",R5:R81)')
    ws.cell(row=86, column=18).number_format = MONEY_FMT
    ws.cell(row=86, column=18).fill = CALC_FILL
    
    for c, w in enumerate([4,16,6,10,12,6,10,8,8,9,9,9,10,8,11,12,11,12], 1): col_w(ws, c, w)
    
    # ========== DUKANDAR_LEDGER ==========
    ws = wb.create_sheet("Dukandar_Ledger", 5)
    ws.sheet_properties.tabColor = "E91E63"
    ws.cell(row=1, column=1, value="DUKANDAR LEDGER (Auto-calculated)").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Dukandar", "Opening", "Purchases", "Discounts", "Net Receivable", "Receipts", "BALANCE"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 8)
    for r in range(4, 81):
        m = r - 1
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2, value=f'=IF(Masters!$I${m}="","",Masters!$I${m})')
        ws.cell(row=r, column=3, value=f'=IF(B{r}="",0,Masters!$J${m})')
        ws.cell(row=r, column=3).number_format = MONEY_FMT
        ws.cell(row=r, column=4, value=f'=SUMIF(Daily_Sales!$D:$D,B{r},Daily_Sales!$G:$G)')
        ws.cell(row=r, column=4).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f'=SUMIF(Daily_Sales!$D:$D,B{r},Daily_Sales!$H:$H)')
        ws.cell(row=r, column=5).number_format = MONEY_FMT
        ws.cell(row=r, column=6, value=f'=IF(B{r}="","",D{r}-E{r}+C{r})')
        ws.cell(row=r, column=6).number_format = MONEY_FMT
        ws.cell(row=r, column=6).fill = CALC_FILL
        ws.cell(row=r, column=7, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"DUKANDAR",Cash_Book!$D:$D,"RECEIPT",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=8, value=f'=IF(B{r}="","",F{r}-G{r})')
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        ws.cell(row=r, column=8).font = Font(bold=True)
    ws.cell(row=81, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=81, column=2, value="PATTI (Total O/S)")
    for c in [3,4,5,6,7,8]:
        ws.cell(row=81, column=c, value=f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}80)')
        ws.cell(row=81, column=c).number_format = MONEY_FMT
        ws.cell(row=81, column=c).font = Font(bold=True)
    
    # Additional totals for Balance Sheet: Receivable (positive) vs Overpaid (negative)
    ws.cell(row=83, column=5, value="For Balance Sheet:").font = Font(bold=True, italic=True)
    ws.cell(row=84, column=6, value="Dukandar Receivable (they owe us)")
    ws.cell(row=84, column=8, value='=SUMIF(H4:H80,">0",H4:H80)')
    ws.cell(row=84, column=8).number_format = MONEY_FMT
    ws.cell(row=84, column=8).fill = CALC_FILL
    ws.cell(row=85, column=6, value="Dukandar Overpaid (we owe them)")
    ws.cell(row=85, column=8, value='=-SUMIF(H4:H80,"<0",H4:H80)')
    ws.cell(row=85, column=8).number_format = MONEY_FMT
    ws.cell(row=85, column=8).fill = CALC_FILL
    
    for c, w in enumerate([4,20,12,14,12,14,12,14], 1): col_w(ws, c, w)
    
    # ========== BEPAARI_AAKDA ==========
    ws = wb.create_sheet("Bepaari_Aakda", 6)
    ws.sheet_properties.tabColor = "795548"
    ws.cell(row=1, column=1, value="BEPAARI SETTLEMENT SLIP (AAKDA)").font = TITLE_FONT
    ws.cell(row=3, column=1, value="SELECT BEPAARI:")
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ak_dv = DataValidation(type="list", formula1='BEPAARI', allow_blank=False)
    ws.add_data_validation(ak_dv)
    ak_dv.add('B3')
    aakda = [["", ""], ["A. SALES", ""], ["Qty Sold", '=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$E:$E)'], ["Gross Sales", '=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$G:$G)'], ["", ""], ["B. DEDUCTIONS", ""], ["Commission %", '=IFERROR(INDEX(Masters!$C:$C,MATCH($B$3,Masters!$B:$B,0)),4)'], ["Commission Amt", '=B9*B12/100'], ["KK (Fixed)", '=IF(B8>0,Masters!$X$3,0)'], ["JB", '=B8*Masters!$X$4'], ["Motor", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"MOTOR",Cash_Book!$E:$E,$B$3)'], ["Bhussa", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"BHUSSA",Cash_Book!$E:$E,$B$3)'], ["Gawali", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"GAWALI",Cash_Book!$E:$E,$B$3)'], ["Cash/Adv", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"CASH_ADV",Cash_Book!$E:$E,$B$3)'], ["TOTAL DEDUCTIONS", '=SUM(B13:B19)'], ["", ""], ["C. CALCULATION", ""], ["Gross Sales", '=B9'], ["(-) Deductions", '=B20'], ["(+) Opening Bal", '=IFERROR(INDEX(Masters!$D:$D,MATCH($B$3,Masters!$B:$B,0)),0)'], ["NET PAYABLE", '=B23-B24+B25'], ["", ""], ["D. PAYMENTS", ""], ["Payments Made", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,$B$3)'], ["", ""], ["BALANCE DUE", '=B26-B29']]
    for r, (lbl, frm) in enumerate(aakda, 6):
        ws.cell(row=r, column=1, value=lbl)
        if frm:
            c = ws.cell(row=r, column=2, value=frm)
            c.number_format = MONEY_FMT if "Qty" not in lbl and "%" not in lbl else NUM_FMT
        if any(x in lbl for x in ["A.", "B.", "C.", "D.", "TOTAL", "NET", "BALANCE"]):
            ws.cell(row=r, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl or "BALANCE" in lbl:
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = CALC_FILL
    col_w(ws, 1, 22)
    col_w(ws, 2, 14)
    
    # ========== BALANCE_SHEET - With Excess Payments ==========
    ws = wb.create_sheet("Balance_Sheet", 7)
    ws.sheet_properties.tabColor = "607D8B"
    ws.cell(row=1, column=1, value="BALANCE SHEET").font = TITLE_FONT
    ws.cell(row=1, column=5, value="LIABILITIES = ASSETS").font = Font(italic=True, size=10, color="FF0000")
    
    # LIABILITIES
    ws.cell(row=3, column=1, value="LIABILITIES (We Owe)").font = SUBTITLE_FONT
    ws.cell(row=3, column=1).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    ws.cell(row=4, column=1, value="Particulars")
    ws.cell(row=4, column=2, value="Amount")
    hdr(ws, 4, 1, 2, HEADER_RED)
    
    # CAPITAL - All parties
    ws.cell(row=5, column=1, value="CAPITAL:").font = Font(bold=True)
    for i in range(1, 32):
        r = 5 + i
        pr = i + 2
        ws.cell(row=r, column=1, value=f'=IF(AND(Masters!$O${pr}<>"",Masters!$Q${pr}="CAPITAL"),"  "&Masters!$O${pr},"")')
        ws.cell(row=r, column=2, value=f'=IF(A{r}="","",Masters!$P${pr}+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"CAPITAL",Cash_Book!$D:$D,"TAKEN",Cash_Book!$E:$E,Masters!$O${pr})-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"CAPITAL",Cash_Book!$D:$D,"WITHDRAWN",Cash_Book!$E:$E,Masters!$O${pr}))')
        ws.cell(row=r, column=2).number_format = MONEY_FMT
    
    cap_total_row = 37
    ws.cell(row=cap_total_row, column=1, value="  Capital Total").font = Font(bold=True)
    ws.cell(row=cap_total_row, column=2, value='=SUMIF(A6:A36,"<>",B6:B36)')
    ws.cell(row=cap_total_row, column=2).number_format = MONEY_FMT
    ws.cell(row=cap_total_row, column=2).fill = CALC_FILL
    
    # LOANS/AMANAT - All parties
    loan_start = 39
    ws.cell(row=loan_start, column=1, value="LOANS/AMANAT:").font = Font(bold=True)
    for i in range(1, 32):
        r = loan_start + i
        pr = i + 2
        ws.cell(row=r, column=1, value=f'=IF(AND(Masters!$O${pr}<>"",OR(Masters!$Q${pr}="LOAN",Masters!$Q${pr}="AMANAT")),"  "&Masters!$O${pr},"")')
        ws.cell(row=r, column=2, value=f'=IF(A{r}="","",Masters!$P${pr}+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"LOAN",Cash_Book!$D:$D,"TAKEN",Cash_Book!$E:$E,Masters!$O${pr})-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"LOAN",Cash_Book!$D:$D,"REPAID",Cash_Book!$E:$E,Masters!$O${pr}))')
        ws.cell(row=r, column=2).number_format = MONEY_FMT
    
    loan_total_row = 71
    ws.cell(row=loan_total_row, column=1, value="  Loans Total").font = Font(bold=True)
    ws.cell(row=loan_total_row, column=2, value=f'=SUMIF(A{loan_start+1}:A{loan_start+31},"<>",B{loan_start+1}:B{loan_start+31})')
    ws.cell(row=loan_total_row, column=2).number_format = MONEY_FMT
    ws.cell(row=loan_total_row, column=2).fill = CALC_FILL
    
    # Other Liabilities
    other_liab_start = 73
    ws.cell(row=other_liab_start, column=1, value="BEPAARI PAYABLES:").font = Font(bold=True)
    ws.cell(row=other_liab_start+1, column=1, value="  Net Payable to Bepaaris")
    ws.cell(row=other_liab_start+1, column=2, value='=Bepaari_Ledger!R85')  # Only positive balances
    ws.cell(row=other_liab_start+1, column=2).number_format = MONEY_FMT
    
    # NEW: Dukandar Overpaid (Advance from Dukandar)
    ws.cell(row=other_liab_start+3, column=1, value="DUKANDAR ADVANCES:").font = Font(bold=True)
    ws.cell(row=other_liab_start+4, column=1, value="  Advance from Dukandars")
    ws.cell(row=other_liab_start+4, column=2, value='=Dukandar_Ledger!H85')  # Overpaid amount
    ws.cell(row=other_liab_start+4, column=2).number_format = MONEY_FMT
    
    ws.cell(row=other_liab_start+6, column=1, value="COMMISSION (Net):").font = Font(bold=True)
    ws.cell(row=other_liab_start+7, column=1, value="  Gross - Discounts")
    ws.cell(row=other_liab_start+7, column=2, value='=Bepaari_Ledger!G82-Dukandar_Ledger!E81')
    ws.cell(row=other_liab_start+7, column=2).number_format = MONEY_FMT
    
    # JB NET
    ws.cell(row=other_liab_start+9, column=1, value="JB (Net):").font = Font(bold=True)
    ws.cell(row=other_liab_start+10, column=1, value="  Collected from Bepaari")
    ws.cell(row=other_liab_start+10, column=2, value='=Bepaari_Ledger!I82')
    ws.cell(row=other_liab_start+10, column=2).number_format = MONEY_FMT
    ws.cell(row=other_liab_start+11, column=1, value="  (-) Paid to Market")
    ws.cell(row=other_liab_start+11, column=2, value='=-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"JB_PAID")')
    ws.cell(row=other_liab_start+11, column=2).number_format = MONEY_FMT
    ws.cell(row=other_liab_start+12, column=1, value="  JB Net").font = Font(bold=True)
    ws.cell(row=other_liab_start+12, column=2, value=f'=B{other_liab_start+10}+B{other_liab_start+11}')
    ws.cell(row=other_liab_start+12, column=2).number_format = MONEY_FMT
    ws.cell(row=other_liab_start+12, column=2).fill = CALC_FILL
    
    ws.cell(row=other_liab_start+14, column=1, value="KK COLLECTED:")
    ws.cell(row=other_liab_start+14, column=2, value='=Bepaari_Ledger!H82')
    ws.cell(row=other_liab_start+14, column=2).number_format = MONEY_FMT
    
    ws.cell(row=other_liab_start+16, column=1, value="ZAKAT PAYABLE:").font = Font(bold=True)
    ws.cell(row=other_liab_start+17, column=1, value="  Open + Provision - Paid")
    ws.cell(row=other_liab_start+17, column=2, value='=Masters!$X$9+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ZAKAT",Cash_Book!$D:$D,"PROVISION")-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ZAKAT",Cash_Book!$D:$D,"PAID")')
    ws.cell(row=other_liab_start+17, column=2).number_format = MONEY_FMT
    
    total_liab_row = other_liab_start + 19
    ws.cell(row=total_liab_row, column=1, value="TOTAL LIABILITIES").font = Font(bold=True)
    ws.cell(row=total_liab_row, column=2, value=f'=B{cap_total_row}+B{loan_total_row}+B{other_liab_start+1}+B{other_liab_start+4}+B{other_liab_start+7}+B{other_liab_start+12}+B{other_liab_start+14}+B{other_liab_start+17}')
    ws.cell(row=total_liab_row, column=2).number_format = MONEY_FMT
    ws.cell(row=total_liab_row, column=2).font = Font(bold=True)
    ws.cell(row=total_liab_row, column=2).fill = CALC_FILL
    
    # ASSETS
    ws.cell(row=3, column=4, value="ASSETS (We Have / Owed to Us)").font = SUBTITLE_FONT
    ws.cell(row=3, column=4).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    ws.cell(row=4, column=4, value="Particulars")
    ws.cell(row=4, column=5, value="Amount")
    hdr(ws, 4, 4, 5, HEADER_GREEN)
    
    ws.cell(row=5, column=4, value="CASH BALANCE:").font = Font(bold=True)
    ws.cell(row=5, column=5, value='=Cash_Book!J10')
    ws.cell(row=5, column=5).number_format = MONEY_FMT
    
    ws.cell(row=7, column=4, value="BANK BALANCE:").font = Font(bold=True)
    ws.cell(row=7, column=5, value='=Cash_Book!J11')
    ws.cell(row=7, column=5).number_format = MONEY_FMT
    
    ws.cell(row=9, column=4, value="PATTI (Dukandar O/S):").font = Font(bold=True)
    ws.cell(row=10, column=4, value="  Net Receivable from Dukandars")
    ws.cell(row=10, column=5, value='=Dukandar_Ledger!H84')  # Only positive balances
    ws.cell(row=10, column=5).number_format = MONEY_FMT
    
    # NEW: Bepaari Overpaid (Advance to Bepaari)
    ws.cell(row=12, column=4, value="BEPAARI ADVANCES:").font = Font(bold=True)
    ws.cell(row=13, column=4, value="  Advance to Bepaaris (Overpaid)")
    ws.cell(row=13, column=5, value='=Bepaari_Ledger!R86')  # Overpaid amount
    ws.cell(row=13, column=5).number_format = MONEY_FMT
    
    ws.cell(row=15, column=4, value="ADVANCES GIVEN:").font = Font(bold=True)
    ws.cell(row=16, column=4, value="  Net (Given - Received)")
    ws.cell(row=16, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"GIVEN")-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"RECEIVED")')
    ws.cell(row=16, column=5).number_format = MONEY_FMT
    
    ws.cell(row=18, column=4, value="EXPENSES:").font = Font(bold=True)
    ws.cell(row=19, column=4, value="  Mandi/Business")
    ws.cell(row=19, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MANDI")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"TRAVEL")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"FOOD")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"SALARY")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"OTHER")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MISC")')
    ws.cell(row=19, column=5).number_format = MONEY_FMT
    
    ws.cell(row=20, column=4, value="  BF Discount")
    ws.cell(row=20, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"BF_DISC")')
    ws.cell(row=20, column=5).number_format = MONEY_FMT
    
    ws.cell(row=total_liab_row, column=4, value="TOTAL ASSETS").font = Font(bold=True)
    ws.cell(row=total_liab_row, column=5, value='=E5+E7+E10+E13+E16+E19+E20')
    ws.cell(row=total_liab_row, column=5).number_format = MONEY_FMT
    ws.cell(row=total_liab_row, column=5).font = Font(bold=True)
    ws.cell(row=total_liab_row, column=5).fill = CALC_FILL
    
    ws.cell(row=total_liab_row+2, column=1, value="DIFFERENCE (Should be 0)").font = Font(bold=True, color="FF0000")
    ws.cell(row=total_liab_row+2, column=2, value=f'=B{total_liab_row}-E{total_liab_row}')
    ws.cell(row=total_liab_row+2, column=2).number_format = MONEY_FMT
    ws.cell(row=total_liab_row+2, column=2).font = Font(bold=True, color="FF0000")
    
    col_w(ws, 1, 28)
    col_w(ws, 2, 14)
    col_w(ws, 3, 3)
    col_w(ws, 4, 30)
    col_w(ws, 5, 14)
    
    # ========== COMMISSION_SUMMARY ==========
    ws = wb.create_sheet("Commission_Summary", 8)
    ws.sheet_properties.tabColor = "FFC107"
    ws.cell(row=1, column=1, value="COMMISSION & PROFIT SUMMARY").font = TITLE_FONT
    items = [["", ""], ["GROSS SALES", "=Bepaari_Ledger!E82"], ["TOTAL GOATS SOLD", "=Bepaari_Ledger!F82"], ["", ""], ["INCOME:", ""], ["  Commission Earned", "=Bepaari_Ledger!G82"], ["  (-) Discounts Given", "=-Dukandar_Ledger!E81"], ["  Net Commission", "=B6+B7"], ["  JB Collected", "=Bepaari_Ledger!I82"], ["  KK Collected", "=Bepaari_Ledger!H82"], ["TOTAL INCOME", "=B8+B9+B10"], ["", ""], ["EXPENSES:", ""], ["  Mandi/Business", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MANDI")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"TRAVEL")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"FOOD")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"SALARY")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"OTHER")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MISC")'], ["  BF Discount", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"BF_DISC")'], ["  JB Paid", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"JB_PAID")'], ["TOTAL EXPENSES", "=B14+B15+B16"], ["", ""], ["NET PROFIT/LOSS", "=B11-B17"]]
    for r, (lbl, frm) in enumerate(items, 2):
        ws.cell(row=r, column=1, value=lbl)
        if frm:
            c = ws.cell(row=r, column=2, value=frm)
            c.number_format = MONEY_FMT if "GOATS" not in lbl else NUM_FMT
        if any(x in lbl for x in ["GROSS", "TOTAL", "NET", "INCOME:", "EXPENSES:"]):
            ws.cell(row=r, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl:
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = CALC_FILL
    col_w(ws, 1, 26)
    col_w(ws, 2, 14)
    
    wb.save(path)
    print(f"Mandi Master V8 created: {path}")

if __name__ == "__main__":
    create_v8("/app/frontend/public/Mandi_Master_V8.xlsx")
