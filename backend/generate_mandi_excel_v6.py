"""Mandi Master V6 - Dynamic Party dropdown, JB Net in Liabilities, Individual Capital parties"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_RED = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
HEADER_GREEN = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="0D47A1")
SUBTITLE_FONT = Font(bold=True, size=11, color="1565C0")
INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
CALC_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
MONEY_FMT = '₹#,##0'
NUM_FMT = '#,##0'
DATE_FMT = 'DD-MMM-YYYY'

def col_w(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def hdr(ws, row, c1, c2, fill=None):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def create_v6(path):
    wb = Workbook()
    wb.remove(wb.active)
    
    # Named ranges for INDIRECT dropdown
    wb.defined_names.add(DefinedName("BEPAARI", attr_text="Masters!$B$3:$B$79"))
    wb.defined_names.add(DefinedName("DUKANDAR", attr_text="Masters!$I$3:$I$79"))
    wb.defined_names.add(DefinedName("CAPITAL", attr_text="Masters!$O$3:$O$24"))
    wb.defined_names.add(DefinedName("LOAN", attr_text="Masters!$O$3:$O$24"))
    wb.defined_names.add(DefinedName("ADVANCE", attr_text="Masters!$T$3:$T$24"))
    
    # ========== MASTERS ==========
    ws = wb.create_sheet("Masters", 0)
    ws.sheet_properties.tabColor = "1565C0"
    
    ws.cell(row=1, column=1, value="BEPAARI MASTER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Bepaari Name", "Comm%", "Opening Bal", "Phone", "Remarks"], 1):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 1, 6)
    bepaaris = ["ABUL HASAN", "SHARAFAT", "SHARIK", "JUNAID", "ANNU SHARIF", "GAYYUR", "ARIF KK", "KISHOR", "SHOEB BYCULLA", "MAULANA", "ISLAM D", "SALIM RAJ", "MEHMOOD MADH", "ANAS RAJ", "MOBIN CHIPLUN", "AYAZ", "JALIL", "NEERAJ", "IRFAN JOGESHWARI", "CHINTU SHUJALPUR"]
    for i, name in enumerate(bepaaris, 1):
        ws.cell(row=i+2, column=1, value=i)
        ws.cell(row=i+2, column=2, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=3, value=4).fill = INPUT_FILL
        ws.cell(row=i+2, column=4, value=0).fill = INPUT_FILL
        ws.cell(row=i+2, column=4).number_format = MONEY_FMT
    for r in range(len(bepaaris)+3, 80):
        ws.cell(row=r, column=1, value=r-2)
        for c in range(2, 7): ws.cell(row=r, column=c).fill = INPUT_FILL
    
    ws.cell(row=1, column=8, value="DUKANDAR MASTER").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Dukandar Name", "Opening Bal", "Phone", "Remarks"], 8):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 8, 12)
    dukandars = ["SAKIM", "SHARIK", "YUNUS", "MUDASSIR", "ARIF", "SHOEB BYCULLA", "ISLAM D", "MEHMOOD MADH", "MOBIN CHIPLUN", "JALIL", "IRFAN JOGESHWARI", "YUNUS DOMBIVLI", "IFTEKHAR PUNE", "LALA/ASIF", "JAFER KASHIMIRA", "CHAND", "ASIF GOREGAON", "CHINTU SHUJALPUR", "VAZIR", "AKRAM", "NADEEM BHIWANDI", "SHAKIL KON", "PARVEZ"]
    for i, name in enumerate(dukandars, 1):
        ws.cell(row=i+2, column=8, value=i)
        ws.cell(row=i+2, column=9, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=10, value=0).fill = INPUT_FILL
        ws.cell(row=i+2, column=10).number_format = MONEY_FMT
    for r in range(len(dukandars)+3, 80):
        ws.cell(row=r, column=8, value=r-2)
        for c in range(9, 13): ws.cell(row=r, column=c).fill = INPUT_FILL
    
    ws.cell(row=1, column=14, value="CAPITAL/LOAN PARTIES").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Party Name", "Opening Bal", "Type"], 14):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 14, 17)
    cap_parties = [["MHN", 3900000, "CAPITAL"], ["SHAKIL GHODEGAON", 2000000, "LOAN"]]
    for i, (name, bal, typ) in enumerate(cap_parties, 1):
        ws.cell(row=i+2, column=14, value=i)
        ws.cell(row=i+2, column=15, value=name).fill = INPUT_FILL
        ws.cell(row=i+2, column=16, value=bal).fill = INPUT_FILL
        ws.cell(row=i+2, column=16).number_format = MONEY_FMT
        ws.cell(row=i+2, column=17, value=typ).fill = INPUT_FILL
    for r in range(len(cap_parties)+3, 25):
        ws.cell(row=r, column=14, value=r-2)
        for c in range(15, 18): ws.cell(row=r, column=c).fill = INPUT_FILL
    type_dv = DataValidation(type="list", formula1='"CAPITAL,LOAN,AMANAT"', allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add('Q3:Q24')
    
    ws.cell(row=1, column=19, value="ADVANCE PARTIES").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Party Name", "Opening Bal"], 19):
        ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 19, 21)
    for r in range(3, 25):
        ws.cell(row=r, column=19, value=r-2)
        for c in range(20, 22): ws.cell(row=r, column=c).fill = INPUT_FILL
    
    ws.cell(row=1, column=23, value="SETTINGS").font = TITLE_FONT
    settings = [["Commission Rate (%)", 4], ["KK Fixed (₹/Bepari)", 100], ["JB Rate (₹/Goat)", 10], ["", ""], ["Opening Cash", 0], ["Opening Bank", 0], ["", ""], ["ZAKAT Opening Payable", 0]]
    for r, (lbl, val) in enumerate(settings, 2):
        ws.cell(row=r, column=23, value=lbl).font = Font(bold=True) if lbl else None
        c = ws.cell(row=r, column=24, value=val)
        c.fill = INPUT_FILL
        if "Cash" in str(lbl) or "Bank" in str(lbl) or "ZAKAT" in str(lbl): c.number_format = MONEY_FMT
    for c, w in {1:4, 2:20, 3:8, 4:12, 5:10, 6:10, 7:2, 8:4, 9:20, 10:12, 11:10, 12:10, 13:2, 14:4, 15:20, 16:12, 17:10, 18:2, 19:4, 20:20, 21:12, 22:2, 23:22, 24:12}.items():
        col_w(ws, c, w)
    
    # ========== DAILY_SALES ==========
    ws = wb.create_sheet("Daily_Sales", 1)
    ws.sheet_properties.tabColor = "4CAF50"
    ws.cell(row=1, column=1, value="DAILY SALES (Bepaari to Dukandar)").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Date", "Bepaari", "Dukandar", "Qty", "Rate", "Gross", "Discount", "Net"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 9)
    bep_dv = DataValidation(type="list", formula1='BEPAARI', allow_blank=True)
    duk_dv = DataValidation(type="list", formula1='DUKANDAR', allow_blank=True)
    ws.add_data_validation(bep_dv)
    ws.add_data_validation(duk_dv)
    for r in range(4, 1004):
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3).fill = INPUT_FILL
        bep_dv.add(f'C{r}')
        ws.cell(row=r, column=4).fill = INPUT_FILL
        duk_dv.add(f'D{r}')
        ws.cell(row=r, column=5).fill = INPUT_FILL
        ws.cell(row=r, column=6).fill = INPUT_FILL
        ws.cell(row=r, column=6).number_format = MONEY_FMT
        ws.cell(row=r, column=7, value=f'=IF(E{r}="","",E{r}*F{r})')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=7).fill = CALC_FILL
        ws.cell(row=r, column=8).fill = INPUT_FILL
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        ws.cell(row=r, column=9, value=f'=IF(G{r}="","",G{r}-IF(H{r}="",0,H{r}))')
        ws.cell(row=r, column=9).number_format = MONEY_FMT
        ws.cell(row=r, column=9).fill = CALC_FILL
    for c, w in enumerate([4, 11, 20, 20, 6, 10, 12, 10, 12], 1): col_w(ws, c, w)
    
    # ========== CASH_BOOK with INDIRECT Party dropdown ==========
    ws = wb.create_sheet("Cash_Book", 2)
    ws.sheet_properties.tabColor = "FF5722"
    ws.cell(row=1, column=1, value="CASH BOOK (All Transactions)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Select TYPE first, then PARTY dropdown shows relevant list (INDIRECT)").font = Font(italic=True, size=9)
    for c, h in enumerate(["Sr", "Date", "Type", "Sub-Type", "Party", "Particulars", "Cash OUT", "Bank OUT", "Cash IN", "Bank IN"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 10)
    
    main_dv = DataValidation(type="list", formula1='"BEPAARI,DUKANDAR,CAPITAL,LOAN,ADVANCE,EXPENSE,ZAKAT,OTHER"', allow_blank=True)
    sub_dv = DataValidation(type="list", formula1='"PAYMENT,MOTOR,BHUSSA,GAWALI,CASH_ADV,OTHER,RECEIPT,TAKEN,WITHDRAWN,REPAID,GIVEN,RECEIVED,MANDI,TRAVEL,FOOD,SALARY,BF_DISC,JB_PAID,PROVISION,PAID,MISC"', allow_blank=True)
    ws.add_data_validation(main_dv)
    ws.add_data_validation(sub_dv)
    
    for r in range(4, 2004):
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3).fill = INPUT_FILL
        main_dv.add(f'C{r}')
        ws.cell(row=r, column=4).fill = INPUT_FILL
        sub_dv.add(f'D{r}')
        ws.cell(row=r, column=5).fill = INPUT_FILL
        # DYNAMIC PARTY DROPDOWN using INDIRECT - shows list based on Type
        party_dv = DataValidation(type="list", formula1=f'INDIRECT($C{r})', allow_blank=True)
        ws.add_data_validation(party_dv)
        party_dv.add(f'E{r}')
        ws.cell(row=r, column=6).fill = INPUT_FILL
        for col in [7, 8, 9, 10]:
            ws.cell(row=r, column=col).fill = INPUT_FILL
            ws.cell(row=r, column=col).number_format = MONEY_FMT
    
    ws.cell(row=1, column=12, value="CASH/BANK SUMMARY").font = SUBTITLE_FONT
    tots = [["Opening Cash", "=Masters!$X$6"], ["Opening Bank", "=Masters!$X$7"], ["Total Cash OUT", "=SUM(G:G)"], ["Total Bank OUT", "=SUM(H:H)"], ["Total Cash IN", "=SUM(I:I)"], ["Total Bank IN", "=SUM(J:J)"], ["CLOSING CASH", "=L2-L4+L6"], ["CLOSING BANK", "=L3-L5+L7"]]
    for r, (lbl, frm) in enumerate(tots, 2):
        ws.cell(row=r, column=11, value=lbl)
        ws.cell(row=r, column=12, value=frm).number_format = MONEY_FMT
        if "CLOSING" in lbl:
            ws.cell(row=r, column=11).font = Font(bold=True)
            ws.cell(row=r, column=12).font = Font(bold=True)
            ws.cell(row=r, column=12).fill = CALC_FILL
    for c, w in enumerate([4, 11, 12, 12, 20, 20, 12, 12, 12, 12, 2, 14, 14], 1): col_w(ws, c, w)
    
    # ========== BEPAARI_LEDGER ==========
    ws = wb.create_sheet("Bepaari_Ledger", 3)
    ws.sheet_properties.tabColor = "9C27B0"
    ws.cell(row=1, column=1, value="BEPAARI LEDGER (Auto-calculated)").font = TITLE_FONT
    hdrs = ["Sr", "Bepaari", "Comm%", "Opening", "Gross Sales", "Qty", "Commission", "KK", "JB", "Motor", "Bhussa", "Gawali", "Cash/Adv", "Other", "Tot.Deduct", "Net Payable", "Payments", "BALANCE"]
    for c, h in enumerate(hdrs, 1): ws.cell(row=4, column=c, value=h)
    hdr(ws, 4, 1, 18)
    for r in range(5, 82):
        m = r - 2
        ws.cell(row=r, column=1, value=r-4)
        ws.cell(row=r, column=2, value=f'=IF(Masters!$B${m}="","",Masters!$B${m})')
        ws.cell(row=r, column=3, value=f'=IF(B{r}="","",Masters!$C${m})')
        ws.cell(row=r, column=4, value=f'=IF(B{r}="",0,Masters!$D${m})')
        ws.cell(row=r, column=4).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f'=SUMIF(Daily_Sales!$C:$C,B{r},Daily_Sales!$G:$G)')
        ws.cell(row=r, column=5).number_format = MONEY_FMT
        ws.cell(row=r, column=6, value=f'=SUMIF(Daily_Sales!$C:$C,B{r},Daily_Sales!$E:$E)')
        ws.cell(row=r, column=6).number_format = NUM_FMT
        ws.cell(row=r, column=7, value=f'=IF(B{r}="",0,E{r}*C{r}/100)')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=8, value=f'=IF(F{r}>0,Masters!$X$3,0)')
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        ws.cell(row=r, column=8).fill = CALC_FILL
        ws.cell(row=r, column=9, value=f'=F{r}*Masters!$X$4')
        ws.cell(row=r, column=9).number_format = MONEY_FMT
        ws.cell(row=r, column=9).fill = CALC_FILL
        for col, sub in [(10,"MOTOR"), (11,"BHUSSA"), (12,"GAWALI"), (13,"CASH_ADV"), (14,"OTHER")]:
            ws.cell(row=r, column=col, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"{sub}",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"{sub}",Cash_Book!$E:$E,B{r})')
            ws.cell(row=r, column=col).number_format = MONEY_FMT
        ws.cell(row=r, column=15, value=f'=SUM(G{r}:N{r})')
        ws.cell(row=r, column=15).number_format = MONEY_FMT
        ws.cell(row=r, column=15).fill = CALC_FILL
        ws.cell(row=r, column=16, value=f'=IF(B{r}="","",E{r}-O{r}+D{r})')
        ws.cell(row=r, column=16).number_format = MONEY_FMT
        ws.cell(row=r, column=16).fill = CALC_FILL
        ws.cell(row=r, column=17, value=f'=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=17).number_format = MONEY_FMT
        ws.cell(row=r, column=18, value=f'=IF(B{r}="","",P{r}-Q{r})')
        ws.cell(row=r, column=18).number_format = MONEY_FMT
        ws.cell(row=r, column=18).font = Font(bold=True)
    ws.cell(row=82, column=1, value="TOTAL").font = Font(bold=True)
    for c in [4,5,6,7,8,9,10,11,12,13,14,15,16,17,18]:
        ws.cell(row=82, column=c, value=f'=SUM({get_column_letter(c)}5:{get_column_letter(c)}81)')
        ws.cell(row=82, column=c).number_format = MONEY_FMT
        ws.cell(row=82, column=c).font = Font(bold=True)
    for c, w in enumerate([4,16,6,10,12,6,10,8,8,9,9,9,10,8,11,12,11,12], 1): col_w(ws, c, w)
    
    # ========== DUKANDAR_LEDGER ==========
    ws = wb.create_sheet("Dukandar_Ledger", 4)
    ws.sheet_properties.tabColor = "E91E63"
    ws.cell(row=1, column=1, value="DUKANDAR LEDGER (Auto-calculated)").font = TITLE_FONT
    for c, h in enumerate(["Sr", "Dukandar", "Opening", "Purchases", "Discounts", "Net Receivable", "Receipts", "BALANCE"], 1):
        ws.cell(row=3, column=c, value=h)
    hdr(ws, 3, 1, 8)
    for r in range(4, 81):
        m = r - 1
        ws.cell(row=r, column=1, value=r-3)
        ws.cell(row=r, column=2, value=f'=IF(Masters!$I${m}="","",Masters!$I${m})')
        ws.cell(row=r, column=3, value=f'=IF(B{r}="",0,Masters!$J${m})')
        ws.cell(row=r, column=3).number_format = MONEY_FMT
        ws.cell(row=r, column=4, value=f'=SUMIF(Daily_Sales!$D:$D,B{r},Daily_Sales!$G:$G)')
        ws.cell(row=r, column=4).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f'=SUMIF(Daily_Sales!$D:$D,B{r},Daily_Sales!$H:$H)')
        ws.cell(row=r, column=5).number_format = MONEY_FMT
        ws.cell(row=r, column=6, value=f'=IF(B{r}="","",D{r}-E{r}+C{r})')
        ws.cell(row=r, column=6).number_format = MONEY_FMT
        ws.cell(row=r, column=6).fill = CALC_FILL
        ws.cell(row=r, column=7, value=f'=SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"DUKANDAR",Cash_Book!$D:$D,"RECEIPT",Cash_Book!$E:$E,B{r})+SUMIFS(Cash_Book!$J:$J,Cash_Book!$C:$C,"DUKANDAR",Cash_Book!$D:$D,"RECEIPT",Cash_Book!$E:$E,B{r})')
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        ws.cell(row=r, column=8, value=f'=IF(B{r}="","",F{r}-G{r})')
        ws.cell(row=r, column=8).number_format = MONEY_FMT
        ws.cell(row=r, column=8).font = Font(bold=True)
    ws.cell(row=81, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=81, column=2, value="PATTI (Total O/S)")
    for c in [3,4,5,6,7,8]:
        ws.cell(row=81, column=c, value=f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}80)')
        ws.cell(row=81, column=c).number_format = MONEY_FMT
        ws.cell(row=81, column=c).font = Font(bold=True)
    for c, w in enumerate([4,20,12,14,12,14,12,14], 1): col_w(ws, c, w)
    
    # ========== BEPAARI_AAKDA ==========
    ws = wb.create_sheet("Bepaari_Aakda", 5)
    ws.sheet_properties.tabColor = "795548"
    ws.cell(row=1, column=1, value="BEPAARI SETTLEMENT SLIP (AAKDA)").font = TITLE_FONT
    ws.cell(row=3, column=1, value="SELECT BEPAARI:")
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ak_dv = DataValidation(type="list", formula1='BEPAARI', allow_blank=False)
    ws.add_data_validation(ak_dv)
    ak_dv.add('B3')
    aakda = [["", ""], ["A. SALES", ""], ["Qty Sold", '=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$E:$E)'], ["Gross Sales", '=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$G:$G)'], ["", ""], ["B. DEDUCTIONS", ""], ["Commission %", '=IFERROR(INDEX(Masters!$C:$C,MATCH($B$3,Masters!$B:$B,0)),4)'], ["Commission Amt", '=B9*B12/100'], ["KK (Fixed)", '=IF(B8>0,Masters!$X$3,0)'], ["JB", '=B8*Masters!$X$4'], ["Motor", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"MOTOR",Cash_Book!$E:$E,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"MOTOR",Cash_Book!$E:$E,$B$3)'], ["Bhussa", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"BHUSSA",Cash_Book!$E:$E,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"BHUSSA",Cash_Book!$E:$E,$B$3)'], ["Gawali", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"GAWALI",Cash_Book!$E:$E,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"GAWALI",Cash_Book!$E:$E,$B$3)'], ["Cash/Adv", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"CASH_ADV",Cash_Book!$E:$E,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"CASH_ADV",Cash_Book!$E:$E,$B$3)'], ["TOTAL DEDUCTIONS", '=SUM(B13:B19)'], ["", ""], ["C. CALCULATION", ""], ["Gross Sales", '=B9'], ["(-) Deductions", '=B20'], ["(+) Opening Bal", '=IFERROR(INDEX(Masters!$D:$D,MATCH($B$3,Masters!$B:$B,0)),0)'], ["NET PAYABLE", '=B23-B24+B25'], ["", ""], ["D. PAYMENTS", ""], ["Payments Made", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,$B$3)+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"BEPAARI",Cash_Book!$D:$D,"PAYMENT",Cash_Book!$E:$E,$B$3)'], ["", ""], ["BALANCE DUE", '=B26-B29']]
    for r, (lbl, frm) in enumerate(aakda, 6):
        ws.cell(row=r, column=1, value=lbl)
        if frm:
            c = ws.cell(row=r, column=2, value=frm)
            c.number_format = MONEY_FMT if "Qty" not in lbl and "%" not in lbl else NUM_FMT
        if any(x in lbl for x in ["A.", "B.", "C.", "D.", "TOTAL", "NET", "BALANCE"]):
            ws.cell(row=r, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl or "BALANCE" in lbl:
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = CALC_FILL
    col_w(ws, 1, 22)
    col_w(ws, 2, 14)
    
    # ========== BALANCE_SHEET (Individual Capital, JB Net in Liabilities) ==========
    ws = wb.create_sheet("Balance_Sheet", 6)
    ws.sheet_properties.tabColor = "607D8B"
    ws.cell(row=1, column=1, value="BALANCE SHEET").font = TITLE_FONT
    ws.cell(row=1, column=5, value="LIABILITIES = ASSETS").font = Font(italic=True, size=10, color="FF0000")
    
    # LIABILITIES
    ws.cell(row=3, column=1, value="LIABILITIES (We Owe)").font = SUBTITLE_FONT
    ws.cell(row=3, column=1).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    ws.cell(row=4, column=1, value="Particulars")
    ws.cell(row=4, column=2, value="Amount")
    hdr(ws, 4, 1, 2, HEADER_RED)
    
    # CAPITAL - Each party separately
    ws.cell(row=5, column=1, value="CAPITAL:").font = Font(bold=True)
    for i in range(1, 11):
        r = 5 + i
        pr = i + 2
        ws.cell(row=r, column=1, value=f'=IF(AND(Masters!$O${pr}<>"",Masters!$Q${pr}="CAPITAL"),"  "&Masters!$O${pr},"")')
        ws.cell(row=r, column=2, value=f'=IF(A{r}="","",Masters!$P${pr}+SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"CAPITAL",Cash_Book!$D:$D,"TAKEN",Cash_Book!$E:$E,Masters!$O${pr})+SUMIFS(Cash_Book!$J:$J,Cash_Book!$C:$C,"CAPITAL",Cash_Book!$D:$D,"TAKEN",Cash_Book!$E:$E,Masters!$O${pr})-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"CAPITAL",Cash_Book!$D:$D,"WITHDRAWN",Cash_Book!$E:$E,Masters!$O${pr})-SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"CAPITAL",Cash_Book!$D:$D,"WITHDRAWN",Cash_Book!$E:$E,Masters!$O${pr}))')
        ws.cell(row=r, column=2).number_format = MONEY_FMT
    ws.cell(row=16, column=1, value="  Capital Total").font = Font(bold=True)
    ws.cell(row=16, column=2, value='=SUM(B6:B15)')
    ws.cell(row=16, column=2).number_format = MONEY_FMT
    ws.cell(row=16, column=2).fill = CALC_FILL
    
    # LOANS - Each party separately
    ws.cell(row=18, column=1, value="LOANS/AMANAT:").font = Font(bold=True)
    for i in range(1, 11):
        r = 18 + i
        pr = i + 2
        ws.cell(row=r, column=1, value=f'=IF(AND(Masters!$O${pr}<>"",OR(Masters!$Q${pr}="LOAN",Masters!$Q${pr}="AMANAT")),"  "&Masters!$O${pr},"")')
        ws.cell(row=r, column=2, value=f'=IF(A{r}="","",Masters!$P${pr}+SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"LOAN",Cash_Book!$D:$D,"TAKEN",Cash_Book!$E:$E,Masters!$O${pr})+SUMIFS(Cash_Book!$J:$J,Cash_Book!$C:$C,"LOAN",Cash_Book!$D:$D,"TAKEN",Cash_Book!$E:$E,Masters!$O${pr})-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"LOAN",Cash_Book!$D:$D,"REPAID",Cash_Book!$E:$E,Masters!$O${pr})-SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"LOAN",Cash_Book!$D:$D,"REPAID",Cash_Book!$E:$E,Masters!$O${pr}))')
        ws.cell(row=r, column=2).number_format = MONEY_FMT
    ws.cell(row=29, column=1, value="  Loans Total").font = Font(bold=True)
    ws.cell(row=29, column=2, value='=SUM(B19:B28)')
    ws.cell(row=29, column=2).number_format = MONEY_FMT
    ws.cell(row=29, column=2).fill = CALC_FILL
    
    ws.cell(row=31, column=1, value="BEPAARI PAYABLES:").font = Font(bold=True)
    ws.cell(row=32, column=1, value="  From Ledger")
    ws.cell(row=32, column=2, value='=Bepaari_Ledger!R82')
    ws.cell(row=32, column=2).number_format = MONEY_FMT
    
    ws.cell(row=34, column=1, value="COMMISSION (Net):").font = Font(bold=True)
    ws.cell(row=35, column=1, value="  Gross - Discounts")
    ws.cell(row=35, column=2, value='=Bepaari_Ledger!G82-Dukandar_Ledger!E81')
    ws.cell(row=35, column=2).number_format = MONEY_FMT
    
    # JB NET = Collected - Paid (in LIABILITIES, not Assets)
    ws.cell(row=37, column=1, value="JB (Net):").font = Font(bold=True)
    ws.cell(row=38, column=1, value="  Collected from Bepaari")
    ws.cell(row=38, column=2, value='=Bepaari_Ledger!I82')
    ws.cell(row=38, column=2).number_format = MONEY_FMT
    ws.cell(row=39, column=1, value="  (-) Paid to Market")
    ws.cell(row=39, column=2, value='=-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"JB_PAID")-SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"JB_PAID")')
    ws.cell(row=39, column=2).number_format = MONEY_FMT
    ws.cell(row=40, column=1, value="  JB Net").font = Font(bold=True)
    ws.cell(row=40, column=2, value='=B38+B39')
    ws.cell(row=40, column=2).number_format = MONEY_FMT
    ws.cell(row=40, column=2).fill = CALC_FILL
    
    ws.cell(row=42, column=1, value="KK COLLECTED:")
    ws.cell(row=42, column=2, value='=Bepaari_Ledger!H82')
    ws.cell(row=42, column=2).number_format = MONEY_FMT
    
    ws.cell(row=44, column=1, value="ZAKAT PAYABLE:").font = Font(bold=True)
    ws.cell(row=45, column=1, value="  Open + Provision - Paid")
    ws.cell(row=45, column=2, value='=Masters!$X$9+SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"ZAKAT",Cash_Book!$D:$D,"PROVISION")+SUMIFS(Cash_Book!$J:$J,Cash_Book!$C:$C,"ZAKAT",Cash_Book!$D:$D,"PROVISION")-SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ZAKAT",Cash_Book!$D:$D,"PAID")-SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"ZAKAT",Cash_Book!$D:$D,"PAID")')
    ws.cell(row=45, column=2).number_format = MONEY_FMT
    
    ws.cell(row=47, column=1, value="TOTAL LIABILITIES").font = Font(bold=True)
    ws.cell(row=47, column=2, value='=B16+B29+B32+B35+B40+B42+B45')
    ws.cell(row=47, column=2).number_format = MONEY_FMT
    ws.cell(row=47, column=2).font = Font(bold=True)
    ws.cell(row=47, column=2).fill = CALC_FILL
    
    # ASSETS (NO JB_PAID here - moved to Liabilities)
    ws.cell(row=3, column=4, value="ASSETS (We Have / Owed to Us)").font = SUBTITLE_FONT
    ws.cell(row=3, column=4).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    ws.cell(row=4, column=4, value="Particulars")
    ws.cell(row=4, column=5, value="Amount")
    hdr(ws, 4, 4, 5, HEADER_GREEN)
    
    ws.cell(row=5, column=4, value="CASH BALANCE:").font = Font(bold=True)
    ws.cell(row=5, column=5, value='=Cash_Book!L8')
    ws.cell(row=5, column=5).number_format = MONEY_FMT
    
    ws.cell(row=7, column=4, value="BANK BALANCE:").font = Font(bold=True)
    ws.cell(row=7, column=5, value='=Cash_Book!L9')
    ws.cell(row=7, column=5).number_format = MONEY_FMT
    
    ws.cell(row=9, column=4, value="PATTI (Dukandar O/S):").font = Font(bold=True)
    ws.cell(row=10, column=4, value="  From Ledger")
    ws.cell(row=10, column=5, value='=Dukandar_Ledger!H81')
    ws.cell(row=10, column=5).number_format = MONEY_FMT
    
    ws.cell(row=12, column=4, value="ADVANCES GIVEN:").font = Font(bold=True)
    ws.cell(row=13, column=4, value="  Net (Given - Received)")
    ws.cell(row=13, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"GIVEN")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"GIVEN")-SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"RECEIVED")-SUMIFS(Cash_Book!$J:$J,Cash_Book!$C:$C,"ADVANCE",Cash_Book!$D:$D,"RECEIVED")')
    ws.cell(row=13, column=5).number_format = MONEY_FMT
    
    ws.cell(row=15, column=4, value="EXPENSES:").font = Font(bold=True)
    ws.cell(row=16, column=4, value="  Mandi/Business")
    ws.cell(row=16, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MANDI")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MANDI")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"TRAVEL")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"TRAVEL")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"FOOD")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"FOOD")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"SALARY")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"SALARY")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"OTHER")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"OTHER")')
    ws.cell(row=16, column=5).number_format = MONEY_FMT
    
    ws.cell(row=17, column=4, value="  BF Discount")
    ws.cell(row=17, column=5, value='=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"BF_DISC")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"BF_DISC")')
    ws.cell(row=17, column=5).number_format = MONEY_FMT
    
    # NOTE: JB_PAID removed from Assets - now in Liabilities as deduction
    
    ws.cell(row=47, column=4, value="TOTAL ASSETS").font = Font(bold=True)
    ws.cell(row=47, column=5, value='=E5+E7+E10+E13+E16+E17')
    ws.cell(row=47, column=5).number_format = MONEY_FMT
    ws.cell(row=47, column=5).font = Font(bold=True)
    ws.cell(row=47, column=5).fill = CALC_FILL
    
    ws.cell(row=49, column=1, value="DIFFERENCE (Should be 0)").font = Font(bold=True, color="FF0000")
    ws.cell(row=49, column=2, value='=B47-E47')
    ws.cell(row=49, column=2).number_format = MONEY_FMT
    ws.cell(row=49, column=2).font = Font(bold=True, color="FF0000")
    
    col_w(ws, 1, 26)
    col_w(ws, 2, 14)
    col_w(ws, 3, 3)
    col_w(ws, 4, 26)
    col_w(ws, 5, 14)
    
    # ========== COMMISSION_SUMMARY ==========
    ws = wb.create_sheet("Commission_Summary", 7)
    ws.sheet_properties.tabColor = "FFC107"
    ws.cell(row=1, column=1, value="COMMISSION & PROFIT SUMMARY").font = TITLE_FONT
    items = [["", ""], ["GROSS SALES", "=Bepaari_Ledger!E82"], ["TOTAL GOATS SOLD", "=Bepaari_Ledger!F82"], ["", ""], ["INCOME:", ""], ["  Commission Earned", "=Bepaari_Ledger!G82"], ["  (-) Discounts Given", "=-Dukandar_Ledger!E81"], ["  Net Commission", "=B6+B7"], ["  JB Collected", "=Bepaari_Ledger!I82"], ["  KK Collected", "=Bepaari_Ledger!H82"], ["TOTAL INCOME", "=B8+B9+B10"], ["", ""], ["EXPENSES:", ""], ["  Mandi/Business", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MANDI")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"MANDI")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"TRAVEL")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"TRAVEL")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"FOOD")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"FOOD")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"SALARY")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"SALARY")+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"OTHER")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"OTHER")'], ["  BF Discount", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"BF_DISC")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"BF_DISC")'], ["  JB Paid", '=SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"JB_PAID")+SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"EXPENSE",Cash_Book!$D:$D,"JB_PAID")'], ["TOTAL EXPENSES", "=B14+B15+B16"], ["", ""], ["NET PROFIT/LOSS", "=B11-B17"]]
    for r, (lbl, frm) in enumerate(items, 2):
        ws.cell(row=r, column=1, value=lbl)
        if frm:
            c = ws.cell(row=r, column=2, value=frm)
            c.number_format = MONEY_FMT if "GOATS" not in lbl else NUM_FMT
        if any(x in lbl for x in ["GROSS", "TOTAL", "NET", "INCOME:", "EXPENSES:"]):
            ws.cell(row=r, column=1).font = Font(bold=True)
            if "TOTAL" in lbl or "NET" in lbl:
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = CALC_FILL
    col_w(ws, 1, 26)
    col_w(ws, 2, 14)
    
    wb.save(path)
    print(f"Mandi Master V6 created: {path}")

if __name__ == "__main__":
    create_v6("/app/Mandi_Master_V6.xlsx")
