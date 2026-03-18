"""
Mandi Master Excel Generator - VERSION 2.0
Complete redesign based on actual paper-based accounting system
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from datetime import datetime
import os

# Style definitions
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")  # Blue
SECTION_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Light blue
TITLE_FONT = Font(bold=True, size=14, color="0D47A1")
SUBTITLE_FONT = Font(bold=True, size=11, color="1565C0")
INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")  # Light yellow
CALC_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Light green
MONEY_FORMAT = '₹#,##0'
NUMBER_FORMAT = '#,##0'
DATE_FORMAT = 'DD-MMM-YYYY'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def style_header(ws, row, start_col, end_col, fill=None):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def add_dropdown(ws, cell_range, formula, error_msg="Invalid entry"):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = error_msg
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv

def create_mandi_excel_v2(output_path):
    wb = Workbook()
    wb.remove(wb.active)
    
    # ================================================================
    # SHEET 1: MASTERS
    # ================================================================
    ws_masters = wb.create_sheet("Masters", 0)
    ws_masters.sheet_properties.tabColor = "1565C0"
    
    # ----- BEPAARI MASTER (Column A-G) -----
    ws_masters.cell(row=1, column=1, value="BEPAARI MASTER").font = TITLE_FONT
    bepaari_headers = ["Sr.", "Bepaari Name", "Comm. Type", "Comm. Value", "Opening Bal.", "Phone", "Remarks"]
    for col, header in enumerate(bepaari_headers, 1):
        ws_masters.cell(row=2, column=col, value=header)
    style_header(ws_masters, 2, 1, 7)
    
    # Sample bepaaris
    sample_bepaaris = [
        [1, "ABUL HASAN", "%", 4, 41470, "", ""],
        [2, "SHARAFAT", "%", 4, 200000, "", ""],
        [3, "JUNAID", "%", 4, 75040, "", ""],
        [4, "CHINTU", "%", 4, 200000, "", ""],
        [5, "GAYYUR", "%", 4, 250000, "", ""],
        [6, "NEERAJ", "Per Piece", 200, 59056, "", "₹200/goat"],
        [7, "ANAS", "%", 4, 150000, "", ""],
        [8, "ANNU SHARIF", "%", 4, 63586, "", ""],
    ]
    for row_idx, data in enumerate(sample_bepaaris, 3):
        for col_idx, value in enumerate(data, 1):
            cell = ws_masters.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = INPUT_FILL
            if col_idx == 5:
                cell.number_format = MONEY_FORMAT
    
    # Empty rows for more bepaaris
    for row_idx in range(11, 60):
        ws_masters.cell(row=row_idx, column=1, value=row_idx - 2)
        for col in range(2, 8):
            ws_masters.cell(row=row_idx, column=col).fill = INPUT_FILL
    
    # Commission type dropdown
    comm_dv = DataValidation(type="list", formula1='"%, Per Piece"', allow_blank=True)
    ws_masters.add_data_validation(comm_dv)
    comm_dv.add('C3:C59')
    
    # ----- DUKANDAR MASTER (Column I-M) -----
    ws_masters.cell(row=1, column=9, value="DUKANDAR MASTER").font = TITLE_FONT
    duk_headers = ["Sr.", "Dukandar Name", "Opening Bal.", "Phone", "Remarks"]
    for col, header in enumerate(duk_headers, 9):
        ws_masters.cell(row=2, column=col, value=header)
    style_header(ws_masters, 2, 9, 13)
    
    sample_dukandars = [
        [1, "IRFAN JOG", 0, "", ""],
        [2, "SALIM", 0, "", ""],
        [3, "YUNUS", 0, "", ""],
        [4, "SHERIF", 0, "", ""],
        [5, "MUDASSIR", 0, "", ""],
    ]
    for row_idx, data in enumerate(sample_dukandars, 3):
        for col_idx, value in enumerate(data, 9):
            cell = ws_masters.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = INPUT_FILL
            if col_idx == 11:
                cell.number_format = MONEY_FORMAT
    
    for row_idx in range(8, 60):
        ws_masters.cell(row=row_idx, column=9, value=row_idx - 2)
        for col in range(10, 14):
            ws_masters.cell(row=row_idx, column=col).fill = INPUT_FILL
    
    # ----- LOAN/AMANAT PARTIES (Column O-R) -----
    ws_masters.cell(row=1, column=15, value="LOAN/AMANAT PARTIES").font = TITLE_FONT
    loan_headers = ["Sr.", "Party Name", "Opening Bal.", "Remarks"]
    for col, header in enumerate(loan_headers, 15):
        ws_masters.cell(row=2, column=col, value=header)
    style_header(ws_masters, 2, 15, 18)
    
    sample_loans = [
        [1, "SHAKIL GHODEGAON", 500000, "Amanat taken"],
    ]
    for row_idx, data in enumerate(sample_loans, 3):
        for col_idx, value in enumerate(data, 15):
            cell = ws_masters.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = INPUT_FILL
            if col_idx == 17:
                cell.number_format = MONEY_FORMAT
    
    for row_idx in range(5, 20):
        ws_masters.cell(row=row_idx, column=15, value=row_idx - 2)
        for col in range(16, 19):
            ws_masters.cell(row=row_idx, column=col).fill = INPUT_FILL
    
    # ----- FIXED VALUES & OPENING BALANCES (Column T-V) -----
    ws_masters.cell(row=1, column=20, value="FIXED VALUES & BALANCES").font = TITLE_FONT
    
    fixed_items = [
        ["COMMISSION RATE (%)", 4],
        ["KARKUNI (KK) - Fixed ₹/Bepari", 100],
        ["JAGAH BHADA (JB) - ₹/Goat", 10],
        ["", ""],
        ["OPENING CASH BALANCE", 453915],
        ["OPENING BANK BALANCE", 218345],
        ["CAPITAL ACCOUNT", 3819981],
    ]
    for row_idx, (label, value) in enumerate(fixed_items, 2):
        ws_masters.cell(row=row_idx, column=20, value=label).font = Font(bold=True) if label else None
        cell = ws_masters.cell(row=row_idx, column=21, value=value)
        cell.fill = INPUT_FILL
        if "BALANCE" in label or "CAPITAL" in label:
            cell.number_format = MONEY_FORMAT
    
    # ----- TRANSACTION TYPES (Column T onwards, row 12+) -----
    ws_masters.cell(row=11, column=20, value="TRANSACTION TYPES").font = TITLE_FONT
    trans_headers = ["Code", "Description"]
    for col, header in enumerate(trans_headers, 20):
        ws_masters.cell(row=12, column=col, value=header)
    style_header(ws_masters, 12, 20, 21)
    
    trans_types = [
        ["BEP_PAY", "Bepaari Payment"],
        ["DUK_REC", "Dukandar Receipt"],
        ["EXP_MOTOR", "Expense - Motor/Transport"],
        ["EXP_CONV", "Expense - Conveyance/OLA"],
        ["EXP_FOOD", "Expense - Food"],
        ["EXP_TRAVEL", "Expense - Traveling"],
        ["EXP_SALARY", "Expense - Salary"],
        ["EXP_OTHER", "Expense - Other"],
        ["LOAN_TAKEN", "Loan/Amanat Taken"],
        ["LOAN_REPAY", "Loan/Amanat Repaid"],
        ["ADV_GIVEN", "Advance Given to Bepaari"],
        ["ADV_RECV", "Advance Recovered"],
        ["ZAKAT", "Zakat Paid"],
        ["BF_DISC", "B/F - Cash Discount on Collection"],
        ["JB_PAID", "Jagah Bhada Paid (Collected from Beparis)"],
        ["CAPITAL", "Capital Contribution/Withdrawal"],
    ]
    for row_idx, data in enumerate(trans_types, 13):
        for col_idx, value in enumerate(data, 20):
            ws_masters.cell(row=row_idx, column=col_idx, value=value)
    
    # Column widths for Masters
    widths = {1:5, 2:22, 3:12, 4:12, 5:15, 6:12, 7:15, 8:2, 
              9:5, 10:22, 11:15, 12:12, 13:15, 14:2,
              15:5, 16:22, 17:15, 18:15, 19:2,
              20:28, 21:18}
    for col, width in widths.items():
        set_col_width(ws_masters, col, width)
    
    # ================================================================
    # SHEET 2: DAILY_SALES
    # ================================================================
    ws_sales = wb.create_sheet("Daily_Sales", 1)
    ws_sales.sheet_properties.tabColor = "4CAF50"
    
    ws_sales.cell(row=1, column=1, value="DAILY SALES ENTRY (Bepaari → Dukandar)").font = TITLE_FONT
    ws_sales.cell(row=2, column=1, value="Enter each goat sale transaction. Formulas auto-calculate totals.").font = Font(italic=True, size=9)
    
    sales_headers = [
        "Sr.", "Date", "Bepaari", "Dukandar", "Qty (Goats)", 
        "Rate (₹)", "Gross Total", "Discount", "Net Amount", "Remarks"
    ]
    for col, header in enumerate(sales_headers, 1):
        ws_sales.cell(row=3, column=col, value=header)
    style_header(ws_sales, 3, 1, 10)
    
    # Create named range for bepaari dropdown
    # Formula references Masters!B3:B59 for bepaari names
    bepaari_dv = DataValidation(type="list", formula1='Masters!$B$3:$B$59', allow_blank=True)
    bepaari_dv.error = "Select Bepaari from list"
    ws_sales.add_data_validation(bepaari_dv)
    
    dukandar_dv = DataValidation(type="list", formula1='Masters!$J$3:$J$59', allow_blank=True)
    dukandar_dv.error = "Select Dukandar from list"
    ws_sales.add_data_validation(dukandar_dv)
    
    for row in range(4, 504):
        ws_sales.cell(row=row, column=1, value=row-3)
        
        # Date
        ws_sales.cell(row=row, column=2).number_format = DATE_FORMAT
        ws_sales.cell(row=row, column=2).fill = INPUT_FILL
        
        # Bepaari dropdown
        ws_sales.cell(row=row, column=3).fill = INPUT_FILL
        bepaari_dv.add(f'C{row}')
        
        # Dukandar dropdown
        ws_sales.cell(row=row, column=4).fill = INPUT_FILL
        dukandar_dv.add(f'D{row}')
        
        # Qty
        ws_sales.cell(row=row, column=5).fill = INPUT_FILL
        
        # Rate
        ws_sales.cell(row=row, column=6).fill = INPUT_FILL
        ws_sales.cell(row=row, column=6).number_format = MONEY_FORMAT
        
        # Gross Total = Qty × Rate
        ws_sales.cell(row=row, column=7, value=f'=IF(E{row}="","",E{row}*F{row})')
        ws_sales.cell(row=row, column=7).number_format = MONEY_FORMAT
        ws_sales.cell(row=row, column=7).fill = CALC_FILL
        
        # Discount (manual entry)
        ws_sales.cell(row=row, column=8).fill = INPUT_FILL
        ws_sales.cell(row=row, column=8).number_format = MONEY_FORMAT
        
        # Net Amount = Gross - Discount
        ws_sales.cell(row=row, column=9, value=f'=IF(G{row}="","",G{row}-IF(H{row}="",0,H{row}))')
        ws_sales.cell(row=row, column=9).number_format = MONEY_FORMAT
        ws_sales.cell(row=row, column=9).fill = CALC_FILL
        
        # Remarks
        ws_sales.cell(row=row, column=10).fill = INPUT_FILL
    
    # Column widths
    sales_widths = [5, 12, 22, 22, 12, 12, 15, 12, 15, 20]
    for col, width in enumerate(sales_widths, 1):
        set_col_width(ws_sales, col, width)
    
    # ================================================================
    # SHEET 3: CASH_BOOK (Single sheet for ALL transactions)
    # ================================================================
    ws_cash = wb.create_sheet("Cash_Book", 2)
    ws_cash.sheet_properties.tabColor = "FF5722"
    
    ws_cash.cell(row=1, column=1, value="DAILY CASH & BANK BOOK").font = TITLE_FONT
    ws_cash.cell(row=2, column=1, value="Record ALL payments, receipts, expenses, loans in ONE place").font = Font(italic=True, size=9)
    
    cash_headers = [
        "Sr.", "Date", "Trans. Type", "Party/Description", "Particulars",
        "Cash OUT", "Bank OUT", "Cash IN", "Bank IN", "Remarks"
    ]
    for col, header in enumerate(cash_headers, 1):
        ws_cash.cell(row=3, column=col, value=header)
    style_header(ws_cash, 3, 1, 10)
    
    # Transaction type dropdown
    trans_type_list = "BEP_PAY,DUK_REC,EXP_MOTOR,EXP_CONV,EXP_FOOD,EXP_TRAVEL,EXP_SALARY,EXP_OTHER,LOAN_TAKEN,LOAN_REPAY,ADV_GIVEN,ADV_RECV,ZAKAT,BF_DISC,JB_PAID,CAPITAL"
    trans_dv = DataValidation(type="list", formula1=f'"{trans_type_list}"', allow_blank=True)
    trans_dv.error = "Select transaction type"
    ws_cash.add_data_validation(trans_dv)
    
    for row in range(4, 1004):  # 1000 transaction rows
        ws_cash.cell(row=row, column=1, value=row-3)
        
        # Date
        ws_cash.cell(row=row, column=2).number_format = DATE_FORMAT
        ws_cash.cell(row=row, column=2).fill = INPUT_FILL
        
        # Transaction Type dropdown
        ws_cash.cell(row=row, column=3).fill = INPUT_FILL
        trans_dv.add(f'C{row}')
        
        # Party/Description (will have conditional dropdown based on trans type)
        ws_cash.cell(row=row, column=4).fill = INPUT_FILL
        
        # Particulars
        ws_cash.cell(row=row, column=5).fill = INPUT_FILL
        
        # Cash OUT
        ws_cash.cell(row=row, column=6).fill = INPUT_FILL
        ws_cash.cell(row=row, column=6).number_format = MONEY_FORMAT
        
        # Bank OUT
        ws_cash.cell(row=row, column=7).fill = INPUT_FILL
        ws_cash.cell(row=row, column=7).number_format = MONEY_FORMAT
        
        # Cash IN
        ws_cash.cell(row=row, column=8).fill = INPUT_FILL
        ws_cash.cell(row=row, column=8).number_format = MONEY_FORMAT
        
        # Bank IN
        ws_cash.cell(row=row, column=9).fill = INPUT_FILL
        ws_cash.cell(row=row, column=9).number_format = MONEY_FORMAT
        
        # Remarks
        ws_cash.cell(row=row, column=10).fill = INPUT_FILL
    
    # Summary section at top-right
    ws_cash.cell(row=1, column=12, value="RUNNING TOTALS").font = SUBTITLE_FONT
    summary_items = [
        ["Opening Cash", "=Masters!$U$6"],
        ["Opening Bank", "=Masters!$U$7"],
        ["Total Cash OUT", "=SUM(F:F)"],
        ["Total Bank OUT", "=SUM(G:G)"],
        ["Total Cash IN", "=SUM(H:H)"],
        ["Total Bank IN", "=SUM(I:I)"],
        ["Closing Cash", "=L2-L4+L6"],
        ["Closing Bank", "=L3-L5+L7"],
    ]
    for row_idx, (label, formula) in enumerate(summary_items, 2):
        ws_cash.cell(row=row_idx, column=11, value=label)
        ws_cash.cell(row=row_idx, column=12, value=formula)
        ws_cash.cell(row=row_idx, column=12).number_format = MONEY_FORMAT
        if "Closing" in label:
            ws_cash.cell(row=row_idx, column=11).font = Font(bold=True)
            ws_cash.cell(row=row_idx, column=12).font = Font(bold=True)
    
    # Column widths
    cash_widths = [5, 12, 14, 25, 20, 14, 14, 14, 14, 20, 2, 15, 15]
    for col, width in enumerate(cash_widths, 1):
        set_col_width(ws_cash, col, width)
    
    # ================================================================
    # SHEET 4: BEPAARI_LEDGER
    # ================================================================
    ws_bep_led = wb.create_sheet("Bepaari_Ledger", 3)
    ws_bep_led.sheet_properties.tabColor = "9C27B0"
    
    ws_bep_led.cell(row=1, column=1, value="BEPAARI LEDGER - Account Summary").font = TITLE_FONT
    ws_bep_led.cell(row=2, column=1, value="Auto-calculated from Daily_Sales, Cash_Book. KK=₹100/Bepari, JB=₹10/Goat").font = Font(italic=True, size=9)
    
    led_headers = [
        "Sr.", "Bepaari Name", "Comm.Type", "Comm.Val", "Opening",
        "Gross Sales", "Qty Sold", "Commission", "KK (₹100)", "JB (₹10/goat)",
        "Motor", "Bhussa", "Gawali", "Cash/Adv", "Other Exp",
        "Total Deductions", "Net Payable", "Payments", "Balance Due"
    ]
    for col, header in enumerate(led_headers, 1):
        ws_bep_led.cell(row=4, column=col, value=header)
    style_header(ws_bep_led, 4, 1, 19)
    
    for row in range(5, 62):  # 57 bepaaris max
        master_row = row - 2
        bep_name_ref = f"Masters!$B${master_row}"
        
        ws_bep_led.cell(row=row, column=1, value=row-4)
        
        # Bepaari Name
        ws_bep_led.cell(row=row, column=2, value=f'=IF({bep_name_ref}="","",{bep_name_ref})')
        
        # Commission Type
        ws_bep_led.cell(row=row, column=3, value=f'=IF(B{row}="","",Masters!$C${master_row})')
        
        # Commission Value
        ws_bep_led.cell(row=row, column=4, value=f'=IF(B{row}="","",Masters!$D${master_row})')
        
        # Opening Balance
        ws_bep_led.cell(row=row, column=5, value=f'=IF(B{row}="",0,Masters!$E${master_row})')
        ws_bep_led.cell(row=row, column=5).number_format = MONEY_FORMAT
        
        # Gross Sales (SUMIF from Daily_Sales)
        ws_bep_led.cell(row=row, column=6, value=f'=SUMIF(Daily_Sales!$C:$C,B{row},Daily_Sales!$G:$G)')
        ws_bep_led.cell(row=row, column=6).number_format = MONEY_FORMAT
        
        # Qty Sold
        ws_bep_led.cell(row=row, column=7, value=f'=SUMIF(Daily_Sales!$C:$C,B{row},Daily_Sales!$E:$E)')
        ws_bep_led.cell(row=row, column=7).number_format = NUMBER_FORMAT
        
        # Commission (% or Per Piece)
        ws_bep_led.cell(row=row, column=8, value=f'=IF(B{row}="","",IF(C{row}="%",F{row}*D{row}/100,G{row}*D{row}))')
        ws_bep_led.cell(row=row, column=8).number_format = MONEY_FORMAT
        
        # KK - Fixed ₹100 per Bepaari (if they have sales)
        ws_bep_led.cell(row=row, column=9, value=f'=IF(G{row}>0,Masters!$U$3,0)')
        ws_bep_led.cell(row=row, column=9).number_format = MONEY_FORMAT
        
        # JB - ₹10 per goat
        ws_bep_led.cell(row=row, column=10, value=f'=G{row}*Masters!$U$4')
        ws_bep_led.cell(row=row, column=10).number_format = MONEY_FORMAT
        
        # Motor (from Cash_Book where type=EXP_MOTOR and party matches)
        ws_bep_led.cell(row=row, column=11, value=f'=SUMIFS(Cash_Book!$F:$F,Cash_Book!$C:$C,"EXP_MOTOR",Cash_Book!$D:$D,B{row})+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXP_MOTOR",Cash_Book!$D:$D,B{row})')
        ws_bep_led.cell(row=row, column=11).number_format = MONEY_FORMAT
        
        # Bhussa - placeholder (can add specific expense type)
        ws_bep_led.cell(row=row, column=12, value=0)
        ws_bep_led.cell(row=row, column=12).number_format = MONEY_FORMAT
        
        # Gawali - placeholder
        ws_bep_led.cell(row=row, column=13, value=0)
        ws_bep_led.cell(row=row, column=13).number_format = MONEY_FORMAT
        
        # Cash/Advance given (ADV_GIVEN)
        ws_bep_led.cell(row=row, column=14, value=f'=SUMIFS(Cash_Book!$F:$F,Cash_Book!$C:$C,"ADV_GIVEN",Cash_Book!$D:$D,B{row})+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ADV_GIVEN",Cash_Book!$D:$D,B{row})')
        ws_bep_led.cell(row=row, column=14).number_format = MONEY_FORMAT
        
        # Other Expenses
        ws_bep_led.cell(row=row, column=15, value=0)
        ws_bep_led.cell(row=row, column=15).number_format = MONEY_FORMAT
        
        # Total Deductions = Commission + KK + JB + Motor + Bhussa + Gawali + Cash + Other
        ws_bep_led.cell(row=row, column=16, value=f'=SUM(H{row}:O{row})')
        ws_bep_led.cell(row=row, column=16).number_format = MONEY_FORMAT
        ws_bep_led.cell(row=row, column=16).fill = CALC_FILL
        
        # Net Payable = Gross Sales - Total Deductions + Opening Balance
        ws_bep_led.cell(row=row, column=17, value=f'=IF(B{row}="","",F{row}-P{row}+E{row})')
        ws_bep_led.cell(row=row, column=17).number_format = MONEY_FORMAT
        ws_bep_led.cell(row=row, column=17).fill = CALC_FILL
        
        # Payments Made (BEP_PAY from Cash_Book)
        ws_bep_led.cell(row=row, column=18, value=f'=SUMIFS(Cash_Book!$F:$F,Cash_Book!$C:$C,"BEP_PAY",Cash_Book!$D:$D,B{row})+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEP_PAY",Cash_Book!$D:$D,B{row})')
        ws_bep_led.cell(row=row, column=18).number_format = MONEY_FORMAT
        
        # Balance Due = Net Payable - Payments
        ws_bep_led.cell(row=row, column=19, value=f'=IF(B{row}="","",Q{row}-R{row})')
        ws_bep_led.cell(row=row, column=19).number_format = MONEY_FORMAT
        ws_bep_led.cell(row=row, column=19).font = Font(bold=True)
    
    # Totals row
    total_row = 62
    ws_bep_led.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]:
        ws_bep_led.cell(row=total_row, column=col, value=f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}61)')
        ws_bep_led.cell(row=total_row, column=col).number_format = MONEY_FORMAT
        ws_bep_led.cell(row=total_row, column=col).font = Font(bold=True)
    
    # Column widths
    led_widths = [4, 18, 8, 8, 12, 14, 10, 12, 10, 12, 10, 10, 10, 12, 10, 14, 14, 14, 14]
    for col, width in enumerate(led_widths, 1):
        set_col_width(ws_bep_led, col, width)
    
    # ================================================================
    # SHEET 5: DUKANDAR_LEDGER
    # ================================================================
    ws_duk_led = wb.create_sheet("Dukandar_Ledger", 4)
    ws_duk_led.sheet_properties.tabColor = "E91E63"
    
    ws_duk_led.cell(row=1, column=1, value="DUKANDAR LEDGER - Account Summary").font = TITLE_FONT
    
    duk_led_headers = [
        "Sr.", "Dukandar Name", "Opening", "Total Purchases", 
        "Discounts Given", "Net Receivable", "Receipts", "Balance Due"
    ]
    for col, header in enumerate(duk_led_headers, 1):
        ws_duk_led.cell(row=3, column=col, value=header)
    style_header(ws_duk_led, 3, 1, 8)
    
    for row in range(4, 61):
        master_row = row - 1
        duk_name_ref = f"Masters!$J${master_row}"
        
        ws_duk_led.cell(row=row, column=1, value=row-3)
        
        # Dukandar Name
        ws_duk_led.cell(row=row, column=2, value=f'=IF({duk_name_ref}="","",{duk_name_ref})')
        
        # Opening Balance
        ws_duk_led.cell(row=row, column=3, value=f'=IF(B{row}="",0,Masters!$K${master_row})')
        ws_duk_led.cell(row=row, column=3).number_format = MONEY_FORMAT
        
        # Total Purchases (SUMIF Net Amount from Daily_Sales)
        ws_duk_led.cell(row=row, column=4, value=f'=SUMIF(Daily_Sales!$D:$D,B{row},Daily_Sales!$G:$G)')
        ws_duk_led.cell(row=row, column=4).number_format = MONEY_FORMAT
        
        # Discounts Given
        ws_duk_led.cell(row=row, column=5, value=f'=SUMIF(Daily_Sales!$D:$D,B{row},Daily_Sales!$H:$H)')
        ws_duk_led.cell(row=row, column=5).number_format = MONEY_FORMAT
        
        # Net Receivable = Purchases - Discounts + Opening
        ws_duk_led.cell(row=row, column=6, value=f'=IF(B{row}="","",D{row}-E{row}+C{row})')
        ws_duk_led.cell(row=row, column=6).number_format = MONEY_FORMAT
        ws_duk_led.cell(row=row, column=6).fill = CALC_FILL
        
        # Receipts (DUK_REC from Cash_Book)
        ws_duk_led.cell(row=row, column=7, value=f'=SUMIFS(Cash_Book!$H:$H,Cash_Book!$C:$C,"DUK_REC",Cash_Book!$D:$D,B{row})+SUMIFS(Cash_Book!$I:$I,Cash_Book!$C:$C,"DUK_REC",Cash_Book!$D:$D,B{row})')
        ws_duk_led.cell(row=row, column=7).number_format = MONEY_FORMAT
        
        # Balance Due = Net Receivable - Receipts
        ws_duk_led.cell(row=row, column=8, value=f'=IF(B{row}="","",F{row}-G{row})')
        ws_duk_led.cell(row=row, column=8).number_format = MONEY_FORMAT
        ws_duk_led.cell(row=row, column=8).font = Font(bold=True)
    
    # Totals
    total_row = 61
    ws_duk_led.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in [3, 4, 5, 6, 7, 8]:
        ws_duk_led.cell(row=total_row, column=col, value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}60)')
        ws_duk_led.cell(row=total_row, column=col).number_format = MONEY_FORMAT
        ws_duk_led.cell(row=total_row, column=col).font = Font(bold=True)
    
    duk_widths = [5, 22, 14, 16, 14, 16, 14, 16]
    for col, width in enumerate(duk_widths, 1):
        set_col_width(ws_duk_led, col, width)
    
    # ================================================================
    # SHEET 6: BEPAARI_AAKDA (Settlement Slip)
    # ================================================================
    ws_aakda = wb.create_sheet("Bepaari_Aakda", 5)
    ws_aakda.sheet_properties.tabColor = "795548"
    
    ws_aakda.cell(row=1, column=1, value="BEPAARI SETTLEMENT SLIP (AAKDA)").font = TITLE_FONT
    
    ws_aakda.cell(row=3, column=1, value="SELECT BEPAARI:")
    ws_aakda.cell(row=3, column=2).fill = INPUT_FILL
    # Dropdown for bepaari selection
    bepaari_aakda_dv = DataValidation(type="list", formula1='Masters!$B$3:$B$59', allow_blank=False)
    ws_aakda.add_data_validation(bepaari_aakda_dv)
    bepaari_aakda_dv.add('B3')
    
    ws_aakda.cell(row=4, column=1, value="DATE:")
    ws_aakda.cell(row=4, column=2).fill = INPUT_FILL
    ws_aakda.cell(row=4, column=2).number_format = DATE_FORMAT
    
    # Settlement details
    ws_aakda.cell(row=6, column=1, value="SETTLEMENT DETAILS").font = SUBTITLE_FONT
    
    aakda_rows = [
        ["", ""],
        ["A. SALES SUMMARY", ""],
        ["Total Qty (Goats)", f'=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$E:$E)'],
        ["Gross Sales Value", f'=SUMIF(Daily_Sales!$C:$C,$B$3,Daily_Sales!$G:$G)'],
        ["", ""],
        ["B. DEDUCTIONS", ""],
        ["Commission Type", f'=IFERROR(VLOOKUP($B$3,Masters!$B:$D,2,FALSE),"")'],
        ["Commission Rate/Value", f'=IFERROR(VLOOKUP($B$3,Masters!$B:$D,3,FALSE),"")'],
        ["Commission Amount", f'=IF(B13="%",B10*B14/100,B9*B14)'],
        ["Karkuni (KK) - Fixed", f'=IF(B9>0,Masters!$U$3,0)'],
        ["Jagah Bhada (JB) @ ₹10/goat", f'=B9*Masters!$U$4'],
        ["Motor Bhada", f'=SUMIFS(Cash_Book!$F:$F,Cash_Book!$C:$C,"EXP_MOTOR",Cash_Book!$D:$D,$B$3)+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"EXP_MOTOR",Cash_Book!$D:$D,$B$3)'],
        ["Cash/Advance Given", f'=SUMIFS(Cash_Book!$F:$F,Cash_Book!$C:$C,"ADV_GIVEN",Cash_Book!$D:$D,$B$3)+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"ADV_GIVEN",Cash_Book!$D:$D,$B$3)'],
        ["Other Expenses", 0],
        ["TOTAL DEDUCTIONS", f'=SUM(B15:B20)'],
        ["", ""],
        ["C. NET CALCULATION", ""],
        ["Gross Sales", f'=B10'],
        ["Less: Total Deductions", f'=B21'],
        ["Add: Opening Balance", f'=IFERROR(VLOOKUP($B$3,Masters!$B:$E,4,FALSE),0)'],
        ["NET PAYABLE TO BEPAARI", f'=B24-B25+B26'],
        ["", ""],
        ["D. PAYMENTS MADE", ""],
        ["Total Payments", f'=SUMIFS(Cash_Book!$F:$F,Cash_Book!$C:$C,"BEP_PAY",Cash_Book!$D:$D,$B$3)+SUMIFS(Cash_Book!$G:$G,Cash_Book!$C:$C,"BEP_PAY",Cash_Book!$D:$D,$B$3)'],
        ["", ""],
        ["BALANCE DUE", f'=B27-B30'],
    ]
    
    for row_idx, (label, formula) in enumerate(aakda_rows, 7):
        ws_aakda.cell(row=row_idx, column=1, value=label)
        if formula:
            ws_aakda.cell(row=row_idx, column=2, value=formula)
            ws_aakda.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        
        if any(x in label for x in ["A.", "B.", "C.", "D.", "TOTAL", "NET PAYABLE", "BALANCE"]):
            ws_aakda.cell(row=row_idx, column=1).font = Font(bold=True)
            if "TOTAL" in label or "NET" in label or "BALANCE" in label:
                ws_aakda.cell(row=row_idx, column=2).font = Font(bold=True)
                ws_aakda.cell(row=row_idx, column=2).fill = CALC_FILL
    
    set_col_width(ws_aakda, 1, 28)
    set_col_width(ws_aakda, 2, 18)
    
    # ================================================================
    # SHEET 7: DAILY_BALANCE_SHEET
    # ================================================================
    ws_bs = wb.create_sheet("Daily_Balance_Sheet", 6)
    ws_bs.sheet_properties.tabColor = "607D8B"
    
    ws_bs.cell(row=1, column=1, value="DAILY BALANCE SHEET").font = TITLE_FONT
    ws_bs.cell(row=1, column=6, value="ASSETS & LIABILITIES - Both Sides Must Tally").font = Font(italic=True, size=10)
    
    # LEFT SIDE - LIABILITIES (What we owe)
    ws_bs.cell(row=3, column=1, value="LIABILITIES (Jama - We Owe)").font = SUBTITLE_FONT
    ws_bs.cell(row=3, column=1).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    ws_bs.merge_cells('A3:B3')
    
    liab_headers = ["Particulars", "Amount"]
    for col, header in enumerate(liab_headers, 1):
        ws_bs.cell(row=4, column=col, value=header)
    style_header(ws_bs, 4, 1, 2, fill=PatternFill(start_color="C62828", end_color="C62828", fill_type="solid"))
    
    liab_items = [
        ["CAPITAL", "=Masters!$U$8"],
        ["Loans/Amanat:", ""],
        ["  SHAKIL GHODEGAON", "=IFERROR(VLOOKUP(\"SHAKIL GHODEGAON\",Masters!$P:$R,2,FALSE),0)"],
        ["", ""],
        ["BEPAARI PAYABLES:", ""],
    ]
    
    row_idx = 5
    for label, formula in liab_items:
        ws_bs.cell(row=row_idx, column=1, value=label)
        if formula:
            ws_bs.cell(row=row_idx, column=2, value=formula)
            ws_bs.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        if ":" in label or label == "CAPITAL":
            ws_bs.cell(row=row_idx, column=1).font = Font(bold=True)
        row_idx += 1
    
    # List bepaari balances (dynamic)
    for i in range(1, 51):  # 50 bepaaris
        bepaari_row = i + 4
        ws_bs.cell(row=row_idx, column=1, value=f'=IF(Bepaari_Ledger!B{bepaari_row}="","",Bepaari_Ledger!B{bepaari_row})')
        ws_bs.cell(row=row_idx, column=2, value=f'=IF(Bepaari_Ledger!B{bepaari_row}="","",Bepaari_Ledger!S{bepaari_row})')
        ws_bs.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        row_idx += 1
    
    # Total Liabilities
    liab_total_row = row_idx
    ws_bs.cell(row=liab_total_row, column=1, value="TOTAL LIABILITIES").font = Font(bold=True)
    ws_bs.cell(row=liab_total_row, column=2, value=f'=SUM(B5:B{liab_total_row-1})')
    ws_bs.cell(row=liab_total_row, column=2).number_format = MONEY_FORMAT
    ws_bs.cell(row=liab_total_row, column=2).font = Font(bold=True)
    ws_bs.cell(row=liab_total_row, column=2).fill = CALC_FILL
    
    # RIGHT SIDE - ASSETS (What we have / What others owe us)
    ws_bs.cell(row=3, column=4, value="ASSETS (Udhaar - We Are Owed)").font = SUBTITLE_FONT
    ws_bs.cell(row=3, column=4).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    ws_bs.merge_cells('D3:E3')
    
    asset_headers = ["Particulars", "Amount"]
    for col, header in enumerate(asset_headers, 4):
        ws_bs.cell(row=4, column=col, value=header)
    style_header(ws_bs, 4, 4, 5, fill=PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"))
    
    asset_items = [
        ["CASH BALANCE", "=Cash_Book!L8"],
        ["BANK BALANCE", "=Cash_Book!L9"],
        ["", ""],
        ["DUKANDAR RECEIVABLES:", ""],
    ]
    
    row_idx = 5
    for label, formula in asset_items:
        ws_bs.cell(row=row_idx, column=4, value=label)
        if formula:
            ws_bs.cell(row=row_idx, column=5, value=formula)
            ws_bs.cell(row=row_idx, column=5).number_format = MONEY_FORMAT
        if ":" in label or "BALANCE" in label:
            ws_bs.cell(row=row_idx, column=4).font = Font(bold=True)
        row_idx += 1
    
    # List dukandar balances
    for i in range(1, 51):
        duk_row = i + 3
        ws_bs.cell(row=row_idx, column=4, value=f'=IF(Dukandar_Ledger!B{duk_row}="","",Dukandar_Ledger!B{duk_row})')
        ws_bs.cell(row=row_idx, column=5, value=f'=IF(Dukandar_Ledger!B{duk_row}="","",Dukandar_Ledger!H{duk_row})')
        ws_bs.cell(row=row_idx, column=5).number_format = MONEY_FORMAT
        row_idx += 1
    
    # Other assets
    other_assets = [
        ["", ""],
        ["Commission Earned", "=Bepaari_Ledger!H62"],
        ["Less: Discounts Given", "=Dukandar_Ledger!E61"],
        ["Net Commission", "=E57-E58"],
        ["", ""],
        ["SUSPENSE", 0],
        ["B/F (Brought Forward)", 0],
        ["", ""],
    ]
    for label, formula in other_assets:
        ws_bs.cell(row=row_idx, column=4, value=label)
        if formula:
            ws_bs.cell(row=row_idx, column=5, value=formula if isinstance(formula, str) and formula.startswith("=") else formula)
            ws_bs.cell(row=row_idx, column=5).number_format = MONEY_FORMAT
        if "Net" in str(label):
            ws_bs.cell(row=row_idx, column=4).font = Font(bold=True)
            ws_bs.cell(row=row_idx, column=5).font = Font(bold=True)
        row_idx += 1
    
    # Total Assets
    asset_total_row = liab_total_row  # Same row as liabilities total for visual alignment
    ws_bs.cell(row=asset_total_row, column=4, value="TOTAL ASSETS").font = Font(bold=True)
    ws_bs.cell(row=asset_total_row, column=5, value=f'=SUM(E5:E{asset_total_row-1})')
    ws_bs.cell(row=asset_total_row, column=5).number_format = MONEY_FORMAT
    ws_bs.cell(row=asset_total_row, column=5).font = Font(bold=True)
    ws_bs.cell(row=asset_total_row, column=5).fill = CALC_FILL
    
    # Tally check
    ws_bs.cell(row=asset_total_row + 2, column=1, value="DIFFERENCE (Should be 0)").font = Font(bold=True, color="FF0000")
    ws_bs.cell(row=asset_total_row + 2, column=2, value=f'=B{liab_total_row}-E{asset_total_row}')
    ws_bs.cell(row=asset_total_row + 2, column=2).number_format = MONEY_FORMAT
    ws_bs.cell(row=asset_total_row + 2, column=2).font = Font(bold=True, color="FF0000")
    
    # Column widths
    set_col_width(ws_bs, 1, 25)
    set_col_width(ws_bs, 2, 15)
    set_col_width(ws_bs, 3, 3)
    set_col_width(ws_bs, 4, 25)
    set_col_width(ws_bs, 5, 15)
    
    # ================================================================
    # SHEET 8: COMMISSION_SUMMARY
    # ================================================================
    ws_comm = wb.create_sheet("Commission_Summary", 7)
    ws_comm.sheet_properties.tabColor = "FFC107"
    
    ws_comm.cell(row=1, column=1, value="COMMISSION SUMMARY").font = TITLE_FONT
    
    comm_items = [
        ["", ""],
        ["GROSS SALES (All Bepaaris)", "=Bepaari_Ledger!F62"],
        ["Commission Rate", "=Masters!$U$2 & \"%\""],
        ["", ""],
        ["COMMISSION EARNED (Gross)", "=Bepaari_Ledger!H62"],
        ["Less: Discounts to Dukandars", "=Dukandar_Ledger!E61"],
        ["", ""],
        ["NET COMMISSION", "=B5-B6"],
        ["", ""],
        ["EXPENSES:", ""],
        ["  Motor/Transport", "=SUMIF(Cash_Book!$C:$C,\"EXP_MOTOR\",Cash_Book!$F:$F)+SUMIF(Cash_Book!$C:$C,\"EXP_MOTOR\",Cash_Book!$G:$G)"],
        ["  Conveyance/OLA", "=SUMIF(Cash_Book!$C:$C,\"EXP_CONV\",Cash_Book!$F:$F)+SUMIF(Cash_Book!$C:$C,\"EXP_CONV\",Cash_Book!$G:$G)"],
        ["  Food", "=SUMIF(Cash_Book!$C:$C,\"EXP_FOOD\",Cash_Book!$F:$F)+SUMIF(Cash_Book!$C:$C,\"EXP_FOOD\",Cash_Book!$G:$G)"],
        ["  Traveling", "=SUMIF(Cash_Book!$C:$C,\"EXP_TRAVEL\",Cash_Book!$F:$F)+SUMIF(Cash_Book!$C:$C,\"EXP_TRAVEL\",Cash_Book!$7:$G)"],
        ["  Salary", "=SUMIF(Cash_Book!$C:$C,\"EXP_SALARY\",Cash_Book!$F:$F)+SUMIF(Cash_Book!$C:$C,\"EXP_SALARY\",Cash_Book!$G:$G)"],
        ["  Other", "=SUMIF(Cash_Book!$C:$C,\"EXP_OTHER\",Cash_Book!$F:$F)+SUMIF(Cash_Book!$C:$C,\"EXP_OTHER\",Cash_Book!$G:$G)"],
        ["  Zakat", "=SUMIF(Cash_Book!$C:$C,\"ZAKAT\",Cash_Book!$F:$F)+SUMIF(Cash_Book!$C:$C,\"ZAKAT\",Cash_Book!$G:$G)"],
        ["TOTAL EXPENSES", "=SUM(B11:B17)"],
        ["", ""],
        ["NET PROFIT/LOSS", "=B8-B18"],
    ]
    
    for row_idx, (label, formula) in enumerate(comm_items, 2):
        ws_comm.cell(row=row_idx, column=1, value=label)
        if formula:
            ws_comm.cell(row=row_idx, column=2, value=formula)
            ws_comm.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        
        if "NET" in label or "TOTAL" in label or "GROSS" in label:
            ws_comm.cell(row=row_idx, column=1).font = Font(bold=True)
            ws_comm.cell(row=row_idx, column=2).font = Font(bold=True)
            if "NET" in label:
                ws_comm.cell(row=row_idx, column=2).fill = CALC_FILL
    
    set_col_width(ws_comm, 1, 30)
    set_col_width(ws_comm, 2, 18)
    
    # ================================================================
    # SAVE
    # ================================================================
    wb.save(output_path)
    print(f"✅ Mandi Master V2 Excel created: {output_path}")
    return output_path


if __name__ == "__main__":
    output_file = "/app/Mandi_Master_V2.xlsx"
    create_mandi_excel_v2(output_file)
