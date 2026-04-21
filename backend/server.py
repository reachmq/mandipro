from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from typing import List, Optional
from datetime import datetime
from models import (
    Bepaari, Dukandar, AdvanceParty, CapitalPartner, Settings,
    DailySale, CashBookEntry, DailySaleCreate, CashBookEntryCreate, MasterCreate,
    AdjustmentEntry, AdjustmentEntryCreate
)
import uuid
import io
import csv

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Mandi Accounting System")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def serialize_doc(doc):
    if doc and "_id" in doc:
        del doc["_id"]
    return doc

def serialize_docs(docs):
    return [serialize_doc(d) for d in docs]


# ============== SETTINGS ==============
@api_router.get("/settings")
async def get_settings():
    settings = await db.settings.find_one({"id": "settings"})
    if not settings:
        default = Settings().model_dump()
        await db.settings.insert_one(default)
        return default
    return serialize_doc(settings)

@api_router.put("/settings")
async def update_settings(updates: dict):
    updates["updated_at"] = datetime.utcnow().isoformat()
    await db.settings.update_one({"id": "settings"}, {"$set": updates}, upsert=True)
    return await get_settings()


# ============== BEPAARI ==============
@api_router.get("/bepaaris")
async def get_bepaaris():
    bepaaris = await db.bepaaris.find({"is_active": True}).to_list(500)
    return serialize_docs(bepaaris)

@api_router.post("/bepaaris")
async def create_bepaari(data: MasterCreate):
    bepaari = Bepaari(
        name=data.name,
        opening_balance=data.opening_balance,
        commission_percent=data.commission_percent or 4.0,
        flat_rate_per_goat=data.flat_rate,
        phone=data.phone
    )
    doc = bepaari.model_dump()
    doc["created_at"] = datetime.utcnow().isoformat()
    await db.bepaaris.insert_one(doc)
    return {"id": bepaari.id, "name": bepaari.name}

@api_router.put("/bepaaris/{bepaari_id}")
async def update_bepaari(bepaari_id: str, data: dict):
    await db.bepaaris.update_one({"id": bepaari_id}, {"$set": data})
    return {"status": "updated"}

@api_router.delete("/bepaaris/{bepaari_id}")
async def delete_bepaari(bepaari_id: str):
    await db.bepaaris.update_one({"id": bepaari_id}, {"$set": {"is_active": False}})
    return {"status": "deleted"}


# ============== DUKANDAR ==============
@api_router.get("/dukandars")
async def get_dukandars():
    dukandars = await db.dukandars.find({"is_active": True}).to_list(500)
    return serialize_docs(dukandars)

@api_router.post("/dukandars")
async def create_dukandar(data: MasterCreate):
    dukandar = Dukandar(
        name=data.name,
        opening_balance=data.opening_balance,
        phone=data.phone
    )
    doc = dukandar.model_dump()
    doc["created_at"] = datetime.utcnow().isoformat()
    await db.dukandars.insert_one(doc)
    return {"id": dukandar.id, "name": dukandar.name}

@api_router.put("/dukandars/{dukandar_id}")
async def update_dukandar(dukandar_id: str, data: dict):
    await db.dukandars.update_one({"id": dukandar_id}, {"$set": data})
    return {"status": "updated"}

@api_router.delete("/dukandars/{dukandar_id}")
async def delete_dukandar(dukandar_id: str):
    await db.dukandars.update_one({"id": dukandar_id}, {"$set": {"is_active": False}})
    return {"status": "deleted"}


# ============== ADVANCE PARTIES ==============
@api_router.get("/advance-parties")
async def get_advance_parties():
    parties = await db.advance_parties.find({"is_active": True}).to_list(100)
    return serialize_docs(parties)

@api_router.post("/advance-parties")
async def create_advance_party(data: MasterCreate):
    party = AdvanceParty(name=data.name, opening_balance=data.opening_balance)
    doc = party.model_dump()
    doc["created_at"] = datetime.utcnow().isoformat()
    await db.advance_parties.insert_one(doc)
    return {"id": party.id, "name": party.name}

@api_router.put("/advance-parties/{party_id}")
async def update_advance_party(party_id: str, data: dict):
    update_fields = {}
    if "name" in data:
        update_fields["name"] = data["name"]
    if "opening_balance" in data:
        update_fields["opening_balance"] = data["opening_balance"]
    await db.advance_parties.update_one({"id": party_id}, {"$set": update_fields})
    return {"status": "updated"}

@api_router.delete("/advance-parties/{party_id}")
async def delete_advance_party(party_id: str):
    await db.advance_parties.update_one({"id": party_id}, {"$set": {"is_active": False}})
    return {"status": "deleted"}


# ============== CAPITAL PARTNERS ==============
@api_router.get("/capital-partners")
async def get_capital_partners():
    partners = await db.capital_partners.find({"is_active": True}).to_list(100)
    return serialize_docs(partners)

@api_router.post("/capital-partners")
async def create_capital_partner(data: MasterCreate):
    partner = CapitalPartner(
        name=data.name,
        partner_type=data.partner_type or "CAPITAL",
        opening_balance=data.opening_balance
    )
    doc = partner.model_dump()
    doc["created_at"] = datetime.utcnow().isoformat()
    await db.capital_partners.insert_one(doc)
    return {"id": partner.id, "name": partner.name}

@api_router.put("/capital-partners/{partner_id}")
async def update_capital_partner(partner_id: str, data: dict):
    update_fields = {}
    if "name" in data:
        update_fields["name"] = data["name"]
    if "opening_balance" in data:
        update_fields["opening_balance"] = data["opening_balance"]
    if "partner_type" in data:
        update_fields["partner_type"] = data["partner_type"]
    await db.capital_partners.update_one({"id": partner_id}, {"$set": update_fields})
    return {"status": "updated"}

@api_router.delete("/capital-partners/{partner_id}")
async def delete_capital_partner(partner_id: str):
    await db.capital_partners.update_one({"id": partner_id}, {"$set": {"is_active": False}})
    return {"status": "deleted"}


# ============== DAILY SALES ==============
@api_router.get("/daily-sales")
async def get_daily_sales(
    date: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    bepaari_id: Optional[str] = None,
    dukandar_id: Optional[str] = None
):
    query = {}
    if date:
        query["date"] = date
    elif from_date and to_date:
        query["date"] = {"$gte": from_date, "$lte": to_date}
    elif from_date:
        query["date"] = {"$gte": from_date}
    elif to_date:
        query["date"] = {"$lte": to_date}
    
    if bepaari_id:
        query["bepaari_id"] = bepaari_id
    if dukandar_id:
        query["dukandar_id"] = dukandar_id
    
    sales = await db.daily_sales.find(query).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    return serialize_docs(sales)

@api_router.post("/daily-sales")
async def create_daily_sale(data: DailySaleCreate):
    bepaari = await db.bepaaris.find_one({"id": data.bepaari_id})
    dukandar = await db.dukandars.find_one({"id": data.dukandar_id})
    
    if not bepaari or not dukandar:
        raise HTTPException(status_code=400, detail="Invalid bepaari or dukandar ID")
    
    gross = data.quantity * data.rate
    net = gross - data.discount
    
    dukandar_rate = data.dukandar_rate if data.dukandar_rate and data.dukandar_rate != data.rate else None
    dukandar_amount = (data.quantity * dukandar_rate) if dukandar_rate else None
    
    sale = DailySale(
        date=data.date,
        bepaari_id=data.bepaari_id,
        bepaari_name=bepaari["name"],
        dukandar_id=data.dukandar_id,
        dukandar_name=dukandar["name"],
        quantity=data.quantity,
        rate=data.rate,
        gross_amount=gross,
        discount=data.discount,
        net_amount=net,
        dukandar_rate=dukandar_rate,
        dukandar_amount=dukandar_amount
    )
    doc = sale.model_dump()
    doc["created_at"] = datetime.utcnow().isoformat()
    await db.daily_sales.insert_one(doc)
    return serialize_doc(doc)

@api_router.delete("/daily-sales/{sale_id}")
async def delete_daily_sale(sale_id: str):
    await db.daily_sales.delete_one({"id": sale_id})
    return {"status": "deleted"}

@api_router.put("/daily-sales/{sale_id}")
async def update_daily_sale(sale_id: str, data: dict):
    update_fields = {}
    
    for field in ["date", "quantity", "rate", "discount"]:
        if field in data:
            update_fields[field] = data[field]
    
    # Recalculate amounts if qty/rate/discount/dukandar_rate changed
    if "quantity" in data or "rate" in data or "discount" in data or "dukandar_rate" in data:
        sale = await db.daily_sales.find_one({"id": sale_id})
        qty = data.get("quantity", sale.get("quantity", 0))
        rate = data.get("rate", sale.get("rate", 0))
        discount = data.get("discount", sale.get("discount", 0))
        update_fields["gross_amount"] = qty * rate
        update_fields["net_amount"] = (qty * rate) - discount
        
        # Handle dukandar_rate
        dukandar_rate = data.get("dukandar_rate", sale.get("dukandar_rate"))
        if dukandar_rate and dukandar_rate != rate:
            update_fields["dukandar_rate"] = dukandar_rate
            update_fields["dukandar_amount"] = qty * dukandar_rate
        else:
            update_fields["dukandar_rate"] = None
            update_fields["dukandar_amount"] = None
    
    # Update bepaari/dukandar if changed
    if "bepaari_id" in data:
        update_fields["bepaari_id"] = data["bepaari_id"]
        bepaari = await db.bepaaris.find_one({"id": data["bepaari_id"]})
        if bepaari:
            update_fields["bepaari_name"] = bepaari.get("name")
    
    if "dukandar_id" in data:
        update_fields["dukandar_id"] = data["dukandar_id"]
        dukandar = await db.dukandars.find_one({"id": data["dukandar_id"]})
        if dukandar:
            update_fields["dukandar_name"] = dukandar.get("name")
    
    await db.daily_sales.update_one({"id": sale_id}, {"$set": update_fields})
    return {"status": "updated"}


# ============== CASH BOOK ==============
@api_router.get("/cash-book")
async def get_cash_book(
    date: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    party_id: Optional[str] = None,
    type: Optional[str] = None,
    sub_type: Optional[str] = None
):
    query = {}
    if date:
        query["date"] = date
    elif from_date and to_date:
        query["date"] = {"$gte": from_date, "$lte": to_date}
    elif from_date:
        query["date"] = {"$gte": from_date}
    elif to_date:
        query["date"] = {"$lte": to_date}
    
    if party_id:
        query["party_id"] = party_id
    if type:
        query["type"] = type
    if sub_type:
        query["sub_type"] = sub_type
    
    entries = await db.cash_book.find(query).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    return serialize_docs(entries)

@api_router.post("/cash-book")
async def create_cash_book_entry(data: CashBookEntryCreate):
    party_name = None
    if data.party_id:
        for coll in ["bepaaris", "dukandars", "advance_parties", "capital_partners"]:
            party = await db[coll].find_one({"id": data.party_id})
            if party:
                party_name = party.get("name")
                break
    
    entry = CashBookEntry(
        date=data.date,
        type=data.type,
        sub_type=data.sub_type,
        party_id=data.party_id,
        party_name=party_name,
        particulars=data.particulars,
        amount=data.amount,
        bf_disc=data.bf_disc,
        mode=data.mode
    )
    doc = entry.model_dump()
    doc["created_at"] = datetime.utcnow().isoformat()
    await db.cash_book.insert_one(doc)
    return serialize_doc(doc)

@api_router.delete("/cash-book/{entry_id}")
async def delete_cash_book_entry(entry_id: str):
    await db.cash_book.delete_one({"id": entry_id})
    return {"status": "deleted"}

@api_router.put("/cash-book/{entry_id}")
async def update_cash_book_entry(entry_id: str, data: dict):
    update_fields = {}
    for field in ["date", "type", "sub_type", "party_id", "amount", "bf_disc", "mode", "particulars"]:
        if field in data:
            update_fields[field] = data[field]
    
    # Update party_name if party_id changed
    if "party_id" in data and data["party_id"]:
        for coll in ["bepaaris", "dukandars", "advance_parties", "capital_partners"]:
            party = await db[coll].find_one({"id": data["party_id"]})
            if party:
                update_fields["party_name"] = party.get("name")
                break
    
    await db.cash_book.update_one({"id": entry_id}, {"$set": update_fields})
    return {"status": "updated"}


# ============== PARTY STATEMENT (FOR DISPUTES) ==============
@api_router.get("/party-statement/{party_type}/{party_id}")
async def get_party_statement(
    party_type: str,
    party_id: str,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Get complete statement for a Bepaari or Dukandar"""
    
    # Get party info
    if party_type == "bepaari":
        party = await db.bepaaris.find_one({"id": party_id})
        cash_type = "BEPAARI"
        transfer_type = "BEPAARI"
    else:
        party = await db.dukandars.find_one({"id": party_id})
        cash_type = "DUKANDAR"
        transfer_type = "DUKANDAR"
    
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Calculate opening balance as of from_date
    # If from_date is specified, we need to calculate balance up to (from_date - 1 day)
    original_opening = party.get("opening_balance", 0)
    calculated_opening = original_opening
    
    if from_date:
        # Get all transactions BEFORE from_date to calculate opening balance
        prev_date_query = {"$lt": from_date}
        
        # Previous sales
        prev_sales_query = {"bepaari_id" if party_type == "bepaari" else "dukandar_id": party_id, "date": prev_date_query}
        prev_sales = await db.daily_sales.find(prev_sales_query).to_list(5000)
        prev_sales_total = sum(
            (s.get("dukandar_amount") or s["gross_amount"]) if party_type == "dukandar" else s["gross_amount"]
            for s in prev_sales
        )
        prev_discount_total = sum(s.get("discount", 0) for s in prev_sales)
        
        # Previous cash entries
        prev_cash_query = {"party_id": party_id, "type": cash_type, "date": prev_date_query}
        prev_cash = await db.cash_book.find(prev_cash_query).to_list(5000)
        
        # Previous adjustments
        prev_adj = await db.adjustments.find({"date": prev_date_query}).to_list(5000)
        
        # Previous balance transfers
        prev_transfers = await db.balance_transfers.find({
            "$or": [{"from_party_id": party_id}, {"to_party_id": party_id}],
            "date": prev_date_query
        }).to_list(5000)
        
        if party_type == "bepaari":
            prev_payments = sum(c["amount"] for c in prev_cash if c.get("sub_type") == "PAYMENT")
            prev_deductions = sum(c["amount"] for c in prev_cash if c.get("sub_type") != "PAYMENT")
            # Party-to-party adj credits reduce payable; expense adj credits increase payable
            expense_heads = ["MANDI_EXPENSE", "BF_DISCOUNT"]
            prev_adj_credits_party = sum(a["amount"] for a in prev_adj if a.get("credit_type") == "BEPAARI" and a.get("credit_party_id") == party_id and a.get("debit_type") not in expense_heads)
            prev_adj_credits_expense = sum(a["amount"] for a in prev_adj if a.get("credit_type") == "BEPAARI" and a.get("credit_party_id") == party_id and a.get("debit_type") in expense_heads)
            prev_adj_debits = sum(a["amount"] for a in prev_adj if a.get("debit_type") == "BEPAARI" and a.get("debit_party_id") == party_id)
            
            # Bepaari balance = Opening + Sales - Deductions - Payments - PartyAdj + ExpenseAdj
            # Note: Balance transfers already modify opening_balance directly, so DON'T include them here
            calculated_opening = original_opening + prev_sales_total - prev_deductions - prev_payments - prev_adj_credits_party + prev_adj_credits_expense + prev_adj_debits
        else:
            prev_receipts = sum(c["amount"] for c in prev_cash if c.get("sub_type") == "RECEIPT")
            prev_bf_disc = sum(c.get("bf_disc", 0) for c in prev_cash)
            prev_adj_debits = sum(a["amount"] for a in prev_adj if a.get("debit_type") == "DUKANDAR" and a.get("debit_party_id") == party_id)
            # Write-offs: CREDIT to Dukandar from expense head = reduces receivable
            expense_heads = ["MANDI_EXPENSE", "BF_DISCOUNT"]
            prev_adj_writeoffs = sum(a["amount"] for a in prev_adj if a.get("credit_type") == "DUKANDAR" and a.get("credit_party_id") == party_id and a.get("debit_type") in expense_heads)
            prev_adj_credits = sum(a["amount"] for a in prev_adj if a.get("credit_type") == "DUKANDAR" and a.get("credit_party_id") == party_id and a.get("debit_type") not in expense_heads)
            
            # Dukandar balance = Opening + Purchases - Discounts - Receipts - BF_Disc - Adj_Debits - Writeoffs
            # Note: Balance transfers already modify opening_balance directly, so DON'T include them here
            calculated_opening = original_opening + prev_sales_total - prev_discount_total - prev_receipts - prev_bf_disc - prev_adj_debits - prev_adj_writeoffs + prev_adj_credits
    
    # Build date query for filtered period
    date_query = {}
    if from_date and to_date:
        date_query = {"$gte": from_date, "$lte": to_date}
    elif from_date:
        date_query = {"$gte": from_date}
    elif to_date:
        date_query = {"$lte": to_date}
    
    # Get sales
    sales_query = {"bepaari_id" if party_type == "bepaari" else "dukandar_id": party_id}
    if date_query:
        sales_query["date"] = date_query
    sales = serialize_docs(await db.daily_sales.find(sales_query).sort("date", 1).to_list(5000))
    
    # Get cash book entries
    cash_query = {"party_id": party_id, "type": cash_type}
    if date_query:
        cash_query["date"] = date_query
    cash_entries = serialize_docs(await db.cash_book.find(cash_query).sort("date", 1).to_list(5000))
    
    # Get adjustments (JVs) where this party is involved
    adj_query = {}
    if date_query:
        adj_query["date"] = date_query
    
    all_adjustments = serialize_docs(await db.adjustments.find(adj_query).sort("date", 1).to_list(5000))
    
    # Filter adjustments for this party (as either debit or credit)
    if party_type == "bepaari":
        # For Bepaari: CREDIT side means someone paid them (reduces our payable)
        party_adjustments = [
            {**a, "direction": "CREDIT", "effect": "Received payment from " + a["debit_party_name"]}
            for a in all_adjustments 
            if a.get("credit_type") == "BEPAARI" and a.get("credit_party_id") == party_id
        ]
        # Also check if Bepaari is on DEBIT side (rare, but possible)
        party_adjustments += [
            {**a, "direction": "DEBIT", "effect": "Paid to " + a["credit_party_name"]}
            for a in all_adjustments 
            if a.get("debit_type") == "BEPAARI" and a.get("debit_party_id") == party_id
        ]
    else:
        # For Dukandar: DEBIT side means they paid someone on our behalf (reduces their receivable)
        party_adjustments = [
            {**a, "direction": "DEBIT", "effect": "Paid to " + a["credit_party_name"]}
            for a in all_adjustments 
            if a.get("debit_type") == "DUKANDAR" and a.get("debit_party_id") == party_id
        ]
        # Also check if Dukandar is on CREDIT side (rare)
        party_adjustments += [
            {**a, "direction": "CREDIT", "effect": "Received payment from " + a["debit_party_name"]}
            for a in all_adjustments 
            if a.get("credit_type") == "DUKANDAR" and a.get("credit_party_id") == party_id
        ]
    
    # Get balance transfers where this party is involved (FROM or TO)
    transfer_query = {"$or": [{"from_party_id": party_id}, {"to_party_id": party_id}]}
    if date_query:
        transfer_query["date"] = date_query
    
    all_transfers = serialize_docs(await db.balance_transfers.find(transfer_query).sort("date", 1).to_list(5000))
    
    # Format transfers with direction
    party_transfers = []
    for t in all_transfers:
        if t.get("from_party_id") == party_id:
            # Balance transferred OUT of this party
            party_transfers.append({
                **t,
                "direction": "OUT",
                "effect": f"Balance transferred to {t.get('to_party_name', 'Unknown')}"
            })
        elif t.get("to_party_id") == party_id:
            # Balance transferred INTO this party
            party_transfers.append({
                **t,
                "direction": "IN",
                "effect": f"Balance received from {t.get('from_party_name', 'Unknown')}"
            })
    
    # Calculate totals
    total_sales = sum(
        (s.get("dukandar_amount") or s["gross_amount"]) if party_type == "dukandar" else s["gross_amount"]
        for s in sales
    )
    total_qty = sum(s["quantity"] for s in sales)
    total_discount = sum(s["discount"] for s in sales)
    
    # Calculate transfer totals
    transfers_in = sum(t["amount"] for t in party_transfers if t["direction"] == "IN")
    transfers_out = sum(t["amount"] for t in party_transfers if t["direction"] == "OUT")
    
    if party_type == "bepaari":
        payments = sum(c["amount"] for c in cash_entries if c["sub_type"] == "PAYMENT")
        deductions = sum(c["amount"] for c in cash_entries if c["sub_type"] != "PAYMENT")
        # JV adjustments: CREDIT to Bepaari = payment received
        adj_received = sum(a["amount"] for a in party_adjustments if a["direction"] == "CREDIT")
        adj_paid = sum(a["amount"] for a in party_adjustments if a["direction"] == "DEBIT")
    else:
        payments = sum(c["amount"] for c in cash_entries if c["sub_type"] == "RECEIPT")
        deductions = sum(c["amount"] for c in cash_entries if c["sub_type"] != "RECEIPT")
        # JV adjustments: DEBIT from Dukandar = payment made on our behalf
        adj_paid = sum(a["amount"] for a in party_adjustments if a["direction"] == "DEBIT")
        adj_received = sum(a["amount"] for a in party_adjustments if a["direction"] == "CREDIT")
    
    return {
        "party": serialize_doc(party),
        "sales": sales,
        "cash_entries": cash_entries,
        "adjustments": party_adjustments,
        "balance_transfers": party_transfers,
        "summary": {
            "total_sales": total_sales,
            "total_quantity": total_qty,
            "total_discount": total_discount,
            "total_payments": payments,
            "total_deductions": deductions,
            "total_adjustments": adj_received + adj_paid,
            "transfers_in": transfers_in,
            "transfers_out": transfers_out,
            "opening_balance": calculated_opening
        }
    }


# ============== EXCEL EXPORT ==============
@api_router.get("/export/party-statement/{party_type}/{party_id}")
async def export_party_statement(
    party_type: str,
    party_id: str,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Export party statement to CSV/Excel"""
    statement = await get_party_statement(party_type, party_id, from_date, to_date)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    party_name = statement["party"]["name"]
    writer.writerow([f"Statement for: {party_name}"])
    writer.writerow([f"Party Type: {party_type.upper()}"])
    writer.writerow([f"Phone: {statement['party'].get('phone', 'N/A')}"])
    writer.writerow([f"Opening Balance: {statement['party'].get('opening_balance', 0)}"])
    if from_date:
        writer.writerow([f"From: {from_date}"])
    if to_date:
        writer.writerow([f"To: {to_date}"])
    writer.writerow([])
    
    # Sales
    writer.writerow(["=== SALES ==="])
    writer.writerow(["Date", "Bepaari" if party_type == "dukandar" else "Dukandar", "Quantity", "Rate", "Gross", "Discount", "Net"])
    for s in statement["sales"]:
        writer.writerow([
            s["date"],
            s["bepaari_name"] if party_type == "dukandar" else s["dukandar_name"],
            s["quantity"],
            s["rate"],
            s["gross_amount"],
            s["discount"],
            s["net_amount"]
        ])
    writer.writerow([])
    
    # Cash Book Entries
    writer.writerow(["=== PAYMENTS/RECEIPTS ==="])
    writer.writerow(["Date", "Type", "Sub-Type", "Amount", "Mode"])
    for c in statement["cash_entries"]:
        writer.writerow([c["date"], c["type"], c["sub_type"], c["amount"], c["mode"]])
    writer.writerow([])
    
    # Adjustments (JV) - NEW SECTION
    if statement.get("adjustments"):
        writer.writerow(["=== ADJUSTMENTS (JV) ==="])
        writer.writerow(["Date", "Direction", "Effect", "Amount", "Narration"])
        for a in statement["adjustments"]:
            writer.writerow([a["date"], a["direction"], a["effect"], a["amount"], a.get("narration", "")])
        writer.writerow([])
    
    # Balance Transfers
    if statement.get("balance_transfers"):
        writer.writerow(["=== BALANCE TRANSFERS ==="])
        writer.writerow(["Date", "Direction", "Effect", "Amount", "Narration"])
        for t in statement["balance_transfers"]:
            writer.writerow([t["date"], t["direction"], t["effect"], t["amount"], t.get("narration", "")])
        writer.writerow([])
    
    # Summary
    writer.writerow(["=== SUMMARY ==="])
    writer.writerow(["Total Sales", statement["summary"]["total_sales"]])
    writer.writerow(["Total Quantity", statement["summary"]["total_quantity"]])
    writer.writerow(["Total Discount", statement["summary"]["total_discount"]])
    writer.writerow(["Total Payments", statement["summary"]["total_payments"]])
    writer.writerow(["Total Deductions", statement["summary"]["total_deductions"]])
    writer.writerow(["Total Adjustments (JV)", statement["summary"].get("total_adjustments", 0)])
    writer.writerow(["Balance Transfers In", statement["summary"].get("transfers_in", 0)])
    writer.writerow(["Balance Transfers Out", statement["summary"].get("transfers_out", 0)])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={party_name}_statement.csv"}
    )


# ============== BEPAARI LEDGER ==============
@api_router.get("/bepaari-ledger")
async def get_bepaari_ledger(as_on_date: Optional[str] = None):
    bepaaris = await db.bepaaris.find({"is_active": True}).to_list(500)
    
    sales_query = {}
    cash_query = {"type": "BEPAARI"}
    adj_query = {}
    if as_on_date:
        sales_query["date"] = {"$lte": as_on_date}
        cash_query["date"] = {"$lte": as_on_date}
        adj_query["date"] = {"$lte": as_on_date}
    
    sales = await db.daily_sales.find(sales_query).to_list(5000)
    cash_entries = await db.cash_book.find(cash_query).to_list(5000)
    adjustments = await db.adjustments.find(adj_query).to_list(5000)
    settings = await get_settings()
    
    ledger = []
    for b in serialize_docs(bepaaris):
        b_sales = [s for s in serialize_docs(sales) if s["bepaari_id"] == b["id"]]
        b_cash = [c for c in serialize_docs(cash_entries) if c.get("party_id") == b["id"]]
        
        gross = sum(s["gross_amount"] for s in b_sales)
        qty = sum(s["quantity"] for s in b_sales)
        
        # Count unique market days for KK calculation
        market_days = len(set(s["date"] for s in b_sales))
        
        if b.get("flat_rate_per_goat"):
            commission = b["flat_rate_per_goat"] * qty
        else:
            commission = gross * (b.get("commission_percent", settings.get("commission_rate", 4)) / 100)
        
        # KK is charged per market day, not flat
        kk = settings.get("kk_fixed", 100) * market_days if market_days > 0 else 0
        jb = qty * settings.get("jb_rate", 10)
        
        motor = sum(c["amount"] for c in b_cash if c.get("sub_type") == "MOTOR")
        bhussa = sum(c["amount"] for c in b_cash if c.get("sub_type") == "BHUSSA")
        gawali = sum(c["amount"] for c in b_cash if c.get("sub_type") == "GAWALI")
        cash_adv = sum(c["amount"] for c in b_cash if c.get("sub_type") == "CASH_ADV")
        payments = sum(c["amount"] for c in b_cash if c.get("sub_type") == "PAYMENT")
        
        # Adjustments: CREDIT to Bepaari from another PARTY = reduces our payable (someone paid them)
        # CREDIT to Bepaari from EXPENSE HEAD = increases our payable (extra payment/write-off)
        expense_heads = ["MANDI_EXPENSE", "BF_DISCOUNT"]
        adj_credit_party = sum(a["amount"] for a in serialize_docs(adjustments) 
                       if a.get("credit_type") == "BEPAARI" and a.get("credit_party_id") == b["id"]
                       and a.get("debit_type") not in expense_heads)
        adj_credit_expense = sum(a["amount"] for a in serialize_docs(adjustments) 
                       if a.get("credit_type") == "BEPAARI" and a.get("credit_party_id") == b["id"]
                       and a.get("debit_type") in expense_heads)
        
        total_ded = commission + kk + jb + motor + bhussa + gawali + cash_adv
        net_payable = b.get("opening_balance", 0) + gross - total_ded
        balance = net_payable - payments - adj_credit_party + adj_credit_expense
        
        ledger.append({
            "id": b["id"],
            "name": b["name"],
            "phone": b.get("phone", ""),
            "opening": b.get("opening_balance", 0),
            "gross_sales": gross,
            "quantity": qty,
            "commission": commission,
            "kk": kk,
            "jb": jb,
            "motor": motor,
            "bhussa": bhussa,
            "gawali": gawali,
            "cash_adv": cash_adv,
            "total_deductions": total_ded,
            "net_payable": net_payable,
            "payments": payments,
            "adjustments": adj_credit_party,
            "expense_writeoffs": adj_credit_expense,
            "balance": balance
        })
    
    return ledger


# ============== DUKANDAR LEDGER ==============
@api_router.get("/dukandar-ledger")
async def get_dukandar_ledger(as_on_date: Optional[str] = None):
    dukandars = await db.dukandars.find({"is_active": True}).to_list(500)
    
    sales_query = {}
    cash_query = {"type": "DUKANDAR"}
    adj_query = {}
    if as_on_date:
        sales_query["date"] = {"$lte": as_on_date}
        cash_query["date"] = {"$lte": as_on_date}
        adj_query["date"] = {"$lte": as_on_date}
    
    sales = await db.daily_sales.find(sales_query).to_list(5000)
    cash_entries = await db.cash_book.find(cash_query).to_list(5000)
    adjustments = await db.adjustments.find(adj_query).to_list(5000)
    
    ledger = []
    for d in serialize_docs(dukandars):
        d_sales = [s for s in serialize_docs(sales) if s["dukandar_id"] == d["id"]]
        d_cash = [c for c in serialize_docs(cash_entries) if c.get("party_id") == d["id"]]
        
        purchases = sum(s.get("dukandar_amount") or s["gross_amount"] for s in d_sales)
        discounts = sum(s["discount"] for s in d_sales)
        receipts = sum(c["amount"] for c in d_cash if c.get("sub_type") == "RECEIPT")
        # BF_DISC: Only from standalone BF_DISC entries reduces balance
        # bf_disc field in RECEIPT is for expense tracking only (Amount already settles full balance)
        bf_disc_standalone = sum(c["amount"] for c in d_cash if c.get("sub_type") == "BF_DISC")
        bf_disc_in_receipt = sum(c.get("bf_disc", 0) for c in d_cash if c.get("sub_type") == "RECEIPT")
        bf_disc_total = bf_disc_standalone + bf_disc_in_receipt  # For display only
        
        # Adjustments: DEBIT to Dukandar = reduces their receivable (they paid someone on our behalf)
        adj_debit = sum(a["amount"] for a in serialize_docs(adjustments) 
                        if a.get("debit_type") == "DUKANDAR" and a.get("debit_party_id") == d["id"])
        # Write-offs: CREDIT to Dukandar from expense head = also reduces receivable
        expense_heads = ["MANDI_EXPENSE", "BF_DISCOUNT"]
        adj_writeoff = sum(a["amount"] for a in serialize_docs(adjustments)
                          if a.get("credit_type") == "DUKANDAR" and a.get("credit_party_id") == d["id"]
                          and a.get("debit_type") in expense_heads)
        
        net_receivable = d.get("opening_balance", 0) + purchases - discounts
        # Only standalone BF_DISC reduces balance, NOT bf_disc from RECEIPT (already in Amount)
        balance = net_receivable - receipts - bf_disc_standalone - adj_debit - adj_writeoff
        
        ledger.append({
            "id": d["id"],
            "name": d["name"],
            "phone": d.get("phone", ""),
            "opening": d.get("opening_balance", 0),
            "purchases": purchases,
            "discounts": discounts,
            "net_receivable": net_receivable,
            "receipts": receipts,
            "bf_disc": bf_disc_total,  # Shows total for reference
            "adjustments": adj_debit,
            "writeoffs": adj_writeoff,
            "balance": balance
        })
    
    return ledger


# ============== BALANCE SHEET ==============
@api_router.get("/balance-sheet")
async def get_balance_sheet(as_on_date: Optional[str] = None):
    settings = await get_settings()
    bepaari_ledger = await get_bepaari_ledger(as_on_date)
    dukandar_ledger = await get_dukandar_ledger(as_on_date)
    
    cash_query = {}
    if as_on_date:
        cash_query["date"] = {"$lte": as_on_date}
    
    cash_entries = serialize_docs(await db.cash_book.find(cash_query).to_list(5000))
    capital_partners = serialize_docs(await db.capital_partners.find({"is_active": True}).to_list(100))
    advance_parties = serialize_docs(await db.advance_parties.find({"is_active": True}).to_list(100))
    
    # Build individual lists for Capital, Loans, and Amanat with their balances
    capital_list = []
    loans_list = []
    amanat_list = []
    
    for p in capital_partners:
        p_type = p.get("partner_type")
        p_name = p.get("name")
        p_opening = p.get("opening_balance", 0)
        
        # Calculate movement for this partner
        partner_taken = sum(c["amount"] for c in cash_entries 
                          if c.get("type") == p_type and c.get("sub_type") == "TAKEN" and c.get("party_name") == p_name)
        partner_repaid = sum(c["amount"] for c in cash_entries 
                           if c.get("type") == p_type and c.get("sub_type") in ["REPAID", "WITHDRAWN"] and c.get("party_name") == p_name)
        
        partner_balance = p_opening + partner_taken - partner_repaid
        
        if partner_balance != 0:  # Only show if there's a balance
            entry = {"id": p.get("id"), "name": p_name, "amount": partner_balance}
            if p_type == "CAPITAL":
                capital_list.append(entry)
            elif p_type == "LOAN":
                loans_list.append(entry)
            elif p_type == "AMANAT":
                amanat_list.append(entry)
    
    # Calculate totals
    capital = sum(p["amount"] for p in capital_list)
    loans = sum(p["amount"] for p in loans_list)
    amanat = sum(p["amount"] for p in amanat_list)
    
    bepaari_payables = sum(b["balance"] for b in bepaari_ledger if b["balance"] > 0)
    bepaari_advances = sum(-b["balance"] for b in bepaari_ledger if b["balance"] < 0)
    patti = sum(d["balance"] for d in dukandar_ledger if d["balance"] > 0)
    dukandar_advances = sum(-d["balance"] for d in dukandar_ledger if d["balance"] < 0)
    
    jb_collected = sum(b["jb"] for b in bepaari_ledger)
    kk_collected = sum(b["kk"] for b in bepaari_ledger)
    commission_earned = sum(b["commission"] for b in bepaari_ledger)
    discounts_given = sum(d["discounts"] for d in dukandar_ledger)
    
    # Rate difference: when dukandar_rate > bepaari rate, the extra goes to commission
    sales_query = {}
    if as_on_date:
        sales_query["date"] = {"$lte": as_on_date}
    all_sales = serialize_docs(await db.daily_sales.find(sales_query).to_list(10000))
    rate_difference = sum(
        (s["dukandar_amount"] - s["gross_amount"])
        for s in all_sales if s.get("dukandar_amount")
    )
    
    jb_paid = sum(c["amount"] for c in cash_entries if c.get("sub_type") == "JB_PAID")
    jb_total = settings.get("jb_opening", 0) + jb_collected - jb_paid
    kk_total = settings.get("kk_opening", 0) + kk_collected
    commission_total = settings.get("commission_opening", 0) + commission_earned + rate_difference - discounts_given
    
    zakat_prov = sum(c["amount"] for c in cash_entries if c.get("type") == "ZAKAT" and c.get("sub_type") == "PROVISION")
    zakat_paid = sum(c["amount"] for c in cash_entries if c.get("type") == "ZAKAT" and c.get("sub_type") == "PAID")
    zakat_total = settings.get("zakat_opening", 0) + zakat_prov - zakat_paid
    
    cash_in_types = ["RECEIPT", "TAKEN", "RECEIVED"]
    # For RECEIPT entries, actual cash received = amount - bf_disc
    cash_in = sum(
        (c["amount"] - c.get("bf_disc", 0)) if c.get("sub_type") == "RECEIPT" else c["amount"]
        for c in cash_entries 
        if c.get("mode") == "CASH" and c.get("sub_type") in cash_in_types
    )
    cash_out = sum(c["amount"] for c in cash_entries if c.get("mode") == "CASH" and c.get("sub_type") not in cash_in_types)
    cash_balance = settings.get("opening_cash", 0) + cash_in - cash_out
    
    bank_modes = ["BANK", "UPI", "TRANSFER"]
    # For RECEIPT entries, actual bank received = amount - bf_disc
    bank_in = sum(
        (c["amount"] - c.get("bf_disc", 0)) if c.get("sub_type") == "RECEIPT" else c["amount"]
        for c in cash_entries 
        if c.get("mode") in bank_modes and c.get("sub_type") in cash_in_types
    )
    bank_out = sum(c["amount"] for c in cash_entries if c.get("mode") in bank_modes and c.get("sub_type") not in cash_in_types)
    bank_balance = settings.get("opening_bank", 0) + bank_in - bank_out
    
    exp_types = ["MANDI", "TRAVEL", "FOOD", "SALARY", "MISC", "OTHER"]
    mandi_exp = sum(c["amount"] for c in cash_entries if c.get("sub_type") in exp_types)
    # BF_DISC: from standalone BF_DISC entries + bf_disc field in RECEIPT entries
    bf_disc_standalone = sum(c["amount"] for c in cash_entries if c.get("sub_type") == "BF_DISC")
    bf_disc_in_receipt = sum(c.get("bf_disc", 0) for c in cash_entries if c.get("sub_type") == "RECEIPT")
    bf_disc = bf_disc_standalone + bf_disc_in_receipt
    mhn_personal = sum(c["amount"] for c in cash_entries if c.get("sub_type") == "MHN_PERSONAL")
    
    mandi_total = settings.get("mandi_exp_opening", 0) + mandi_exp
    bf_disc_total = settings.get("bf_disc_opening", 0) + bf_disc
    mhn_total = settings.get("mhn_personal_opening", 0) + mhn_personal
    
    # Get adjustments for advance party and expense write-off calculations
    adj_query = {}
    if as_on_date:
        adj_query["date"] = {"$lte": as_on_date}
    adjustments = serialize_docs(await db.adjustments.find(adj_query).to_list(5000))
    
    # Add expense write-offs from JV entries to expense totals
    jv_mandi_exp = sum(a["amount"] for a in adjustments if a.get("debit_type") == "MANDI_EXPENSE")
    jv_bf_disc = sum(a["amount"] for a in adjustments if a.get("debit_type") == "BF_DISCOUNT")
    mandi_total += jv_mandi_exp
    bf_disc_total += jv_bf_disc
    
    adv_receivables = []
    for ap in advance_parties:
        given = sum(c["amount"] for c in cash_entries if c.get("type") == "ADVANCE" and c.get("sub_type") == "GIVEN" and c.get("party_name") == ap["name"])
        received = sum(c["amount"] for c in cash_entries if c.get("type") == "ADVANCE" and c.get("sub_type") == "RECEIVED" and c.get("party_name") == ap["name"])
        # JV: When ADVANCE party is on DEBIT side, they paid someone (reduces their receivable)
        jv_paid = sum(a["amount"] for a in adjustments if a.get("debit_type") == "ADVANCE" and a.get("debit_party_id") == ap["id"])
        bal = ap.get("opening_balance", 0) + given - received - jv_paid
        if bal != 0:
            adv_receivables.append({"name": ap["name"], "amount": bal})
    
    adv_total = sum(a["amount"] for a in adv_receivables if a["amount"] > 0)
    
    total_liab = capital + loans + amanat + bepaari_payables + dukandar_advances + jb_total + kk_total + commission_total + zakat_total
    total_assets = cash_balance + bank_balance + patti + bepaari_advances + mandi_total + bf_disc_total + mhn_total + adv_total
    
    return {
        "as_on_date": as_on_date or "Current",
        "liabilities": {
            "capital": capital, "capital_list": capital_list,
            "loans": loans, "loans_list": loans_list,
            "amanat": amanat, "amanat_list": amanat_list,
            "bepaari_payables": bepaari_payables, "dukandar_advances": dukandar_advances,
            "jb": {"bf": settings.get("jb_opening", 0), "collected": jb_collected, "paid": jb_paid, "total": jb_total},
            "kk": {"bf": settings.get("kk_opening", 0), "collected": kk_collected, "total": kk_total},
            "commission": {"bf": settings.get("commission_opening", 0), "earned": commission_earned, "rate_diff": rate_difference, "discounts": discounts_given, "total": commission_total},
            "zakat": zakat_total, "total": total_liab
        },
        "assets": {
            "cash_balance": cash_balance, "bank_balance": bank_balance,
            "patti": patti, "bepaari_advances": bepaari_advances,
            "mandi_expenses": {"bf": settings.get("mandi_exp_opening", 0), "current": mandi_exp, "total": mandi_total},
            "bf_discount": {"bf": settings.get("bf_disc_opening", 0), "current": bf_disc, "total": bf_disc_total},
            "mhn_personal": {"bf": settings.get("mhn_personal_opening", 0), "current": mhn_personal, "total": mhn_total},
            "advance_receivables": adv_receivables, "total": total_assets
        },
        "difference": total_liab - total_assets
    }


@api_router.get("/dashboard")
async def get_dashboard():
    today = datetime.utcnow().strftime("%Y-%m-%d")
    today_sales = serialize_docs(await db.daily_sales.find({"date": today}).to_list(100))
    balance_sheet = await get_balance_sheet()
    
    return {
        "today": {"sales_count": len(today_sales), "sales_amount": sum(s["gross_amount"] for s in today_sales)},
        "summary": {
            "cash_balance": balance_sheet["assets"]["cash_balance"],
            "bank_balance": balance_sheet["assets"]["bank_balance"],
            "patti": balance_sheet["assets"]["patti"],
            "bepaari_payable": balance_sheet["liabilities"]["bepaari_payables"],
            "balance_diff": balance_sheet["difference"]
        }
    }


# ============== JOURNAL VOUCHER / ADJUSTMENTS ==============
@api_router.get("/adjustments")
async def get_adjustments(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Get all adjustment/JV entries"""
    query = {}
    if from_date and to_date:
        query["date"] = {"$gte": from_date, "$lte": to_date}
    elif from_date:
        query["date"] = {"$gte": from_date}
    elif to_date:
        query["date"] = {"$lte": to_date}
    
    adjustments = await db.adjustments.find(query).sort("date", -1).to_list(1000)
    return serialize_docs(adjustments)


@api_router.post("/adjustments")
async def create_adjustment(data: AdjustmentEntryCreate):
    """Create a new adjustment/JV entry"""
    EXPENSE_HEADS = {
        "MANDI_EXPENSE": {"id": "__MANDI_EXPENSE__", "name": "Mandi Expense"},
        "BF_DISCOUNT": {"id": "__BF_DISCOUNT__", "name": "BF Discount"},
    }
    
    # Get party names
    debit_party_name = None
    credit_party_name = None
    
    # Find debit party (or expense head)
    if data.debit_type in EXPENSE_HEADS:
        debit_party_name = EXPENSE_HEADS[data.debit_type]["name"]
    else:
        for coll in ["bepaaris", "dukandars", "advance_parties", "capital_partners"]:
            party = await db[coll].find_one({"id": data.debit_party_id})
            if party:
                debit_party_name = party.get("name")
                break
    
    # Find credit party (or expense head)
    if data.credit_type in EXPENSE_HEADS:
        credit_party_name = EXPENSE_HEADS[data.credit_type]["name"]
    else:
        for coll in ["bepaaris", "dukandars", "advance_parties", "capital_partners"]:
            party = await db[coll].find_one({"id": data.credit_party_id})
            if party:
                credit_party_name = party.get("name")
                break
    
    if not debit_party_name or not credit_party_name:
        raise HTTPException(status_code=400, detail="Invalid party ID")
    
    # For expense heads, use the fixed synthetic ID
    debit_party_id = EXPENSE_HEADS[data.debit_type]["id"] if data.debit_type in EXPENSE_HEADS else data.debit_party_id
    credit_party_id = EXPENSE_HEADS[data.credit_type]["id"] if data.credit_type in EXPENSE_HEADS else data.credit_party_id
    
    adjustment = AdjustmentEntry(
        date=data.date,
        debit_type=data.debit_type,
        debit_party_id=debit_party_id,
        debit_party_name=debit_party_name,
        credit_type=data.credit_type,
        credit_party_id=credit_party_id,
        credit_party_name=credit_party_name,
        amount=data.amount,
        narration=data.narration
    )
    
    doc = adjustment.model_dump()
    doc["created_at"] = datetime.utcnow().isoformat()
    await db.adjustments.insert_one(doc)
    return serialize_doc(doc)


@api_router.delete("/adjustments/{adjustment_id}")
async def delete_adjustment(adjustment_id: str):
    """Delete an adjustment entry"""
    await db.adjustments.delete_one({"id": adjustment_id})
    return {"status": "deleted"}

@api_router.put("/adjustments/{adjustment_id}")
async def update_adjustment(adjustment_id: str, data: dict):
    """Update an adjustment entry"""
    update_fields = {}
    for field in ["date", "amount", "narration"]:
        if field in data:
            update_fields[field] = data[field]
    
    # Update party info if changed
    if "debit_party_id" in data:
        update_fields["debit_party_id"] = data["debit_party_id"]
        update_fields["debit_type"] = data.get("debit_type", "")
        for coll in ["bepaaris", "dukandars", "advance_parties", "capital_partners"]:
            party = await db[coll].find_one({"id": data["debit_party_id"]})
            if party:
                update_fields["debit_party_name"] = party.get("name")
                break
    
    if "credit_party_id" in data:
        update_fields["credit_party_id"] = data["credit_party_id"]
        update_fields["credit_type"] = data.get("credit_type", "")
        for coll in ["bepaaris", "dukandars", "advance_parties", "capital_partners"]:
            party = await db[coll].find_one({"id": data["credit_party_id"]})
            if party:
                update_fields["credit_party_name"] = party.get("name")
                break
    
    await db.adjustments.update_one({"id": adjustment_id}, {"$set": update_fields})
    return {"status": "updated"}


# ============== BALANCE TRANSFER ==============
@api_router.post("/balance-transfer")
async def transfer_balance(data: dict):
    """Transfer balance from one party to another (corrects opening balances)"""
    from_type = data.get("from_type")  # BEPAARI, DUKANDAR, etc.
    from_party_id = data.get("from_party_id")
    to_type = data.get("to_type")
    to_party_id = data.get("to_party_id")  # Can be None if creating new
    to_party_name = data.get("to_party_name")  # For new party
    amount = float(data.get("amount", 0))
    transfer_date = data.get("date", datetime.utcnow().strftime("%Y-%m-%d"))
    narration = data.get("narration", "")
    
    if not from_type or not from_party_id or not to_type or amount <= 0:
        raise HTTPException(status_code=400, detail="Invalid transfer data")
    
    # Map type to collection
    type_to_collection = {
        "BEPAARI": "bepaaris",
        "DUKANDAR": "dukandars",
        "ADVANCE": "advance_parties"
    }
    
    from_coll = type_to_collection.get(from_type)
    to_coll = type_to_collection.get(to_type)
    
    if not from_coll or not to_coll:
        raise HTTPException(status_code=400, detail="Invalid party type")
    
    # Get source party
    from_party = await db[from_coll].find_one({"id": from_party_id})
    if not from_party:
        raise HTTPException(status_code=404, detail="Source party not found")
    
    # Reduce source party's opening balance
    new_from_balance = from_party.get("opening_balance", 0) - amount
    await db[from_coll].update_one({"id": from_party_id}, {"$set": {"opening_balance": new_from_balance}})
    
    # Handle destination party
    if to_party_id:
        # Existing party - increase their opening balance
        to_party = await db[to_coll].find_one({"id": to_party_id})
        if not to_party:
            raise HTTPException(status_code=404, detail="Destination party not found")
        new_to_balance = to_party.get("opening_balance", 0) + amount
        await db[to_coll].update_one({"id": to_party_id}, {"$set": {"opening_balance": new_to_balance}})
        to_name = to_party.get("name")
    else:
        # Create new party with the transfer amount as opening balance
        if not to_party_name:
            raise HTTPException(status_code=400, detail="New party name required")
        
        new_id = str(uuid.uuid4())
        new_party = {
            "id": new_id,
            "name": to_party_name,
            "opening_balance": amount,
            "is_active": True,
            "created_at": datetime.utcnow().isoformat()
        }
        
        if to_type == "BEPAARI":
            new_party["commission_percent"] = 4.0
        
        await db[to_coll].insert_one(new_party)
        to_name = to_party_name
        to_party_id = new_id
    
    # CREATE A TRANSFER RECORD for audit trail
    transfer_record = {
        "id": str(uuid.uuid4()),
        "date": transfer_date,
        "from_type": from_type,
        "from_party_id": from_party_id,
        "from_party_name": from_party.get("name"),
        "to_type": to_type,
        "to_party_id": to_party_id,
        "to_party_name": to_name,
        "amount": amount,
        "narration": narration,
        "created_at": datetime.utcnow().isoformat()
    }
    await db.balance_transfers.insert_one(transfer_record)
    
    return {
        "status": "transferred",
        "from_party": from_party.get("name"),
        "to_party": to_name,
        "amount": amount,
        "new_to_party_id": to_party_id,
        "transfer_id": transfer_record["id"]
    }


# ============== GET BALANCE TRANSFERS ==============
@api_router.get("/balance-transfers")
async def get_balance_transfers(
    party_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Get all balance transfer records, optionally filtered by party"""
    query = {}
    
    if party_id:
        query["$or"] = [
            {"from_party_id": party_id},
            {"to_party_id": party_id}
        ]
    
    if from_date and to_date:
        query["date"] = {"$gte": from_date, "$lte": to_date}
    elif from_date:
        query["date"] = {"$gte": from_date}
    elif to_date:
        query["date"] = {"$lte": to_date}
    
    transfers = await db.balance_transfers.find(query).sort("date", -1).to_list(1000)
    return serialize_docs(transfers)


@api_router.post("/balance-transfers/backfill")
async def backfill_balance_transfer(data: dict):
    """Create a transfer record for past transfers (audit trail backfill - does NOT modify balances)"""
    from_type = data.get("from_type")
    from_party_id = data.get("from_party_id")
    to_type = data.get("to_type")
    to_party_id = data.get("to_party_id")
    amount = float(data.get("amount", 0))
    transfer_date = data.get("date")
    narration = data.get("narration", "Backfilled transfer record")
    
    if not from_type or not from_party_id or not to_type or not to_party_id or amount <= 0 or not transfer_date:
        raise HTTPException(status_code=400, detail="All fields required: from_type, from_party_id, to_type, to_party_id, amount, date")
    
    # Map type to collection
    type_to_collection = {
        "BEPAARI": "bepaaris",
        "DUKANDAR": "dukandars",
        "ADVANCE": "advance_parties"
    }
    
    from_coll = type_to_collection.get(from_type)
    to_coll = type_to_collection.get(to_type)
    
    if not from_coll or not to_coll:
        raise HTTPException(status_code=400, detail="Invalid party type")
    
    # Get party names
    from_party = await db[from_coll].find_one({"id": from_party_id})
    to_party = await db[to_coll].find_one({"id": to_party_id})
    
    if not from_party or not to_party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Create transfer record (does NOT modify balances - just creates audit trail)
    transfer_record = {
        "id": str(uuid.uuid4()),
        "date": transfer_date,
        "from_type": from_type,
        "from_party_id": from_party_id,
        "from_party_name": from_party.get("name"),
        "to_type": to_type,
        "to_party_id": to_party_id,
        "to_party_name": to_party.get("name"),
        "amount": amount,
        "narration": narration,
        "created_at": datetime.utcnow().isoformat(),
        "is_backfill": True
    }
    await db.balance_transfers.insert_one(transfer_record)
    
    return {
        "status": "backfilled",
        "from_party": from_party.get("name"),
        "to_party": to_party.get("name"),
        "amount": amount,
        "transfer_id": transfer_record["id"]
    }


@api_router.put("/cash-book/reassign")
async def reassign_cash_book_entries(data: dict):
    """Reassign cash book entries from one party to another"""
    entry_ids = data.get("entry_ids", [])
    new_party_id = data.get("new_party_id")
    new_party_type = data.get("new_party_type")
    
    if not entry_ids or not new_party_id:
        raise HTTPException(status_code=400, detail="Invalid reassignment data")
    
    # Get new party name
    new_party_name = None
    for coll in ["bepaaris", "dukandars", "advance_parties", "capital_partners"]:
        party = await db[coll].find_one({"id": new_party_id})
        if party:
            new_party_name = party.get("name")
            break
    
    if not new_party_name:
        raise HTTPException(status_code=404, detail="New party not found")
    
    # Update all specified entries
    for entry_id in entry_ids:
        await db.cash_book.update_one(
            {"id": entry_id},
            {"$set": {"party_id": new_party_id, "party_name": new_party_name}}
        )
    
    return {"status": "reassigned", "count": len(entry_ids)}


# ============== BEPAARI AAKDA (Daily Settlement Slip) ==============
@api_router.get("/bepaari-aakda")
async def get_bepaari_aakda(date: str):
    """Get Aakda (settlement slip) for all Bepaaris for a specific market day"""
    
    bepaaris = serialize_docs(await db.bepaaris.find({"is_active": True}).to_list(500))
    sales = serialize_docs(await db.daily_sales.find({"date": date}).to_list(1000))
    cash_entries = serialize_docs(await db.cash_book.find({"date": date, "type": "BEPAARI"}).to_list(1000))
    settings = await get_settings()
    
    # Get previous balance (all transactions before this date)
    prev_sales = serialize_docs(await db.daily_sales.find({"date": {"$lt": date}}).to_list(5000))
    prev_cash = serialize_docs(await db.cash_book.find({"date": {"$lt": date}, "type": "BEPAARI"}).to_list(5000))
    
    # Get ALL adjustments (for opening balance) and today's adjustments
    all_adjustments = serialize_docs(await db.adjustments.find({}).to_list(5000))
    prev_adjustments = [a for a in all_adjustments if a["date"] < date]
    today_adjustments = [a for a in all_adjustments if a["date"] == date]
    
    aakda_list = []
    
    for b in bepaaris:
        # Today's sales
        b_sales_today = [s for s in sales if s["bepaari_id"] == b["id"]]
        b_cash_today = [c for c in cash_entries if c.get("party_id") == b["id"]]
        b_adj_today = [a for a in today_adjustments if a.get("credit_type") == "BEPAARI" and a.get("credit_party_id") == b["id"]]
        
        if not b_sales_today and not b_cash_today and not b_adj_today:
            continue  # Skip Bepaaris with no activity today
        
        # Previous data for opening balance calculation
        b_prev_sales = [s for s in prev_sales if s["bepaari_id"] == b["id"]]
        b_prev_cash = [c for c in prev_cash if c.get("party_id") == b["id"]]
        b_prev_adj = [a for a in prev_adjustments if a.get("credit_type") == "BEPAARI" and a.get("credit_party_id") == b["id"]]
        
        # Calculate previous balance (opening for today)
        prev_gross = sum(s["gross_amount"] for s in b_prev_sales)
        prev_qty = sum(s["quantity"] for s in b_prev_sales)
        prev_market_days = len(set(s["date"] for s in b_prev_sales))
        
        if b.get("flat_rate_per_goat"):
            prev_comm = b["flat_rate_per_goat"] * prev_qty
        else:
            prev_comm = prev_gross * (b.get("commission_percent", settings.get("commission_rate", 4)) / 100)
        
        prev_kk = settings.get("kk_fixed", 100) * prev_market_days if prev_market_days > 0 else 0
        prev_jb = prev_qty * settings.get("jb_rate", 10)
        prev_motor = sum(c["amount"] for c in b_prev_cash if c.get("sub_type") == "MOTOR")
        prev_bhussa = sum(c["amount"] for c in b_prev_cash if c.get("sub_type") == "BHUSSA")
        prev_gawali = sum(c["amount"] for c in b_prev_cash if c.get("sub_type") == "GAWALI")
        prev_cash_adv = sum(c["amount"] for c in b_prev_cash if c.get("sub_type") == "CASH_ADV")
        prev_payments = sum(c["amount"] for c in b_prev_cash if c.get("sub_type") == "PAYMENT")
        prev_adj_amount = sum(a["amount"] for a in b_prev_adj)  # JV credits to this Bepaari
        
        prev_total_ded = prev_comm + prev_kk + prev_jb + prev_motor + prev_bhussa + prev_gawali + prev_cash_adv
        opening_balance = b.get("opening_balance", 0) + prev_gross - prev_total_ded - prev_payments - prev_adj_amount
        
        # Today's calculations
        today_gross = sum(s["gross_amount"] for s in b_sales_today)
        today_qty = sum(s["quantity"] for s in b_sales_today)
        
        if b.get("flat_rate_per_goat"):
            today_comm = b["flat_rate_per_goat"] * today_qty
        else:
            today_comm = today_gross * (b.get("commission_percent", settings.get("commission_rate", 4)) / 100)
        
        today_kk = settings.get("kk_fixed", 100) if today_qty > 0 else 0
        today_jb = today_qty * settings.get("jb_rate", 10)
        today_motor = sum(c["amount"] for c in b_cash_today if c.get("sub_type") == "MOTOR")
        today_bhussa = sum(c["amount"] for c in b_cash_today if c.get("sub_type") == "BHUSSA")
        today_gawali = sum(c["amount"] for c in b_cash_today if c.get("sub_type") == "GAWALI")
        today_cash_adv = sum(c["amount"] for c in b_cash_today if c.get("sub_type") == "CASH_ADV")
        today_payments = sum(c["amount"] for c in b_cash_today if c.get("sub_type") == "PAYMENT")
        today_adj_amount = sum(a["amount"] for a in b_adj_today)  # JV credits today
        
        today_total_ded = today_comm + today_kk + today_jb + today_motor + today_bhussa + today_gawali + today_cash_adv
        today_net = today_gross - today_total_ded
        closing_balance = opening_balance + today_net - today_payments - today_adj_amount
        
        # Sales breakdown by Dukandar
        sales_detail = []
        today_rate_diff = 0
        for s in b_sales_today:
            duk_amt = s.get("dukandar_amount")
            if duk_amt:
                today_rate_diff += (duk_amt - s["gross_amount"])
            sales_detail.append({
                "dukandar": s["dukandar_name"],
                "quantity": s["quantity"],
                "rate": s["rate"],
                "amount": s["gross_amount"],
                "discount": s["discount"],
                "dukandar_rate": s.get("dukandar_rate"),
                "dukandar_amount": duk_amt
            })
        
        aakda_list.append({
            "bepaari_id": b["id"],
            "bepaari_name": b["name"],
            "phone": b.get("phone", ""),
            "date": date,
            "opening_balance": opening_balance,
            "sales_detail": sales_detail,
            "summary": {
                "gross_sales": today_gross,
                "quantity": today_qty,
                "commission": today_comm,
                "commission_pct": b.get("commission_percent", settings.get("commission_rate", 4)),
                "rate_diff": today_rate_diff,
                "kk": today_kk,
                "jb": today_jb,
                "jb_rate": settings.get("jb_rate", 10),
                "motor": today_motor,
                "bhussa": today_bhussa,
                "gawali": today_gawali,
                "cash_advance": today_cash_adv,
                "total_deductions": today_total_ded,
                "net_amount": today_net,
                "payments": today_payments,
                "jv_adjustment": today_adj_amount,
                "closing_balance": closing_balance
            }
        })
    
    return aakda_list


@api_router.get("/bepaari-aakda/{bepaari_id}")
async def get_single_bepaari_aakda(bepaari_id: str, date: str):
    """Get Aakda for a single Bepaari"""
    all_aakda = await get_bepaari_aakda(date)
    for a in all_aakda:
        if a["bepaari_id"] == bepaari_id:
            return a
    raise HTTPException(status_code=404, detail="No transactions found for this Bepaari on this date")


# ============== EXPORT ALL DATA ==============
@api_router.get("/export-all")
async def export_all_data():
    """Export all data as multi-sheet Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    gold_fill = PatternFill(start_color="C5A55A", end_color="C5A55A", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="E2E0D8"))
    num_fmt = '#,##0'

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 25)

    # 1. Daily Sales
    sales = serialize_docs(await db.daily_sales.find().sort([("date", -1), ("created_at", -1)]).to_list(10000))
    ws1 = wb.active
    ws1.title = "Daily Sales"
    style_header(ws1, ["Date", "Bepaari", "Dukandar", "Qty", "Rate", "Gross", "Discount", "Net", "Duk Rate", "Duk Amount"])
    for i, s in enumerate(sales, 2):
        ws1.cell(i, 1, s["date"])
        ws1.cell(i, 2, s.get("bepaari_name", ""))
        ws1.cell(i, 3, s.get("dukandar_name", ""))
        ws1.cell(i, 4, s["quantity"])
        ws1.cell(i, 5, s["rate"]).number_format = num_fmt
        ws1.cell(i, 6, s["gross_amount"]).number_format = num_fmt
        ws1.cell(i, 7, s.get("discount", 0)).number_format = num_fmt
        ws1.cell(i, 8, s.get("net_amount", 0)).number_format = num_fmt
        ws1.cell(i, 9, s.get("dukandar_rate") or "").number_format = num_fmt
        ws1.cell(i, 10, s.get("dukandar_amount") or "").number_format = num_fmt
    auto_width(ws1)

    # 2. Cash & Bank
    cash = serialize_docs(await db.cash_book.find().sort([("date", -1), ("created_at", -1)]).to_list(10000))
    ws2 = wb.create_sheet("Cash & Bank")
    style_header(ws2, ["Date", "Type", "Sub-Type", "Party", "Amount", "BF Disc", "Mode", "Comments"])
    for i, c in enumerate(cash, 2):
        ws2.cell(i, 1, c["date"])
        ws2.cell(i, 2, c["type"])
        ws2.cell(i, 3, c["sub_type"])
        ws2.cell(i, 4, c.get("party_name", ""))
        ws2.cell(i, 5, c["amount"]).number_format = num_fmt
        ws2.cell(i, 6, c.get("bf_disc", 0)).number_format = num_fmt
        ws2.cell(i, 7, c["mode"])
        ws2.cell(i, 8, c.get("particulars", ""))
    auto_width(ws2)

    # 3. Bepaari Ledger
    bepaari_ledger = await get_bepaari_ledger()
    ws3 = wb.create_sheet("Bepaari Ledger")
    style_header(ws3, ["Name", "Phone", "Opening", "Sales", "Qty", "Commission", "KK", "JB", "Total Ded", "Payments", "Adj", "Balance"])
    for i, b in enumerate(bepaari_ledger, 2):
        ws3.cell(i, 1, b["name"])
        ws3.cell(i, 2, b.get("phone", ""))
        ws3.cell(i, 3, b["opening"]).number_format = num_fmt
        ws3.cell(i, 4, b["gross_sales"]).number_format = num_fmt
        ws3.cell(i, 5, b["quantity"])
        ws3.cell(i, 6, b["commission"]).number_format = num_fmt
        ws3.cell(i, 7, b["kk"]).number_format = num_fmt
        ws3.cell(i, 8, b["jb"]).number_format = num_fmt
        ws3.cell(i, 9, b["total_deductions"]).number_format = num_fmt
        ws3.cell(i, 10, b["payments"]).number_format = num_fmt
        ws3.cell(i, 11, b.get("adjustments", 0)).number_format = num_fmt
        ws3.cell(i, 12, b["balance"]).number_format = num_fmt
    auto_width(ws3)

    # 4. Dukandar Ledger
    dukandar_ledger = await get_dukandar_ledger()
    ws4 = wb.create_sheet("Dukandar Ledger")
    style_header(ws4, ["Name", "Phone", "Opening", "Purchases", "Discounts", "Net Receivable", "Receipts", "BF Disc", "Adj", "Balance"])
    for i, d in enumerate(dukandar_ledger, 2):
        ws4.cell(i, 1, d["name"])
        ws4.cell(i, 2, d.get("phone", ""))
        ws4.cell(i, 3, d["opening"]).number_format = num_fmt
        ws4.cell(i, 4, d["purchases"]).number_format = num_fmt
        ws4.cell(i, 5, d["discounts"]).number_format = num_fmt
        ws4.cell(i, 6, d["net_receivable"]).number_format = num_fmt
        ws4.cell(i, 7, d["receipts"]).number_format = num_fmt
        ws4.cell(i, 8, d.get("bf_disc", 0)).number_format = num_fmt
        ws4.cell(i, 9, d.get("adjustments", 0)).number_format = num_fmt
        ws4.cell(i, 10, d["balance"]).number_format = num_fmt
    auto_width(ws4)

    # 5. Balance Sheet
    bs = await get_balance_sheet()
    ws5 = wb.create_sheet("Balance Sheet")
    L = bs["liabilities"]
    A = bs["assets"]
    gold_font = Font(bold=True, color="1B2A4A", size=11)

    ws5.cell(1, 1, "LIABILITIES").font = Font(bold=True, size=12, color="1B2A4A")
    ws5.cell(1, 3, "ASSETS").font = Font(bold=True, size=12, color="1B2A4A")
    row = 2
    liab_items = [
        ("Capital", L["capital"]),
        ("Loans", L["loans"]),
        ("Amanat", L["amanat"]),
        ("Bepaari Payables", L["bepaari_payables"]),
        ("Dukandar Advances", L.get("dukandar_advances", 0)),
        ("JB Total", L["jb"]["total"]),
        ("KK Total", L["kk"]["total"]),
        ("Commission (Net)", L["commission"]["total"]),
        ("  Gross Earned", L["commission"]["earned"]),
        ("  Rate Difference", L["commission"].get("rate_diff", 0)),
        ("  Less: Discounts", L["commission"]["discounts"]),
        ("Zakat", L.get("zakat", 0)),
    ]
    asset_items = [
        ("Cash Balance", A["cash_balance"]),
        ("Bank Balance", A["bank_balance"]),
        ("Patti (Receivable)", A["patti"]),
        ("Bepaari Advances", A.get("bepaari_advances", 0)),
        ("Mandi Expenses", A["mandi_expenses"]["total"]),
        ("BF Discount", A["bf_discount"]["total"]),
        ("MHN Personal", A["mhn_personal"]["total"]),
    ]
    for name, val in liab_items:
        ws5.cell(row, 1, name)
        ws5.cell(row, 2, val).number_format = num_fmt
        row += 1
    ws5.cell(row, 1, "TOTAL LIABILITIES").font = gold_font
    ws5.cell(row, 2, L["total"]).number_format = num_fmt
    ws5.cell(row, 2).font = gold_font

    row = 2
    for name, val in asset_items:
        ws5.cell(row, 3, name)
        ws5.cell(row, 4, val).number_format = num_fmt
        row += 1
    # Advance receivables
    for ar in A.get("advance_receivables", []):
        ws5.cell(row, 3, f"Adv: {ar['name']}")
        ws5.cell(row, 4, ar["amount"]).number_format = num_fmt
        row += 1
    ws5.cell(row, 3, "TOTAL ASSETS").font = gold_font
    ws5.cell(row, 4, A["total"]).number_format = num_fmt
    ws5.cell(row, 4).font = gold_font
    row += 1
    ws5.cell(row, 3, "DIFFERENCE")
    ws5.cell(row, 4, bs["difference"]).number_format = num_fmt
    auto_width(ws5)

    # 6. Adjustments
    adj = serialize_docs(await db.adjustments.find().sort("date", -1).to_list(10000))
    ws6 = wb.create_sheet("Adjustments")
    style_header(ws6, ["Date", "Debit Type", "Debit Party", "Credit Type", "Credit Party", "Amount", "Narration"])
    for i, a in enumerate(adj, 2):
        ws6.cell(i, 1, a["date"])
        ws6.cell(i, 2, a["debit_type"])
        ws6.cell(i, 3, a.get("debit_party_name", ""))
        ws6.cell(i, 4, a["credit_type"])
        ws6.cell(i, 5, a.get("credit_party_name", ""))
        ws6.cell(i, 6, a["amount"]).number_format = num_fmt
        ws6.cell(i, 7, a.get("narration", ""))
    auto_width(ws6)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Mandi_Data_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============== REGISTER ROUTER & MIDDLEWARE ==============
app.include_router(api_router)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
