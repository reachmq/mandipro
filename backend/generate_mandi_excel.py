"""
Mandi Master Excel Generator
Generates a comprehensive Excel workbook for livestock commission agent accounting
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import os

# Style definitions
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Dark green
SUBHEADER_FILL = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
TITLE_FONT = Font(bold=True, size=16, color="1B5E20")
MONEY_FORMAT = '₹#,##0.00'
NUMBER_FORMAT = '#,##0'
DATE_FORMAT = 'DD-MMM-YYYY'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def style_header_row(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def create_mandi_excel(output_path):
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # ========================================
    # SHEET 1: INSTRUCTIONS
    # ========================================
    ws_instructions = wb.create_sheet("Instructions", 0)
    ws_instructions.sheet_properties.tabColor = "FF9800"
    
    instructions = [
        ["MANDI MASTER - Commission Agent Accounting System"],
        [""],
        ["HOW TO USE THIS WORKBOOK:"],
        [""],
        ["STEP 1: Setup Masters (One-time)"],
        ["  → Go to 'Masters' sheet"],
        ["  → Add all your Bepaari names with their commission type (% or Per Piece)"],
        ["  → Add all your Dukandar names"],
        ["  → Enter Opening Balances for existing parties"],
        [""],
        ["STEP 2: Daily Entry (Each Mandi Day)"],
        ["  → Go to 'Daily_Sales' sheet"],
        ["  → Enter each sale: Date, Bepaari, Dukandar, Qty, Rate, Discount (if any)"],
        ["  → Go to 'Bepaari_Expenses' sheet"],
        ["  → Enter Bepaari expenses: Motor, Bhussa, Jagah Bhada, Gawali, Karkuni, Cash/Advance"],
        [""],
        ["STEP 3: Payments & Receipts"],
        ["  → 'Bepaari_Payments': When you pay a Bepaari"],
        ["  → 'Dukandar_Receipts': When a Dukandar pays you"],
        [""],
        ["STEP 4: View Reports"],
        ["  → 'Bepaari_Ledger': See each Bepaari's account summary"],
        ["  → 'Dukandar_Ledger': See each Dukandar's account summary"],
        ["  → 'Bepaari_Aakda': Generate settlement slip for any Bepaari"],
        ["  → 'Daily_Summary': Day-wise business summary"],
        [""],
        ["IMPORTANT NOTES:"],
        ["  • Commission Type: '%' means percentage (e.g., 4 = 4%)"],
        ["  • Commission Type: 'Per Piece' means per goat (e.g., 200 = ₹200/goat)"],
        ["  • Opening Balance: Positive = You owe them, Negative = They owe you"],
        ["  • All YELLOW cells are for data entry, other cells have formulas"],
        [""],
        ["Created: " + datetime.now().strftime("%d-%b-%Y")],
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_idx, column=1, value=row_data[0] if row_data else "")
        if row_idx == 1:
            cell.font = TITLE_FONT
        elif "STEP" in str(row_data[0]) or "IMPORTANT" in str(row_data[0]) or "HOW TO" in str(row_data[0]):
            cell.font = Font(bold=True, size=11)
    
    set_column_width(ws_instructions, 1, 80)
    
    # ========================================
    # SHEET 2: MASTERS
    # ========================================
    ws_masters = wb.create_sheet("Masters", 1)
    ws_masters.sheet_properties.tabColor = "2196F3"
    
    # Bepaari Master Section
    ws_masters.cell(row=1, column=1, value="BEPAARI MASTER").font = TITLE_FONT
    bepaari_headers = ["Sr.", "Bepaari Name", "Commission Type", "Commission Value", "Opening Balance", "Remarks"]
    for col, header in enumerate(bepaari_headers, 1):
        ws_masters.cell(row=2, column=col, value=header)
    style_header_row(ws_masters, 2, 1, len(bepaari_headers))
    
    # Add sample data and empty rows for Bepaaris
    sample_bepaaris = [
        [1, "ABUL HASAN", "%", 4, 0, "Sample - Delete or edit"],
        [2, "SHARAFAT", "%", 4, 0, ""],
        [3, "JUNAID", "Per Piece", 200, 0, "₹200 per goat"],
    ]
    for row_idx, data in enumerate(sample_bepaaris, 3):
        for col_idx, value in enumerate(data, 1):
            cell = ws_masters.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in [4, 5]:  # Commission Value and Opening Balance
                cell.number_format = MONEY_FORMAT if col_idx == 5 else NUMBER_FORMAT
    
    # Add empty rows for more bepaaris (up to 50)
    for row_idx in range(6, 52):
        ws_masters.cell(row=row_idx, column=1, value=row_idx - 2)
    
    # Commission Type validation
    comm_type_validation = DataValidation(type="list", formula1='"%, Per Piece"', allow_blank=True)
    comm_type_validation.error = "Please select % or Per Piece"
    ws_masters.add_data_validation(comm_type_validation)
    comm_type_validation.add(f'C3:C51')
    
    # Dukandar Master Section (Column H onwards)
    ws_masters.cell(row=1, column=8, value="DUKANDAR MASTER").font = TITLE_FONT
    dukandar_headers = ["Sr.", "Dukandar Name", "Opening Balance", "Remarks"]
    for col, header in enumerate(dukandar_headers, 8):
        ws_masters.cell(row=2, column=col, value=header)
    style_header_row(ws_masters, 2, 8, 11)
    
    # Sample Dukandars
    sample_dukandars = [
        [1, "IRFAN JOG", 0, "Sample"],
        [2, "SALIM", 0, ""],
        [3, "YUNUS", 0, ""],
    ]
    for row_idx, data in enumerate(sample_dukandars, 3):
        for col_idx, value in enumerate(data, 8):
            cell = ws_masters.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 10:  # Opening Balance
                cell.number_format = MONEY_FORMAT
    
    # Add empty rows for more dukandars
    for row_idx in range(6, 52):
        ws_masters.cell(row=row_idx, column=8, value=row_idx - 2)
    
    # Expense Types Section (Column M onwards)
    ws_masters.cell(row=1, column=13, value="EXPENSE TYPES").font = TITLE_FONT
    expense_headers = ["Sr.", "Expense Code", "Expense Name"]
    for col, header in enumerate(expense_headers, 13):
        ws_masters.cell(row=2, column=col, value=header)
    style_header_row(ws_masters, 2, 13, 15)
    
    expense_types = [
        [1, "MOTOR", "Motor Bhada (Transport)"],
        [2, "BHUSSA", "Bhussa (Fodder)"],
        [3, "JB", "Jagah Bhada (Space Rent)"],
        [4, "GAWALI", "Gawali (Handlers)"],
        [5, "KK", "Karkuni (Clerical)"],
        [6, "CASH", "Cash Given / Advance"],
        [7, "OTHER", "Other Expenses"],
    ]
    for row_idx, data in enumerate(expense_types, 3):
        for col_idx, value in enumerate(data, 13):
            ws_masters.cell(row=row_idx, column=col_idx, value=value)
    
    # Set column widths for Masters
    widths = [5, 25, 15, 15, 18, 20, 3, 5, 25, 18, 20, 3, 5, 12, 25]
    for col, width in enumerate(widths, 1):
        set_column_width(ws_masters, col, width)
    
    # ========================================
    # SHEET 3: DAILY_SALES
    # ========================================
    ws_sales = wb.create_sheet("Daily_Sales", 2)
    ws_sales.sheet_properties.tabColor = "4CAF50"
    
    ws_sales.cell(row=1, column=1, value="DAILY SALES ENTRY").font = TITLE_FONT
    sales_headers = [
        "Sr.", "Date", "Bepaari Name", "Dukandar Name", "Qty (Goats)", 
        "Rate (₹/Goat)", "Gross Total", "Dukandar Discount", "Net Amount", "Remarks"
    ]
    for col, header in enumerate(sales_headers, 1):
        ws_sales.cell(row=2, column=col, value=header)
    style_header_row(ws_sales, 2, 1, len(sales_headers))
    
    # Add formulas for calculated columns and format
    input_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")  # Light yellow
    for row in range(3, 503):  # 500 rows for entries
        ws_sales.cell(row=row, column=1, value=row-2)  # Sr. No.
        
        # Date column
        ws_sales.cell(row=row, column=2).number_format = DATE_FORMAT
        ws_sales.cell(row=row, column=2).fill = input_fill
        
        # Input columns highlighting
        for col in [3, 4, 5, 6, 8, 10]:
            ws_sales.cell(row=row, column=col).fill = input_fill
        
        # Gross Total = Qty × Rate
        ws_sales.cell(row=row, column=7, value=f"=IF(E{row}=\"\",\"\",E{row}*F{row})")
        ws_sales.cell(row=row, column=7).number_format = MONEY_FORMAT
        
        # Net Amount = Gross - Discount
        ws_sales.cell(row=row, column=9, value=f"=IF(G{row}=\"\",\"\",G{row}-IF(H{row}=\"\",0,H{row}))")
        ws_sales.cell(row=row, column=9).number_format = MONEY_FORMAT
        
        # Discount format
        ws_sales.cell(row=row, column=8).number_format = MONEY_FORMAT
    
    # Column widths
    sales_widths = [5, 12, 20, 20, 12, 15, 15, 18, 15, 25]
    for col, width in enumerate(sales_widths, 1):
        set_column_width(ws_sales, col, width)
    
    # ========================================
    # SHEET 4: BEPAARI_EXPENSES
    # ========================================
    ws_bep_exp = wb.create_sheet("Bepaari_Expenses", 3)
    ws_bep_exp.sheet_properties.tabColor = "F44336"
    
    ws_bep_exp.cell(row=1, column=1, value="BEPAARI EXPENSES ENTRY").font = TITLE_FONT
    exp_headers = ["Sr.", "Date", "Bepaari Name", "Expense Type", "Amount (₹)", "Remarks"]
    for col, header in enumerate(exp_headers, 1):
        ws_bep_exp.cell(row=2, column=col, value=header)
    style_header_row(ws_bep_exp, 2, 1, len(exp_headers))
    
    # Expense type validation
    exp_validation = DataValidation(type="list", formula1='"MOTOR,BHUSSA,JB,GAWALI,KK,CASH,OTHER"', allow_blank=True)
    ws_bep_exp.add_data_validation(exp_validation)
    
    for row in range(3, 503):
        ws_bep_exp.cell(row=row, column=1, value=row-2)
        ws_bep_exp.cell(row=row, column=2).number_format = DATE_FORMAT
        ws_bep_exp.cell(row=row, column=2).fill = input_fill
        ws_bep_exp.cell(row=row, column=3).fill = input_fill
        ws_bep_exp.cell(row=row, column=4).fill = input_fill
        ws_bep_exp.cell(row=row, column=5).number_format = MONEY_FORMAT
        ws_bep_exp.cell(row=row, column=5).fill = input_fill
        ws_bep_exp.cell(row=row, column=6).fill = input_fill
        exp_validation.add(f'D{row}')
    
    exp_widths = [5, 12, 25, 15, 15, 30]
    for col, width in enumerate(exp_widths, 1):
        set_column_width(ws_bep_exp, col, width)
    
    # ========================================
    # SHEET 5: BEPAARI_PAYMENTS
    # ========================================
    ws_bep_pay = wb.create_sheet("Bepaari_Payments", 4)
    ws_bep_pay.sheet_properties.tabColor = "9C27B0"
    
    ws_bep_pay.cell(row=1, column=1, value="BEPAARI PAYMENTS (Money Paid to Bepaari)").font = TITLE_FONT
    pay_headers = ["Sr.", "Date", "Bepaari Name", "Mode", "Amount (₹)", "Reference/Remarks"]
    for col, header in enumerate(pay_headers, 1):
        ws_bep_pay.cell(row=2, column=col, value=header)
    style_header_row(ws_bep_pay, 2, 1, len(pay_headers))
    
    mode_validation = DataValidation(type="list", formula1='"CASH,BANK,UPI,CHEQUE"', allow_blank=True)
    ws_bep_pay.add_data_validation(mode_validation)
    
    for row in range(3, 503):
        ws_bep_pay.cell(row=row, column=1, value=row-2)
        ws_bep_pay.cell(row=row, column=2).number_format = DATE_FORMAT
        ws_bep_pay.cell(row=row, column=2).fill = input_fill
        ws_bep_pay.cell(row=row, column=3).fill = input_fill
        ws_bep_pay.cell(row=row, column=4).fill = input_fill
        ws_bep_pay.cell(row=row, column=5).number_format = MONEY_FORMAT
        ws_bep_pay.cell(row=row, column=5).fill = input_fill
        ws_bep_pay.cell(row=row, column=6).fill = input_fill
        mode_validation.add(f'D{row}')
    
    pay_widths = [5, 12, 25, 12, 15, 35]
    for col, width in enumerate(pay_widths, 1):
        set_column_width(ws_bep_pay, col, width)
    
    # ========================================
    # SHEET 6: DUKANDAR_RECEIPTS
    # ========================================
    ws_duk_rec = wb.create_sheet("Dukandar_Receipts", 5)
    ws_duk_rec.sheet_properties.tabColor = "FF9800"
    
    ws_duk_rec.cell(row=1, column=1, value="DUKANDAR RECEIPTS (Money Received from Dukandar)").font = TITLE_FONT
    rec_headers = ["Sr.", "Date", "Dukandar Name", "Mode", "Amount (₹)", "Reference/Remarks"]
    for col, header in enumerate(rec_headers, 1):
        ws_duk_rec.cell(row=2, column=col, value=header)
    style_header_row(ws_duk_rec, 2, 1, len(rec_headers))
    
    ws_duk_rec.add_data_validation(mode_validation)
    
    for row in range(3, 503):
        ws_duk_rec.cell(row=row, column=1, value=row-2)
        ws_duk_rec.cell(row=row, column=2).number_format = DATE_FORMAT
        ws_duk_rec.cell(row=row, column=2).fill = input_fill
        ws_duk_rec.cell(row=row, column=3).fill = input_fill
        ws_duk_rec.cell(row=row, column=4).fill = input_fill
        ws_duk_rec.cell(row=row, column=5).number_format = MONEY_FORMAT
        ws_duk_rec.cell(row=row, column=5).fill = input_fill
        ws_duk_rec.cell(row=row, column=6).fill = input_fill
    
    rec_widths = [5, 12, 25, 12, 15, 35]
    for col, width in enumerate(rec_widths, 1):
        set_column_width(ws_duk_rec, col, width)
    
    # ========================================
    # SHEET 7: BEPAARI_LEDGER
    # ========================================
    ws_bep_ledger = wb.create_sheet("Bepaari_Ledger", 6)
    ws_bep_ledger.sheet_properties.tabColor = "00BCD4"
    
    ws_bep_ledger.cell(row=1, column=1, value="BEPAARI LEDGER SUMMARY").font = TITLE_FONT
    ws_bep_ledger.cell(row=2, column=1, value="(Auto-calculated from Daily_Sales, Bepaari_Expenses & Bepaari_Payments)").font = Font(italic=True, size=10)
    
    ledger_headers = [
        "Sr.", "Bepaari Name", "Comm. Type", "Comm. Value", "Opening Bal.",
        "Total Sales", "Total Qty", "Commission", "Total Expenses", 
        "Gross Payable", "Payments Made", "Closing Balance"
    ]
    for col, header in enumerate(ledger_headers, 1):
        ws_bep_ledger.cell(row=4, column=col, value=header)
    style_header_row(ws_bep_ledger, 4, 1, len(ledger_headers))
    
    # Add formulas for each bepaari (referencing Masters sheet)
    for row in range(5, 54):  # 49 bepaaris max
        master_row = row - 2  # Corresponding row in Masters
        
        ws_bep_ledger.cell(row=row, column=1, value=row-4)  # Sr.
        
        # Bepaari Name from Masters
        ws_bep_ledger.cell(row=row, column=2, value=f"=IF(Masters!B{master_row}=\"\",\"\",Masters!B{master_row})")
        
        # Commission Type from Masters
        ws_bep_ledger.cell(row=row, column=3, value=f"=IF(Masters!C{master_row}=\"\",\"\",Masters!C{master_row})")
        
        # Commission Value from Masters
        ws_bep_ledger.cell(row=row, column=4, value=f"=IF(Masters!D{master_row}=\"\",\"\",Masters!D{master_row})")
        
        # Opening Balance from Masters
        ws_bep_ledger.cell(row=row, column=5, value=f"=IF(Masters!E{master_row}=\"\",0,Masters!E{master_row})")
        ws_bep_ledger.cell(row=row, column=5).number_format = MONEY_FORMAT
        
        # Total Sales (SUMIF from Daily_Sales)
        ws_bep_ledger.cell(row=row, column=6, value=f"=SUMIF(Daily_Sales!C:C,B{row},Daily_Sales!G:G)")
        ws_bep_ledger.cell(row=row, column=6).number_format = MONEY_FORMAT
        
        # Total Qty (SUMIF from Daily_Sales)
        ws_bep_ledger.cell(row=row, column=7, value=f"=SUMIF(Daily_Sales!C:C,B{row},Daily_Sales!E:E)")
        ws_bep_ledger.cell(row=row, column=7).number_format = NUMBER_FORMAT
        
        # Commission calculation (% or Per Piece)
        ws_bep_ledger.cell(row=row, column=8, 
            value=f"=IF(B{row}=\"\",\"\",IF(C{row}=\"%\",F{row}*D{row}/100,G{row}*D{row}))")
        ws_bep_ledger.cell(row=row, column=8).number_format = MONEY_FORMAT
        
        # Total Expenses (SUMIF from Bepaari_Expenses)
        ws_bep_ledger.cell(row=row, column=9, value=f"=SUMIF(Bepaari_Expenses!C:C,B{row},Bepaari_Expenses!E:E)")
        ws_bep_ledger.cell(row=row, column=9).number_format = MONEY_FORMAT
        
        # Gross Payable = Sales - Commission - Expenses + Opening Balance
        ws_bep_ledger.cell(row=row, column=10, value=f"=IF(B{row}=\"\",\"\",F{row}-H{row}-I{row}+E{row})")
        ws_bep_ledger.cell(row=row, column=10).number_format = MONEY_FORMAT
        
        # Payments Made (SUMIF from Bepaari_Payments)
        ws_bep_ledger.cell(row=row, column=11, value=f"=SUMIF(Bepaari_Payments!C:C,B{row},Bepaari_Payments!E:E)")
        ws_bep_ledger.cell(row=row, column=11).number_format = MONEY_FORMAT
        
        # Closing Balance = Gross Payable - Payments Made
        ws_bep_ledger.cell(row=row, column=12, value=f"=IF(B{row}=\"\",\"\",J{row}-K{row})")
        ws_bep_ledger.cell(row=row, column=12).number_format = MONEY_FORMAT
    
    # Totals row
    total_row = 54
    ws_bep_ledger.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in [5, 6, 8, 9, 10, 11, 12]:
        ws_bep_ledger.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}53)")
        ws_bep_ledger.cell(row=total_row, column=col).number_format = MONEY_FORMAT
        ws_bep_ledger.cell(row=total_row, column=col).font = Font(bold=True)
    ws_bep_ledger.cell(row=total_row, column=7, value=f"=SUM(G5:G53)")
    ws_bep_ledger.cell(row=total_row, column=7).font = Font(bold=True)
    
    ledger_widths = [5, 22, 12, 12, 15, 15, 12, 15, 15, 15, 15, 15]
    for col, width in enumerate(ledger_widths, 1):
        set_column_width(ws_bep_ledger, col, width)
    
    # ========================================
    # SHEET 8: DUKANDAR_LEDGER
    # ========================================
    ws_duk_ledger = wb.create_sheet("Dukandar_Ledger", 7)
    ws_duk_ledger.sheet_properties.tabColor = "E91E63"
    
    ws_duk_ledger.cell(row=1, column=1, value="DUKANDAR LEDGER SUMMARY").font = TITLE_FONT
    ws_duk_ledger.cell(row=2, column=1, value="(Auto-calculated from Daily_Sales & Dukandar_Receipts)").font = Font(italic=True, size=10)
    
    duk_ledger_headers = [
        "Sr.", "Dukandar Name", "Opening Bal.", "Total Purchases", 
        "Total Discounts", "Net Receivable", "Receipts", "Closing Balance"
    ]
    for col, header in enumerate(duk_ledger_headers, 1):
        ws_duk_ledger.cell(row=4, column=col, value=header)
    style_header_row(ws_duk_ledger, 4, 1, len(duk_ledger_headers))
    
    for row in range(5, 54):
        master_row = row - 2
        
        ws_duk_ledger.cell(row=row, column=1, value=row-4)
        
        # Dukandar Name from Masters (Column I)
        ws_duk_ledger.cell(row=row, column=2, value=f"=IF(Masters!I{master_row}=\"\",\"\",Masters!I{master_row})")
        
        # Opening Balance from Masters (Column J)
        ws_duk_ledger.cell(row=row, column=3, value=f"=IF(Masters!J{master_row}=\"\",0,Masters!J{master_row})")
        ws_duk_ledger.cell(row=row, column=3).number_format = MONEY_FORMAT
        
        # Total Purchases (SUMIF Gross from Daily_Sales)
        ws_duk_ledger.cell(row=row, column=4, value=f"=SUMIF(Daily_Sales!D:D,B{row},Daily_Sales!G:G)")
        ws_duk_ledger.cell(row=row, column=4).number_format = MONEY_FORMAT
        
        # Total Discounts (SUMIF from Daily_Sales)
        ws_duk_ledger.cell(row=row, column=5, value=f"=SUMIF(Daily_Sales!D:D,B{row},Daily_Sales!H:H)")
        ws_duk_ledger.cell(row=row, column=5).number_format = MONEY_FORMAT
        
        # Net Receivable = Purchases - Discounts + Opening Balance
        ws_duk_ledger.cell(row=row, column=6, value=f"=IF(B{row}=\"\",\"\",D{row}-E{row}+C{row})")
        ws_duk_ledger.cell(row=row, column=6).number_format = MONEY_FORMAT
        
        # Receipts (SUMIF from Dukandar_Receipts)
        ws_duk_ledger.cell(row=row, column=7, value=f"=SUMIF(Dukandar_Receipts!C:C,B{row},Dukandar_Receipts!E:E)")
        ws_duk_ledger.cell(row=row, column=7).number_format = MONEY_FORMAT
        
        # Closing Balance = Net Receivable - Receipts
        ws_duk_ledger.cell(row=row, column=8, value=f"=IF(B{row}=\"\",\"\",F{row}-G{row})")
        ws_duk_ledger.cell(row=row, column=8).number_format = MONEY_FORMAT
    
    # Totals row
    total_row = 54
    ws_duk_ledger.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in [3, 4, 5, 6, 7, 8]:
        ws_duk_ledger.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}53)")
        ws_duk_ledger.cell(row=total_row, column=col).number_format = MONEY_FORMAT
        ws_duk_ledger.cell(row=total_row, column=col).font = Font(bold=True)
    
    duk_widths = [5, 22, 15, 18, 15, 18, 15, 18]
    for col, width in enumerate(duk_widths, 1):
        set_column_width(ws_duk_ledger, col, width)
    
    # ========================================
    # SHEET 9: BEPAARI_AAKDA (Settlement Slip)
    # ========================================
    ws_aakda = wb.create_sheet("Bepaari_Aakda", 8)
    ws_aakda.sheet_properties.tabColor = "795548"
    
    ws_aakda.cell(row=1, column=1, value="BEPAARI SETTLEMENT SLIP (AAKDA)").font = TITLE_FONT
    
    # Input section
    ws_aakda.cell(row=3, column=1, value="SELECT BEPAARI:")
    ws_aakda.cell(row=3, column=2).fill = input_fill
    ws_aakda.cell(row=3, column=2, value="")  # User will type bepaari name
    
    ws_aakda.cell(row=4, column=1, value="DATE:")
    ws_aakda.cell(row=4, column=2).fill = input_fill
    ws_aakda.cell(row=4, column=2).number_format = DATE_FORMAT
    
    # Settlement details header
    ws_aakda.cell(row=6, column=1, value="SETTLEMENT DETAILS").font = Font(bold=True, size=12)
    
    aakda_items = [
        ["A. SALES SUMMARY", ""],
        ["Total Qty (Goats)", f"=SUMIF(Daily_Sales!C:C,B3,Daily_Sales!E:E)"],
        ["Gross Sales Value", f"=SUMIF(Daily_Sales!C:C,B3,Daily_Sales!G:G)"],
        ["", ""],
        ["B. DEDUCTIONS", ""],
        ["Commission", f"=IFERROR(VLOOKUP(B3,Masters!B:E,4,FALSE),0)"],
        ["Commission Type", f"=IFERROR(VLOOKUP(B3,Masters!B:D,3,FALSE),\"\")"],
        ["Commission Amount", f"=IF(A13=\"%\",A9*A12/100,A8*A12)"],
        ["Motor Bhada", f"=SUMIFS(Bepaari_Expenses!E:E,Bepaari_Expenses!C:C,B3,Bepaari_Expenses!D:D,\"MOTOR\")"],
        ["Bhussa", f"=SUMIFS(Bepaari_Expenses!E:E,Bepaari_Expenses!C:C,B3,Bepaari_Expenses!D:D,\"BHUSSA\")"],
        ["Jagah Bhada", f"=SUMIFS(Bepaari_Expenses!E:E,Bepaari_Expenses!C:C,B3,Bepaari_Expenses!D:D,\"JB\")"],
        ["Gawali", f"=SUMIFS(Bepaari_Expenses!E:E,Bepaari_Expenses!C:C,B3,Bepaari_Expenses!D:D,\"GAWALI\")"],
        ["Karkuni", f"=SUMIFS(Bepaari_Expenses!E:E,Bepaari_Expenses!C:C,B3,Bepaari_Expenses!D:D,\"KK\")"],
        ["Cash/Advance", f"=SUMIFS(Bepaari_Expenses!E:E,Bepaari_Expenses!C:C,B3,Bepaari_Expenses!D:D,\"CASH\")"],
        ["Other Expenses", f"=SUMIFS(Bepaari_Expenses!E:E,Bepaari_Expenses!C:C,B3,Bepaari_Expenses!D:D,\"OTHER\")"],
        ["TOTAL DEDUCTIONS", f"=SUM(A14:A21)"],
        ["", ""],
        ["C. NET CALCULATION", ""],
        ["Gross Sales", f"=A9"],
        ["Less: Total Deductions", f"=A22"],
        ["Opening Balance", f"=IFERROR(VLOOKUP(B3,Masters!B:E,4,FALSE),0)"],
        ["NET PAYABLE TO BEPAARI", f"=A25-A26+A27"],
        ["", ""],
        ["D. PAYMENTS MADE", ""],
        ["Total Payments", f"=SUMIF(Bepaari_Payments!C:C,B3,Bepaari_Payments!E:E)"],
        ["", ""],
        ["BALANCE DUE", f"=A28-A31"],
    ]
    
    for row_idx, (label, formula) in enumerate(aakda_items, 7):
        ws_aakda.cell(row=row_idx, column=1, value=label)
        if formula:
            ws_aakda.cell(row=row_idx, column=2, value=formula)
            ws_aakda.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        
        # Bold section headers
        if label.startswith(("A.", "B.", "C.", "D.")) or "TOTAL" in label or "NET PAYABLE" in label or "BALANCE" in label:
            ws_aakda.cell(row=row_idx, column=1).font = Font(bold=True)
            if "TOTAL" in label or "NET PAYABLE" in label or "BALANCE" in label:
                ws_aakda.cell(row=row_idx, column=2).font = Font(bold=True)
    
    set_column_width(ws_aakda, 1, 25)
    set_column_width(ws_aakda, 2, 20)
    
    # ========================================
    # SHEET 10: DAILY_SUMMARY
    # ========================================
    ws_summary = wb.create_sheet("Daily_Summary", 9)
    ws_summary.sheet_properties.tabColor = "607D8B"
    
    ws_summary.cell(row=1, column=1, value="DAILY BUSINESS SUMMARY").font = TITLE_FONT
    ws_summary.cell(row=2, column=1, value="Enter date to see that day's summary").font = Font(italic=True, size=10)
    
    ws_summary.cell(row=4, column=1, value="SELECT DATE:")
    ws_summary.cell(row=4, column=2).fill = input_fill
    ws_summary.cell(row=4, column=2).number_format = DATE_FORMAT
    
    summary_items = [
        ["", ""],
        ["TODAY'S TRADING SUMMARY", ""],
        ["Total Goats Sold", f"=SUMIF(Daily_Sales!B:B,B4,Daily_Sales!E:E)"],
        ["Gross Sales Value", f"=SUMIF(Daily_Sales!B:B,B4,Daily_Sales!G:G)"],
        ["Total Discounts Given", f"=SUMIF(Daily_Sales!B:B,B4,Daily_Sales!H:H)"],
        ["Net Sales", f"=B9-B10"],
        ["", ""],
        ["EXPENSES (for the day)", ""],
        ["Total Bepaari Expenses", f"=SUMIF(Bepaari_Expenses!B:B,B4,Bepaari_Expenses!E:E)"],
        ["", ""],
        ["COMMISSION EARNED (Approx.)", ""],
        ["Estimated @ 4%", f"=B9*0.04"],
        ["", ""],
        ["PAYMENTS/RECEIPTS", ""],
        ["Paid to Bepaaris", f"=SUMIF(Bepaari_Payments!B:B,B4,Bepaari_Payments!E:E)"],
        ["Received from Dukandars", f"=SUMIF(Dukandar_Receipts!B:B,B4,Dukandar_Receipts!E:E)"],
        ["Net Cash Flow", f"=B21-B20"],
    ]
    
    for row_idx, (label, formula) in enumerate(summary_items, 5):
        ws_summary.cell(row=row_idx, column=1, value=label)
        if formula:
            ws_summary.cell(row=row_idx, column=2, value=formula)
            ws_summary.cell(row=row_idx, column=2).number_format = MONEY_FORMAT
        
        if "SUMMARY" in label or "EXPENSES" in label or "COMMISSION" in label or "PAYMENTS" in label:
            ws_summary.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    
    set_column_width(ws_summary, 1, 30)
    set_column_width(ws_summary, 2, 18)
    
    # ========================================
    # SAVE WORKBOOK
    # ========================================
    wb.save(output_path)
    print(f"✅ Mandi Master Excel created: {output_path}")
    return output_path


if __name__ == "__main__":
    output_file = "/app/Mandi_Master_Accounting.xlsx"
    create_mandi_excel(output_file)
